"""
工资汇总与明细合并工具（多Sheet版）
- 汇总.xls: 合同制工资条、代理制工资条
- 明细.xlsx: 传统、长价
遍历汇总所有sheet的人员，用工号去明细所有sheet查询，
每个人员生成一个独立Excel，包含汇总 + 传统明细 + 长价明细。
文件命名格式：序号_姓名_工号.xlsx
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import os
import sys
import re
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

# ======================== 路径处理（兼容exe运行） ========================
if getattr(sys, 'frozen', False):
    BASE_DIR = os.path.dirname(sys.executable)
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

OUTPUT_DIR = os.path.join(BASE_DIR, 'output')


def log(msg):
    """线程安全的日志输出"""
    text_log.insert(tk.END, msg + '\n')
    text_log.see(tk.END)
    root.update_idletasks()


def safe_filename(name):
    return re.sub(r'[\\/*?:"<>|]', '', name)


def safe_cell_value(val):
    """将长数字转为文本，避免Excel显示为科学计数法（如 8.026E+15）"""
    if pd.isna(val):
        return val
    if isinstance(val, float):
        s = str(val)
        if 'e' in s.lower() or ('e' not in s.lower() and len(s.replace('.', '')) >= 12):
            # 科学计数法 或 超过12位的数字，转为完整数字字符串
            try:
                return str(int(val))
            except (ValueError, OverflowError):
                return val
    if isinstance(val, int):
        s = str(val)
        if len(s) >= 12:
            return s
    return val


def browse_summary():
    path = filedialog.askopenfilename(
        title='请选择汇总Excel文件',
        filetypes=[('Excel文件', '*.xls *.xlsx'), ('所有文件', '*.*')],
        initialdir=BASE_DIR
    )
    if path:
        entry_summary.delete(0, tk.END)
        entry_summary.insert(0, path)


def browse_detail():
    path = filedialog.askopenfilename(
        title='请选择明细Excel文件',
        filetypes=[('Excel文件', '*.xls *.xlsx'), ('所有文件', '*.*')],
        initialdir=BASE_DIR
    )
    if path:
        entry_detail.delete(0, tk.END)
        entry_detail.insert(0, path)


def do_process():
    summary_file = entry_summary.get().strip()
    detail_file = entry_detail.get().strip()
    summary_key_col = entry_summary_key.get().strip()
    detail_key_col = entry_detail_key.get().strip()
    detail_filter_col = entry_detail_filter.get().strip()
    detail_filter_val = entry_detail_filter_val.get().strip()

    if not summary_file:
        messagebox.showwarning('提示', '请先选择汇总Excel文件！')
        return
    if not detail_file:
        messagebox.showwarning('提示', '请先选择明细Excel文件！')
        return
    if not summary_key_col:
        messagebox.showwarning('提示', '请输入汇总对比列名！')
        return
    if not detail_key_col:
        messagebox.showwarning('提示', '请输入明细对比列名！')
        return
    if not os.path.exists(summary_file):
        messagebox.showerror('错误', f'汇总文件不存在：\n{summary_file}')
        return
    if not os.path.exists(detail_file):
        messagebox.showerror('错误', f'明细文件不存在：\n{detail_file}')
        return

    btn_start.config(state=tk.DISABLED)
    text_log.delete(1.0, tk.END)

    def run():
        try:
            SUMMARY_FILE = summary_file
            DETAIL_FILE = detail_file
            SUMMARY_KEY = summary_key_col
            DETAIL_KEY = detail_key_col
            DETAIL_FILTER = detail_filter_col
            DETAIL_FILTER_VAL = detail_filter_val

            # ======================== 辅助函数：在DataFrame中查找列 ========================
            def find_col(df, keywords):
                """在DataFrame列中查找包含指定关键词的列名"""
                for col in df.columns:
                    col_str = str(col).strip()
                    for kw in keywords:
                        if kw in col_str:
                            return col
                return None

            def find_col_index(df_raw, keywords, search_rows=3):
                """在无表头DataFrame中查找包含关键词的列索引"""
                for r in range(min(search_rows, len(df_raw))):
                    for c in range(df_raw.shape[1]):
                        val = str(df_raw.iloc[r, c]).strip()
                        for kw in keywords:
                            if kw in val:
                                return c
                return None

            # ======================== 读取明细数据（所有sheet） ========================
            log('正在读取明细数据...')
            detail_xl = pd.ExcelFile(DETAIL_FILE)
            detail_sheet_names = detail_xl.sheet_names
            log(f'  明细共有 {len(detail_sheet_names)} 个sheet: {", ".join(detail_sheet_names)}')

            detail_sheets = []  # [{name, df, cols, biz_col}]
            for sname in detail_sheet_names:
                df = pd.read_excel(DETAIL_FILE, sheet_name=sname)
                # 修复Unnamed首列
                if str(df.columns[0]).startswith('Unnamed'):
                    df.rename(columns={df.columns[0]: '序号'}, inplace=True)
                cols = df.columns.tolist()
                # 查找用户指定的对比列
                biz_col = find_col(df, [DETAIL_KEY])
                if biz_col is None:
                    log(f'  警告: sheet "{sname}" 中未找到"{DETAIL_KEY}"列，跳过')
                    continue
                # 如果指定了剔除列，过滤掉等于指定值的行
                if DETAIL_FILTER and DETAIL_FILTER_VAL:
                    filter_col = find_col(df, [DETAIL_FILTER])
                    if filter_col is not None:
                        before = len(df)
                        # 尝试转为数值进行比较
                        try:
                            filter_num = float(DETAIL_FILTER_VAL)
                            df[filter_col] = pd.to_numeric(df[filter_col], errors='coerce')
                            df = df[~(df[filter_col].isna() | (df[filter_col] == filter_num))]
                        except ValueError:
                            # 非数值，按文本匹配
                            df[filter_col] = df[filter_col].astype(str).str.strip()
                            df = df[df[filter_col] != DETAIL_FILTER_VAL]
                        after = len(df)
                        log(f'  {sname}: 剔除"{DETAIL_FILTER}"={DETAIL_FILTER_VAL}的数据，{before}行 → {after}行')
                    else:
                        log(f'  警告: sheet "{sname}" 中未找到剔除列"{DETAIL_FILTER}"，跳过过滤')
                # 创建字符串版业务员代码用于匹配
                df['_biz_code_str'] = df[biz_col].apply(
                    lambda x: str(int(float(x))) if pd.notna(x) and not isinstance(x, str) else str(x)
                )
                detail_sheets.append({'name': sname, 'df': df, 'cols': cols, 'biz_col': biz_col})
                log(f'  {sname}: {len(df)} 行, {len(cols)} 列')
            if not detail_sheets:
                raise ValueError(f'明细文件中未找到任何包含"{DETAIL_KEY}"列的sheet')

            # ======================== 解析汇总数据（所有sheet） ========================
            def parse_summary_sheet(sheet_name):
                df_raw = pd.read_excel(SUMMARY_FILE, sheet_name=sheet_name, header=None)
                # 动态查找对比列和姓名列
                key_col = find_col_index(df_raw, [SUMMARY_KEY])
                name_col = find_col_index(df_raw, ['姓名'])
                if key_col is None or name_col is None:
                    log(f'  警告: sheet "{sheet_name}" 中未找到"{SUMMARY_KEY}"或姓名列，跳过')
                    return None
                main_headers = [str(df_raw.iloc[0, c]) if pd.notna(df_raw.iloc[0, c]) else '' for c in range(df_raw.shape[1])]
                sub_headers = [str(df_raw.iloc[1, c]) if pd.notna(df_raw.iloc[1, c]) else '' for c in range(df_raw.shape[1])]
                persons = []
                for idx in range(2, len(df_raw), 4):
                    if idx >= len(df_raw):
                        break
                    row = df_raw.iloc[idx]
                    key_val = row.iloc[key_col]
                    if pd.notna(key_val) and str(key_val) != SUMMARY_KEY:
                        data = [row.iloc[c] for c in range(len(row))]
                        kv = str(int(float(key_val))) if not isinstance(key_val, str) else str(key_val)
                        xm = str(row.iloc[name_col])
                        persons.append({
                            'gonghao': kv, 'name': xm, 'data': data,
                            'sheet': sheet_name, 'main_headers': main_headers, 'sub_headers': sub_headers
                        })
                return {'name': sheet_name, 'main_headers': main_headers, 'sub_headers': sub_headers, 'persons': persons}

            log('\n正在解析汇总数据...')
            sum_xl = pd.ExcelFile(SUMMARY_FILE)
            sum_sheet_names = sum_xl.sheet_names
            log(f'  汇总共有 {len(sum_sheet_names)} 个sheet: {", ".join(sum_sheet_names)}')
            all_persons = []
            sum_max_cols = 0
            for sname in sum_sheet_names:
                result = parse_summary_sheet(sname)
                if result is None:
                    continue
                all_persons.extend(result['persons'])
                sum_max_cols = max(sum_max_cols, len(result['main_headers']))
                log(f'  {sname}: {len(result["persons"])} 人')
            log(f'共计 {len(all_persons)} 人')
            if not all_persons:
                raise ValueError('汇总文件中未找到任何人员数据')

            # ======================== 匹配明细 ========================
            log('\n正在匹配明细...')
            for p in all_persons:
                p['details'] = []
                for ds in detail_sheets:
                    matched = ds['df'][ds['df']['_biz_code_str'] == p['gonghao']].drop(columns=['_biz_code_str'])
                    p['details'].append({
                        'label': ds['name'],
                        'df': matched,
                        'cols': ds['cols']
                    })
                parts = []
                for d in p['details']:
                    if len(d['df']) > 0:
                        parts.append(f'{d["label"]}{len(d["df"])}条')
                detail_info = ', '.join(parts) if parts else '无明细'
                log(f'  {SUMMARY_KEY} {p["gonghao"]} ({p["name"]}): {detail_info}')

            # ======================== 样式定义 ========================
            header_font = Font(name='微软雅黑', bold=True, size=10)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            sub_header_font = Font(name='微软雅黑', bold=True, size=9)
            sub_header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            # 明细各sheet颜色轮换
            detail_palette = [
                PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
                PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
                PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
                PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
                PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid'),
                PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
            ]
            detail_fills = {ds['name']: detail_palette[i % len(detail_palette)]
                            for i, ds in enumerate(detail_sheets)}
            data_font = Font(name='微软雅黑', size=9)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            def is_id_column(col_name):
                """判断列名是否为ID类列（需要文本格式避免科学计数法）"""
                col_str = str(col_name)
                for kw in ['代码', '工号', '序号', 'ID', '编号', '卡号', '账号',
                           '保单号', '单号', '合同号', '凭证号', '流水号',
                           '手机', '电话', '身份证']:
                    if kw in col_str:
                        return True
                # 任何以"号"、"码"结尾的列名都强制文本格式
                if col_str.endswith('号') or col_str.endswith('码'):
                    return True
                # 用户指定对比列也强制文本格式
                if SUMMARY_KEY in col_str or DETAIL_KEY in col_str:
                    return True
                return False

            def write_detail_section(ws, start_row, df, cols, label):
                if len(df) == 0:
                    return start_row
                r = start_row + 5
                fill = detail_fills.get(label, detail_palette[0])
                for col_idx, col_name in enumerate(cols):
                    cell = ws.cell(row=r, column=col_idx + 1)
                    cell.value = col_name
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = center_align
                    cell.border = thin_border
                r += 1
                for _, detail_row in df.iterrows():
                    for col_idx, col_name in enumerate(cols):
                        cell = ws.cell(row=r, column=col_idx + 1)
                        val = detail_row[col_name]
                        if pd.notna(val):
                            cell.value = safe_cell_value(val)
                        # ID类列强制设为文本格式，避免科学计数法
                        if is_id_column(col_name) and pd.notna(val):
                            cell.number_format = '@'
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = thin_border
                    r += 1
                return r

            # 计算最大列数
            detail_max_cols = max((len(ds['cols']) for ds in detail_sheets), default=0)
            max_cols = max(sum_max_cols, detail_max_cols)

            # ======================== 生成输出文件 ========================
            log('\n正在生成输出文件...')
            os.makedirs(OUTPUT_DIR, exist_ok=True)

            for pi, p in enumerate(all_persons):
                seq = pi + 1
                gh = p['gonghao']
                name = p['name']
                main_headers = p['main_headers']
                sub_headers = p['sub_headers']
                person_data = p['data']

                safe_name = safe_filename(name)
                out_filename = f'{seq}_{safe_name}_{gh}.xlsx'
                out_path = os.path.join(OUTPUT_DIR, out_filename)

                wb = openpyxl.Workbook()
                ws = wb.active
                ws.title = f'{safe_name}_{gh}'

                current_row = 1
                # 一级表头
                for col_idx in range(len(main_headers)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = main_headers[col_idx].replace('\n', '')
                    if val:
                        cell.value = val
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

                # 二级表头
                for col_idx in range(len(sub_headers)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = sub_headers[col_idx].replace('\n', '')
                    if val:
                        cell.value = val
                        cell.font = sub_header_font
                        cell.fill = sub_header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    else:
                        cell.border = thin_border
                current_row += 1

                # 汇总数据行
                for col_idx in range(len(person_data)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = person_data[col_idx]
                    if pd.notna(val):
                        cell.value = safe_cell_value(val)
                    # ID类列强制设为文本格式
                    header_name = main_headers[col_idx] if col_idx < len(main_headers) else ''
                    if is_id_column(header_name) and pd.notna(val):
                        cell.number_format = '@'
                    cell.font = data_font
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

                # 遍历所有明细sheet
                for d in p['details']:
                    current_row = write_detail_section(ws, current_row, d['df'], d['cols'], d['label'])

                # 调整列宽
                for col_idx in range(1, max_cols + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 14

                wb.save(out_path)
                log(f'  [{seq}/{len(all_persons)}] {out_filename}')

            log(f'\n完成！共生成 {len(all_persons)} 个文件')
            log(f'保存在 {OUTPUT_DIR} 目录下')
            messagebox.showinfo('完成', f'处理完成！\n共生成 {len(all_persons)} 个文件\n保存在 output 目录下')

        except Exception as e:
            log(f'\n错误: {e}')
            import traceback
            traceback.print_exc()
            messagebox.showerror('错误', f'处理失败：\n{e}')
        finally:
            btn_start.config(state=tk.NORMAL)

    threading.Thread(target=run, daemon=True).start()


# ======================== 构建GUI界面 ========================
root = tk.Tk()
root.title('工资汇总明细合并工具')
root.geometry('700x660')
root.resizable(True, True)

# 居中窗口
root.update_idletasks()
sw = root.winfo_screenwidth()
sh = root.winfo_screenheight()
x = (sw - 700) // 2
y = (sh - 660) // 2
root.geometry(f'700x660+{x}+{y}')

# 标题
lbl_title = tk.Label(root, text='工资汇总明细合并工具', font=('微软雅黑', 16, 'bold'))
lbl_title.pack(pady=(15, 10))

# 汇总文件选择行
frame1 = tk.Frame(root)
frame1.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame1, text='汇总文件：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_summary = tk.Entry(frame1, font=('微软雅黑', 10))
entry_summary.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
tk.Button(frame1, text='浏览...', font=('微软雅黑', 10), command=browse_summary, width=8).pack(side=tk.LEFT)

# 明细文件选择行
frame2 = tk.Frame(root)
frame2.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame2, text='明细文件：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_detail = tk.Entry(frame2, font=('微软雅黑', 10))
entry_detail.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
tk.Button(frame2, text='浏览...', font=('微软雅黑', 10), command=browse_detail, width=8).pack(side=tk.LEFT)

# 汇总对比列输入行
frame3 = tk.Frame(root)
frame3.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame3, text='汇总对比列：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_summary_key = tk.Entry(frame3, font=('微软雅黑', 10))
entry_summary_key.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
entry_summary_key.insert(0, '工号')
tk.Label(frame3, text='（汇总中用于匹配的列名）', font=('微软雅黑', 9), fg='gray').pack(side=tk.LEFT)

# 明细对比列输入行
frame4 = tk.Frame(root)
frame4.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame4, text='明细对比列：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_detail_key = tk.Entry(frame4, font=('微软雅黑', 10))
entry_detail_key.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
entry_detail_key.insert(0, '业务员代码')
tk.Label(frame4, text='（明细中用于匹配的列名）', font=('微软雅黑', 9), fg='gray').pack(side=tk.LEFT)

# 明细剔除列输入行
frame5 = tk.Frame(root)
frame5.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame5, text='明细剔除列：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_detail_filter = tk.Entry(frame5, font=('微软雅黑', 10))
entry_detail_filter.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
entry_detail_filter.insert(0, '实发佣金')
tk.Label(frame5, text='（指定要过滤的列名）', font=('微软雅黑', 9), fg='gray').pack(side=tk.LEFT)

# 明细剔除值输入行
frame6 = tk.Frame(root)
frame6.pack(fill=tk.X, padx=20, pady=5)
tk.Label(frame6, text='明细剔除值：', font=('微软雅黑', 11), width=10, anchor='e').pack(side=tk.LEFT)
entry_detail_filter_val = tk.Entry(frame6, font=('微软雅黑', 10))
entry_detail_filter_val.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 5))
entry_detail_filter_val.insert(0, '0')
tk.Label(frame6, text='（剔除等于此值的行，留空则不剔除）', font=('微软雅黑', 9), fg='gray').pack(side=tk.LEFT)

# 开始按钮
btn_start = tk.Button(root, text='开始处理', font=('微软雅黑', 12, 'bold'),
                       bg='#4CAF50', fg='white', activebackground='#45a049',
                       command=do_process, width=15, height=2)
btn_start.pack(pady=15)

# 日志输出区域
frame_log = tk.Frame(root)
frame_log.pack(fill=tk.BOTH, expand=True, padx=20, pady=(0, 15))
tk.Label(frame_log, text='处理日志：', font=('微软雅黑', 10), anchor='w').pack(anchor='w')
text_log = tk.Text(frame_log, font=('Consolas', 9), wrap=tk.WORD, bg='#1e1e1e', fg='#d4d4d4',
                    insertbackground='white', relief=tk.SUNKEN, borderwidth=1)
text_log.pack(fill=tk.BOTH, expand=True)
scrollbar = tk.Scrollbar(text_log)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
text_log.config(yscrollcommand=scrollbar.set)
scrollbar.config(command=text_log.yview)

root.mainloop()
