import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import pandas as pd
import os
import threading
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side



# ========== 业务逻辑 ==========

def find_consecutive_groups(nums_with_docs):
    """从已排序的(发票号码, 单据编号)列表中找出连号组"""
    if not nums_with_docs:
        return []

    groups = []
    i = 0
    while i < len(nums_with_docs):
        j = i
        while j + 1 < len(nums_with_docs) and nums_with_docs[j + 1][0] == nums_with_docs[j][0] + 1:
            j += 1
        if j > i:
            group = nums_with_docs[i:j + 1]
            groups.append({
                '前序发票号码': str(group[0][0]),
                '连号发票号码': str(group[-1][0]),
                '条数': len(group),
                '报销单号': ','.join(dict.fromkeys(str(p[1]) for p in group))
            })
        i = j + 1
    return groups


def process_excel(input_path, status_callback=None, log_callback=None):
    """处理Excel文件，返回结果DataFrame"""
    def log(msg):
        if log_callback:
            log_callback(msg)
        if status_callback:
            status_callback(msg)

    log("正在读取 Excel 文件...")
    df = pd.read_excel(input_path)

    if '发票号码' not in df.columns:
        raise ValueError("文件中未找到「发票号码」列")
    if '单据编号' not in df.columns:
        raise ValueError("文件中未找到「单据编号」列")

    log(f"已读取数据：{len(df)} 行，{len(df.columns)} 列")

    # 收集所有(发票号码, 单据编号)对
    pairs = []
    for _, row in df.iterrows():
        inv_str = str(row['发票号码'])
        if inv_str in ('nan', '', 'None'):
            continue
        doc_no = str(row['单据编号']) if pd.notna(row['单据编号']) else ''
        for num in inv_str.split(','):
            num = num.strip()
            if num:
                try:
                    pairs.append((int(num), doc_no))
                except ValueError:
                    continue

    log(f"拆分后共 {len(pairs)} 条发票记录，正在去重...")

    # 按发票号码去重（保留第一个出现的单据编号）
    seen = set()
    unique_pairs = []
    for num, doc in pairs:
        if num not in seen:
            seen.add(num)
            unique_pairs.append((num, doc))

    unique_pairs.sort(key=lambda x: x[0])
    log(f"去重后剩余 {len(unique_pairs)} 个唯一发票号码，正在检测连号...")

    groups = find_consecutive_groups(unique_pairs)
    log(f"检测完成，共发现 {len(groups)} 组连号发票")

    result_df = pd.DataFrame(groups, columns=['前序发票号码', '连号发票号码', '条数', '报销单号'])
    return result_df


# ========== GUI 页面 ==========

class SidebarButton(tk.Frame):
    """左侧导航按钮"""
    def __init__(self, parent, text, command, selected=False):
        super().__init__(parent, bg='#2c3e50', cursor='hand2')
        self.selected = selected
        self.command = command
        self.label = tk.Label(
            self,
            text=text,
            font=('Microsoft YaHei', 11),
            bg='#2c3e50',
            fg='white',
            padx=25,
            pady=12,
            anchor='w',
            width=14
        )
        self.label.pack(fill='x')
        self.bind('<Button-1>', self.on_click)
        self.label.bind('<Button-1>', self.on_click)
        self.bind('<Enter>', self.on_hover)
        self.bind('<Leave>', self.on_leave)
        self.label.bind('<Enter>', self.on_hover)
        self.label.bind('<Leave>', self.on_leave)
        self.set_selected(selected)

    def on_click(self, event=None):
        self.command()

    def on_hover(self, event=None):
        if not self.selected:
            self.label.config(bg='#34495e')
            self.config(bg='#34495e')

    def on_leave(self, event=None):
        if not self.selected:
            self.label.config(bg='#2c3e50')
            self.config(bg='#2c3e50')

    def set_selected(self, selected):
        self.selected = selected
        if selected:
            self.label.config(bg='#3498db', fg='white')
            self.config(bg='#3498db')
        else:
            self.label.config(bg='#2c3e50', fg='white')
            self.config(bg='#2c3e50')


class InvoicePage(tk.Frame):
    """发票连号检测页面"""
    def __init__(self, parent, root):
        super().__init__(parent, bg='#f5f5f5')
        self.root = root
        self.setup_ui()

    def setup_ui(self):
        # 主内容容器
        content = tk.Frame(self, bg='#f0f0f0')
        content.pack(fill='both', expand=True, padx=20, pady=20)

        # 黄色提示横幅
        banner = tk.Frame(content, bg='#fff3cd', padx=15, pady=10)
        banner.pack(fill='x', pady=(0, 20))
        tk.Label(
            banner,
            text='一键完成：读取报账单费用行 → 拆分逗号分隔发票号 → 检测连号 → 输出结果文件',
            font=('Microsoft YaHei', 10),
            bg='#fff3cd',
            fg='#856404'
        ).pack()

        # 输入文件区域
        input_frame = tk.LabelFrame(
            content,
            text=' 账单费用行文件 ',
            font=('Microsoft YaHei', 10),
            bg='#f0f0f0',
            fg='#000000',
            relief='groove',
            bd=1
        )
        input_frame.pack(fill='x', pady=(0, 15))

        file_row = tk.Frame(input_frame, bg='#f0f0f0')
        file_row.pack(fill='x', padx=12, pady=(2, 12))
        self.file_path_var = tk.StringVar()
        tk.Entry(file_row, textvariable=self.file_path_var, font=('Microsoft YaHei', 10), relief='solid', bd=1).pack(side='left', fill='x', expand=True, padx=(0, 10))
        tk.Button(
            file_row,
            text='浏览...',
            command=self.browse_file,
            font=('Microsoft YaHei', 10),
            bg='#e9ecef',
            fg='#333',
            relief='solid',
            bd=1,
            padx=15,
            pady=3
        ).pack(side='right')

        # 输出目录区域
        output_frame = tk.LabelFrame(
            content,
            text=' 输出目录 ',
            font=('Microsoft YaHei', 10),
            bg='#f0f0f0',
            fg='#000000',
            relief='groove',
            bd=1
        )
        output_frame.pack(fill='x', pady=(0, 20))

        out_row = tk.Frame(output_frame, bg='#f0f0f0')
        out_row.pack(fill='x', padx=12, pady=(2, 12))
        self.output_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(out_row, textvariable=self.output_dir_var, font=('Microsoft YaHei', 10), relief='solid', bd=1).pack(side='left', fill='x', expand=True, padx=(0, 10))
        tk.Button(
            out_row,
            text='浏览...',
            command=self.browse_output_dir,
            font=('Microsoft YaHei', 10),
            bg='#e9ecef',
            fg='#333',
            relief='solid',
            bd=1,
            padx=15,
            pady=3
        ).pack(side='right')

        # 转换按钮
        self.convert_btn = tk.Button(
            content,
            text='一键转换',
            command=self.start_conversion,
            font=('Microsoft YaHei', 12, 'bold'),
            bg='#9b59b6',
            fg='white',
            activebackground='#8e44ad',
            activeforeground='white',
            relief='flat',
            bd=0,
            padx=80,
            pady=12
        )
        self.convert_btn.pack(pady=(0, 18))

        # 进度条
        self.progress = ttk.Progressbar(content, mode='determinate', maximum=100, length=600)
        self.progress.pack(fill='x', pady=(0, 18))

        # 处理日志
        log_frame = tk.Frame(content, bg='white', relief='solid', bd=1)
        log_frame.pack(fill='both', expand=True, padx=2)
        tk.Label(log_frame, text='处理日志', font=('Microsoft YaHei', 11, 'bold'), bg='white', fg='#333').pack(anchor='w', padx=15, pady=(10, 5))
        self.log_text = scrolledtext.ScrolledText(
            log_frame,
            font=('Consolas', 10),
            bg='white',
            fg='#333',
            relief='flat',
            padx=10,
            pady=8,
            state='disabled'
        )
        self.log_text.pack(fill='both', expand=True, padx=10, pady=(0, 10))

        # 初始日志
        self.log('系统就绪，请选择报账单费用行文件后点击「一键转换」')

    def log(self, msg):
        timestamp = datetime.now().strftime('%H:%M:%S')
        line = f'[{timestamp}] {msg}\n'
        self.log_text.configure(state='normal')
        self.log_text.insert('end', line)
        self.log_text.see('end')
        self.log_text.configure(state='disabled')

    def browse_file(self):
        filepath = filedialog.askopenfilename(
            title='选择报账单费用行文件',
            filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')]
        )
        if filepath:
            self.file_path_var.set(filepath)
            self.output_dir_var.set(os.path.dirname(filepath))
            self.log(f'已选择文件：{filepath}')

    def browse_output_dir(self):
        dirpath = filedialog.askdirectory(title='选择输出目录')
        if dirpath:
            self.output_dir_var.set(dirpath)
            self.log(f'已选择输出目录：{dirpath}')

    def update_progress(self, value):
        self.root.after(0, lambda: self.progress.configure(value=value))

    def start_conversion(self):
        input_path = self.file_path_var.get().strip()
        output_dir = self.output_dir_var.get().strip()

        if not input_path:
            messagebox.showwarning('提示', '请先选择报账单费用行文件')
            return
        if not os.path.exists(input_path):
            messagebox.showerror('错误', f'文件不存在：\n{input_path}')
            return
        if not output_dir or not os.path.exists(output_dir):
            messagebox.showerror('错误', f'输出目录不存在：\n{output_dir}')
            return

        self.convert_btn.configure(state='disabled')
        self.progress.configure(value=0)
        self.log('开始处理...')

        def run():
            try:
                self.update_progress(10)
                result_df = process_excel(
                    input_path,
                    status_callback=lambda msg: self.update_progress(min(self.progress['value'] + 15, 90)),
                    log_callback=self.log
                )
                self.update_progress(90)

                if result_df.empty:
                    self.root.after(0, lambda: self.show_complete('未发现连号发票数据', result_df))
                    return

                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                output_path = os.path.join(output_dir, f'连号发票结果_{timestamp}.xlsx')
                result_df.to_excel(output_path, index=False, engine='openpyxl')

                # 设置表头样式：珊瑚橙色背景 + 白色字体
                wb = load_workbook(output_path)
                ws = wb.active
                header_fill = PatternFill(start_color='F4B183', end_color='F4B183', fill_type='solid')
                header_font = Font(name='Microsoft YaHei', size=11, bold=True, color='FFFFFF')
                header_alignment = Alignment(horizontal='center', vertical='center')
                thin_border = Border(
                    left=Side(style='thin', color='D9D9D9'),
                    right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'),
                    bottom=Side(style='thin', color='D9D9D9')
                )

                for cell in ws[1]:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = thin_border

                # 自动调整列宽并设置数据区域边框/对齐（报销单号左对齐，其他居中）
                expense_col = None
                for idx, cell in enumerate(ws[1]):
                    if cell.value == '报销单号':
                        expense_col = idx
                        break

                for col_idx, col in enumerate(ws.columns):
                    max_length = 0
                    col_letter = col[0].column_letter
                    for cell in col:
                        cell.border = thin_border
                        if cell.row != 1:
                            if col_idx == expense_col:
                                cell.alignment = Alignment(horizontal='left', vertical='center')
                            else:
                                cell.alignment = Alignment(horizontal='center', vertical='center')
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except Exception:
                            pass
                    ws.column_dimensions[col_letter].width = min(max_length + 4, 50)

                # 固定首行并设置表头行高
                ws.freeze_panes = 'A2'
                ws.row_dimensions[1].height = 24

                wb.save(output_path)
                wb.close()

                self.update_progress(100)


                self.root.after(0, lambda: self.show_complete(
                    f'转换完成！共发现 {len(result_df)} 组连号发票', result_df, output_path
                ))
            except Exception as e:
                self.root.after(0, lambda: self.show_error(str(e)))

        threading.Thread(target=run, daemon=True).start()

    def show_complete(self, msg, result_df, output_path=None):
        self.convert_btn.configure(state='normal')
        self.log(msg)
        if output_path:
            self.log(f'输出文件：{output_path}')

        detail = f'{msg}\n\n共 {len(result_df)} 条记录。'
        if output_path:
            detail += f'\n\n输出文件：\n{output_path}'
            result = messagebox.askyesno('转换完成', detail + '\n\n是否打开文件所在目录？')
            if result:
                os.startfile(os.path.dirname(output_path))
        else:
            messagebox.showinfo('转换完成', detail)

    def show_error(self, msg):
        self.progress.configure(value=0)
        self.convert_btn.configure(state='normal')
        self.log(f'处理失败：{msg}')
        messagebox.showerror('错误', f'处理失败：\n{msg}')


class HelpPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg='#f5f5f5')

        content = tk.Frame(self, bg='#f5f5f5')
        content.pack(fill='both', expand=True, padx=30, pady=25)

        tk.Label(
            content,
            text='使用说明',
            font=('Microsoft YaHei', 16, 'bold'),
            bg='#f5f5f5',
            fg='#2c3e50',
            anchor='center'
        ).pack(fill='x', pady=(0, 12))

        text = scrolledtext.ScrolledText(
            content,
            font=('Microsoft YaHei', 10),
            bg='white',
            fg='#333',
            relief='solid',
            bd=1,
            padx=20,
            pady=20,
            state='normal'
        )
        text.pack(fill='both', expand=True)
        text.insert('1.0', '''1. 准备数据
   请确保 Excel 文件包含以下列：
   • 发票号码：单个或多个发票号码，多个号码用逗号（,）分隔
   • 单据编号：每个费用行对应的报销单据编号

2. 选择文件
   点击「报账单费用行文件」右侧的「浏览...」按钮，选择需要处理的 Excel 文件。

3. 选择输出目录
   默认输出到输入文件所在目录，也可点击「浏览...」修改。

4. 开始转换
   点击紫色的「一键转换」按钮，程序会自动：
   • 拆分逗号分隔的发票号码
   • 对全部发票号码去重并排序
   • 检测相邻号码差值为 1 的连号组
   • 输出前序发票号码、连号发票号码、条数、报销单号

5. 查看结果
   处理完成后会在输出目录生成「连号发票结果.xlsx」。
''')
        text.configure(state='disabled')


class AboutPage(tk.Frame):
    def __init__(self, parent):
        super().__init__(parent, bg='#f5f5f5')

        content = tk.Frame(self, bg='#f5f5f5')
        content.pack(fill='both', expand=True, padx=60, pady=40)

        # 居中标题
        tk.Label(content, text='关于', font=('Microsoft YaHei', 20, 'bold'), bg='#f5f5f5', fg='#2c3e50').pack(pady=(20, 35))

        # 信息表单（左对齐）
        form = tk.Frame(content, bg='#f5f5f5')
        form.pack(anchor='w')

        info = [
            ('软件名称：', '发票连号检测工具'),
            ('版本号：', 'v2.0'),
            ('上线日期：', '2026年7月15日'),
            ('开发环境：', 'Python 3.x + Tkinter'),
            ('主要功能：', '自动检测报账单中连号的发票号码'),
            ('适用场景：', '企业财务发票连号核对与报销审计'),
        ]

        for label, value in info:
            row = tk.Frame(form, bg='#f5f5f5')
            row.pack(fill='x', pady=10)
            tk.Label(row, text=label, font=('Microsoft YaHei', 11, 'bold'), bg='#f5f5f5', fg='#333', anchor='w').pack(side='left')
            tk.Label(row, text=value, font=('Microsoft YaHei', 11), bg='#f5f5f5', fg='#333', anchor='w').pack(side='left', padx=(15, 0))

        # 版权信息
        footer = tk.Frame(self, bg='#f5f5f5')
        footer.pack(side='bottom', fill='x', pady=30)
        tk.Label(footer, text='© 2026 发票连号检测工具 保留所有权利', font=('Microsoft YaHei', 10), bg='#f5f5f5', fg='#999').pack()



class App:
    def __init__(self, root):
        self.root = root
        self.root.title('报销单发票连号检测工具')
        self.root.geometry('900x650')
        self.root.resizable(False, False)
        self.root.configure(bg='white')

        # 公共顶部标题栏（始终显示软件名称）
        header = tk.Frame(root, bg='#2c3e50', height=60)
        header.pack(fill='x', side='top')
        header.pack_propagate(False)
        tk.Label(
            header,
            text='发票连号检测工具',
            font=('Microsoft YaHei', 16, 'bold'),
            bg='#2c3e50',
            fg='white'
        ).pack(side='left', padx=25, pady=10)

        # 主体区域（左侧导航栏 + 主内容区）
        body_frame = tk.Frame(root, bg='#f5f5f5')
        body_frame.pack(side='top', fill='both', expand=True)

        # 左侧导航栏
        self.sidebar = tk.Frame(body_frame, bg='#2c3e50', width=180)
        self.sidebar.pack(side='left', fill='y')
        self.sidebar.pack_propagate(False)

        self.nav_items = []
        self.pages = {}

        self.nav_invoice = SidebarButton(self.sidebar, '发票连号检测', lambda: self.show_page('invoice'), selected=True)
        self.nav_invoice.pack(fill='x', pady=2)
        self.nav_items.append(self.nav_invoice)

        self.nav_help = SidebarButton(self.sidebar, '使用说明', lambda: self.show_page('help'))
        self.nav_help.pack(fill='x', pady=2)
        self.nav_items.append(self.nav_help)

        self.nav_about = SidebarButton(self.sidebar, '关于', lambda: self.show_page('about'))
        self.nav_about.pack(fill='x', pady=2)
        self.nav_items.append(self.nav_about)

        # 主内容区
        self.main_frame = tk.Frame(body_frame, bg='#f5f5f5')
        self.main_frame.pack(side='left', fill='both', expand=True)

        # 页面容器
        self.page_container = tk.Frame(self.main_frame, bg='#f5f5f5')
        self.page_container.pack(fill='both', expand=True)

        self.pages['invoice'] = InvoicePage(self.page_container, root)
        self.pages['help'] = HelpPage(self.page_container)
        self.pages['about'] = AboutPage(self.page_container)

        self.current_page = None
        self.show_page('invoice')

    def show_page(self, page_name):
        for name, page in self.pages.items():
            page.pack_forget()
        self.pages[page_name].pack(fill='both', expand=True)
        self.current_page = page_name

        for item in self.nav_items:
            item.set_selected(False)
        if page_name == 'invoice':
            self.nav_invoice.set_selected(True)
        elif page_name == 'help':
            self.nav_help.set_selected(True)
        elif page_name == 'about':
            self.nav_about.set_selected(True)


if __name__ == '__main__':
    import sys
    import traceback

    def show_startup_error(exc_info):
        """当 GUI 初始化失败时，用临时窗口显示错误"""
        try:
            root = tk.Tk()
            root.title('启动错误')
            root.geometry('700x300')
            root.configure(bg='#f5f5f5')
            msg = traceback.format_exception(*exc_info)
            text = tk.Text(root, font=('Consolas', 10), wrap='word')
            text.insert('1.0', ''.join(msg))
            text.pack(fill='both', expand=True, padx=10, pady=10)
            tk.Button(root, text='关闭', command=root.destroy,
                       font=('Microsoft YaHei', 11), padx=30).pack(pady=(0, 10))
            root.mainloop()
        except Exception:
            # 最后兜底：写入临时文件
            with open(os.path.join(os.path.expanduser('~'), 'Desktop', '发票检测工具_错误日志.txt'), 'w', encoding='utf-8') as f:
                f.write(''.join(traceback.format_exception(*exc_info)))

    try:
        root = tk.Tk()
        app = App(root)
        # 确保窗口显示在最前面
        root.lift()
        root.attributes('-topmost', True)
        root.after(100, lambda: root.attributes('-topmost', False))
        root.mainloop()
    except Exception:
        show_startup_error(sys.exc_info())
