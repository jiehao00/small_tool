# -*- coding: utf-8 -*-
"""
发票验重 & 连号检测 页面模块

功能：
  - 选择 Excel 文件，指定发票号码列
  - 重复检测：标记号码完全相同的行
  - 连号检测：标记号码连续的行组
  - 输出三个 Sheet：检测结果（原序）、重复明细、连号明细
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import re
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

# 样式常量
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")

# 高亮色
DUPLICATE_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 浅黄
CONSECUTIVE_FILL_A = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 浅橙
CONSECUTIVE_FILL_B = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")  # 浅蓝


def _extract_numeric(invoice_str: str):
    """提取发票号中的连续数字部分，用于连号比较。
    返回 (纯数字字符串, 整数值) 或 None。
    例如 '044001900111' → ('044001900111', 44001900111)
         'FP-2024-0001'  → ('20240001', 20240001)
    """
    if not invoice_str:
        return None
    s = str(invoice_str).strip()
    # 尝试找最长连续数字
    numbers = re.findall(r'\d+', s)
    if not numbers:
        return None
    # 取最长的数字段（最常见场景：号码就是主体数字）
    longest = max(numbers, key=len)
    try:
        return (longest, int(longest))
    except ValueError:
        return None


def _to_str(val):
    """安全转字符串"""
    if val is None:
        return ""
    return str(val).strip()


def create_page(parent, app):
    return _InvoiceCheckPage(parent, app).frame


class _InvoiceCheckPage:
    """发票验重 & 连号检测页面"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self._columns = []  # 文件列名
        self._filepath = ""

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # 页面标题
        tk.Label(
            self.frame, text="★ 发票验重 & 连号检测",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
                                     padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        frow = tk.Frame(file_frame, bg=bg)
        frow.pack(fill=tk.X)
        tk.Label(frow, text="待检测文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_file = ttk.Entry(frow, style="Readonly.TEntry", state="readonly")
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(frow, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_file).pack(side=tk.LEFT)

        # ---- 列映射 ----
        col_frame = ttk.LabelFrame(self.frame, text=" 列映射（重复/连号比对依据）", style="Card.TLabelframe",
                                    padding=(16, 10))
        col_frame.pack(fill=tk.X, pady=(0, 12))

        crow = tk.Frame(col_frame, bg=bg)
        crow.pack(fill=tk.X)
        tk.Label(crow, text="比对列", font=("Microsoft YaHei", 10),
                 bg=bg, width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_cols = ttk.Entry(crow, style="Normal.TEntry", width=36)
        self.entry_cols.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(crow, text="📋", width=3,
                   command=self._show_column_picker).pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(crow,
                 text="多选列组合匹配，仅选发票号时连号检测效果最佳",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6").pack(side=tk.LEFT)

        # ---- 检测选项 ----
        opt_frame = ttk.LabelFrame(self.frame, text=" 检测选项 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        opt_frame.pack(fill=tk.X, pady=(0, 12))

        # 重复检测
        self.dup_var = tk.BooleanVar(value=True)
        tk.Checkbutton(opt_frame, text="重复检测：标记所选列组合值完全相同的行（如 发票号+品名+金额 全同=重复）",
                       variable=self.dup_var,
                       bg=bg, fg="#2c3e50",
                       activebackground=bg, activeforeground="#2c3e50",
                       selectcolor="white", font=("Microsoft YaHei", 10),
                       anchor=tk.W).pack(anchor=tk.W)

        # 连号检测
        self.seq_var = tk.BooleanVar(value=False)
        seq_top = tk.Frame(opt_frame, bg=bg)
        seq_top.pack(fill=tk.X, anchor=tk.W, pady=(6, 0))
        tk.Checkbutton(seq_top, text="连号检测：标记发票号码连续的行组",
                       variable=self.seq_var,
                       bg=bg, fg="#2c3e50",
                       activebackground=bg, activeforeground="#2c3e50",
                       selectcolor="white", font=("Microsoft YaHei", 10),
                       anchor=tk.W).pack(side=tk.LEFT)
        tk.Label(seq_top, text="  连续阈值：",
                 font=("Microsoft YaHei", 10), bg=bg, fg="#2c3e50").pack(side=tk.LEFT, padx=(16, 4))
        self.spin_threshold = tk.Spinbox(seq_top, from_=2, to=999, width=4,
                                          font=("Microsoft YaHei", 10),
                                          justify="center")
        self.spin_threshold.delete(0, tk.END)
        self.spin_threshold.insert(0, "2")
        self.spin_threshold.pack(side=tk.LEFT)
        tk.Label(seq_top, text=" 张及以上标出",
                 font=("Microsoft YaHei", 10), bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        orow = tk.Frame(out_frame, bg=bg)
        orow.pack(fill=tk.X)
        tk.Label(orow, text="输出文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(orow, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(orow, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=14)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始检测",
                                     style="Action.TButton", width=22,
                                     command=self._start_check)
        self.btn_start.pack()

        # 快捷操作提示
        tk.Label(self.frame,
                 text="提示：检测结果 Sheet 保持原序；重复明细 / 连号明细 Sheet 按分组集中排序，便于审计查阅",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 anchor=tk.CENTER).pack(pady=(0, 8))

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
                                   font=("Microsoft YaHei", 11, "bold"),
                                   bg=bg, fg="#2c3e50",
                                   padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== 浏览 / 选择 ====================
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择待检测的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if path:
            self._set_readonly(self.entry_file, path)
            self._filepath = path
            self._load_columns(path)

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="发票检测结果.xlsx"
        )
        if path:
            self._set_readonly(self.entry_out, path)

    def _set_readonly(self, entry, value):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.configure(state="readonly")

    def _load_columns(self, filepath):
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            self._columns = [str(c) for c in row if c is not None]
            # 清空已选列
            self.entry_cols.delete(0, tk.END)
            self._log(f"已加载 {len(self._columns)} 个列名")
        except Exception as e:
            self._log(f"[ERROR] 读取列名失败: {e}")

    def _set_entry_text(self, entry, text):
        """只读 Entry 写入文本（需临时解只读）"""
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, text)
        entry.configure(state="readonly")

    # ==================== 弹窗选列 ====================
    def _show_column_picker(self):
        if not self._columns:
            messagebox.showwarning("提示", "请先选择待检测的 Excel 文件")
            return

        existing = [c.strip() for c in self.entry_cols.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择比对列")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()

        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.EXTENDED,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in self._columns:
            lb.insert(tk.END, col)

        # 恢复已选列
        sel_order = []
        for i, col in enumerate(self._columns):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            sel_order[:] = [i for i in sel_order if i not in removed]
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            self._set_entry_text(self.entry_cols, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10),
                  width=8, bg="#3498db", fg="white", relief="flat",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

    # ==================== 日志 ====================
    def _log(self, msg):
        ts = datetime.now().strftime("%H:%M:%S")
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, f"[{ts}] {msg}\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)
        self.app.log(msg)

    # ==================== 开始检测 ====================
    def _start_check(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return
        filepath = self.entry_file.get().strip()
        out_path = self.entry_out.get().strip()

        # 获取比对列（从 Entry 文本解析）
        col_text = self.entry_cols.get().strip()
        if not col_text:
            messagebox.showwarning("提示", "请点击 📋 按钮选择至少一列")
            return
        dup_cols = [c.strip() for c in col_text.split(",") if c.strip()]

        if not filepath:
            messagebox.showwarning("提示", "请选择待检测的 Excel 文件")
            return
        if not out_path:
            messagebox.showwarning("提示", "请设置输出路径")
            return
        if not self.dup_var.get() and not self.seq_var.get():
            messagebox.showwarning("提示", "请至少勾选一项检测")
            return

        try:
            threshold = int(self.spin_threshold.get())
            if threshold < 2:
                threshold = 2
        except ValueError:
            threshold = 2

        self.btn_start.configure(state=tk.DISABLED, text="检测中，请稍候...")
        threading.Thread(
            target=self._do_check,
            args=(filepath, dup_cols, out_path, self.dup_var.get(), self.seq_var.get(), threshold),
            daemon=True
        ).start()

    def _do_check(self, filepath, dup_cols, out_path, do_dup, do_seq, threshold):
        try:
            self._log("=" * 60)
            self._log(f"开始发票验重 & 连号检测")
            self._log(f"  文件: {os.path.basename(filepath)}")
            self._log(f"  重复比对列({len(dup_cols)}列): {', '.join(dup_cols)}")
            self._log(f"  重复检测: {'是' if do_dup else '否'}")
            self._log(f"  连号检测: {'是' if do_seq else '否'}, 阈值 ≥ {threshold} 张")

            # ---- 读取数据 ----
            self._log("读取文件...")
            wb = load_workbook(filepath)
            ws = wb.active
            all_rows = list(ws.iter_rows(min_row=1, values_only=True))
            wb.close()

            if not all_rows:
                self._log("[ERROR] 文件为空")
                self._reset_btn("⚡ 开始检测")
                return

            headers = [_to_str(c) for c in all_rows[0]]
            data_rows = all_rows[1:]

            # 验证列名
            dup_col_indices = []
            for col in dup_cols:
                if col not in headers:
                    self._log(f"[ERROR] 文件中不存在列名: '{col}'")
                    self._log(f"  可用列名: {', '.join(headers)}")
                    self._reset_btn("⚡ 开始检测")
                    return
                dup_col_indices.append(headers.index(col))

            self._log(f"  共 {len(headers)} 列，{len(data_rows)} 行数据")

            # ---- 构建组合键索引 ----
            # 重复检测: 组合键 = 所有选中列的值用 || 连接
            number_map = defaultdict(list)  # 组合键 → [行号列表]
            empty_indices = set()  # 记录组合键为空的行号
            for row_idx, row in enumerate(data_rows):
                key_parts = [_to_str(row[i]) for i in dup_col_indices]
                key = "||".join(key_parts)
                # 所有选中列全空才算空
                if all(p == "" for p in key_parts):
                    empty_indices.add(row_idx)
                else:
                    number_map[key].append(row_idx)

            # ---- 重复检测（基于组合键） ----
            dup_group_map = {}  # row_idx → (group_id, occurrence_count)
            dup_groups = defaultdict(list)  # 组合键 → [row_idx, ...]
            if do_dup:
                dup_id = 0
                for key, indices in number_map.items():
                    if len(indices) >= 2:
                        dup_id += 1
                        dup_groups[key] = indices
                        for ri in indices:
                            dup_group_map[ri] = (f"C{dup_id}", len(indices))
                self._log(f"  重复检测完成: 发现 {len(dup_groups)} 组重复（基于 {len(dup_cols)} 列组合）")
            else:
                self._log("  重复检测: 已跳过")

            # ---- 连号检测（基于组合键提取数值） ----
            seq_group_map = {}
            seq_groups = []
            if do_seq:
                row_numeric = []
                for row_idx, row in enumerate(data_rows):
                    if row_idx in empty_indices:
                        continue
                    # 同样用组合键作为基数，提取最长数值段
                    key_parts = [_to_str(row[i]) for i in dup_col_indices]
                    composite = "||".join(key_parts)
                    n = _extract_numeric(composite)
                    if n:
                        row_numeric.append((row_idx, n[1], n[0]))

                sorted_rows = sorted(
                    [(ri, n_int, n_str) for ri, n_int, n_str in row_numeric],
                    key=lambda x: x[1]
                )

                if sorted_rows:
                    groups_raw = []
                    current_group = [sorted_rows[0]]
                    for i in range(1, len(sorted_rows)):
                        prev = sorted_rows[i - 1]
                        curr = sorted_rows[i]
                        if curr[1] == prev[1] + 1:
                            current_group.append(curr)
                        else:
                            groups_raw.append(current_group)
                            current_group = [curr]
                    groups_raw.append(current_group)

                    seq_id = 0
                    for grp in groups_raw:
                        if len(grp) >= threshold:
                            seq_id += 1
                            gid = f"L{seq_id}"
                            group_info = []
                            for ri, n_int, n_str in grp:
                                seq_group_map[ri] = (gid, len(grp))
                                group_info.append((ri, n_int))
                            seq_groups.append((gid, len(grp), group_info))

                self._log(f"  连号检测完成: 发现 {len(seq_groups)} 组连号（≥{threshold} 张）")
            else:
                self._log("  连号检测: 已跳过")

            # ---- 写入输出 ----
            self._log("正在生成输出文件...")
            out_wb = Workbook()

            # -- Sheet1: 检测结果（原序） --
            ws_result = out_wb.active
            ws_result.title = "检测结果"

            # 追加标记列
            mark_headers = []
            if do_dup:
                mark_headers.append("_验重结果")
                mark_headers.append("_验重组号")
            if do_seq:
                mark_headers.append("_连号结果")
                mark_headers.append("_连号组号")
                mark_headers.append("_连号张数")

            output_headers = headers + mark_headers

            # 写表头
            for j, h in enumerate(output_headers, 1):
                cell = ws_result.cell(row=1, column=j, value=h)
                self._style_header_cell(cell)

            # 写数据（原序）
            for row_idx, row in enumerate(data_rows):
                out_row_num = row_idx + 2
                for j, val in enumerate(row):
                    cell = ws_result.cell(row=out_row_num, column=j + 1, value=val)
                    cell.alignment = CELL_ALIGNMENT

                # 填充标记列
                mark_col_offset = len(headers) + 1
                if do_dup:
                    if row_idx in dup_group_map:
                        gid, cnt = dup_group_map[row_idx]
                        ws_result.cell(row=out_row_num, column=mark_col_offset,
                                       value=f"⚠ 重复({gid})").alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 1,
                                       value=gid).alignment = CELL_ALIGNMENT
                        # 高亮该行
                        for j in range(len(headers)):
                            ws_result.cell(row=out_row_num, column=j + 1).fill = DUPLICATE_FILL
                    else:
                        ws_result.cell(row=out_row_num, column=mark_col_offset,
                                       value="正常").alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 1,
                                       value="—").alignment = CELL_ALIGNMENT
                    mark_col_offset += 2

                if do_seq:
                    if row_idx in seq_group_map:
                        gid, gsize = seq_group_map[row_idx]
                        ws_result.cell(row=out_row_num, column=mark_col_offset,
                                       value=f"⚠ 连号({gid})").alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 1,
                                       value=gid).alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 2,
                                       value=gsize).alignment = CELL_ALIGNMENT
                    else:
                        ws_result.cell(row=out_row_num, column=mark_col_offset,
                                       value="正常").alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 1,
                                       value="—").alignment = CELL_ALIGNMENT
                        ws_result.cell(row=out_row_num, column=mark_col_offset + 2,
                                       value="—").alignment = CELL_ALIGNMENT

            # -- Sheet2: 重复明细（按组集中） --
            if do_dup and dup_groups:
                ws_dup = out_wb.create_sheet("重复明细")
                dup_headers = headers + ["_验重组号", "_重复次数"]
                for j, h in enumerate(dup_headers, 1):
                    cell = ws_dup.cell(row=1, column=j, value=h)
                    self._style_header_cell(cell)

                out_row = 2
                for inv_val, indices in dup_groups.items():
                    gid = dup_group_map[indices[0]][0]
                    cnt = len(indices)
                    for ri in sorted(indices):
                        row_data = data_rows[ri]
                        for j, val in enumerate(row_data):
                            ws_dup.cell(row=out_row, column=j + 1, value=val).alignment = CELL_ALIGNMENT
                        ws_dup.cell(row=out_row, column=len(headers) + 1, value=gid).alignment = CELL_ALIGNMENT
                        ws_dup.cell(row=out_row, column=len(headers) + 2, value=cnt).alignment = CELL_ALIGNMENT
                        # 高亮
                        for j in range(len(headers)):
                            ws_dup.cell(row=out_row, column=j + 1).fill = DUPLICATE_FILL
                        out_row += 1
                self._log(f"  重复明细: {out_row - 2} 行")
            elif do_dup:
                ws_dup = out_wb.create_sheet("重复明细")
                ws_dup.cell(row=1, column=1, value="未发现重复发票").alignment = CELL_ALIGNMENT

            # -- Sheet3: 连号明细（按组集中，组内排序） --
            if do_seq and seq_groups:
                ws_seq = out_wb.create_sheet("连号明细")
                seq_headers = headers + ["_连号组号", "_连号张数", "_发票号码(数值)"]
                for j, h in enumerate(seq_headers, 1):
                    cell = ws_seq.cell(row=1, column=j, value=h)
                    self._style_header_cell(cell)

                out_row = 2
                color_toggle = True
                for gid, gsize, group_info in seq_groups:
                    fill = CONSECUTIVE_FILL_A if color_toggle else CONSECUTIVE_FILL_B
                    color_toggle = not color_toggle

                    # 组内按发票号码排序
                    for ri, n_int in sorted(group_info, key=lambda x: x[1]):
                        row_data = data_rows[ri]
                        for j, val in enumerate(row_data):
                            ws_seq.cell(row=out_row, column=j + 1, value=val).alignment = CELL_ALIGNMENT
                        ws_seq.cell(row=out_row, column=len(headers) + 1, value=gid).alignment = CELL_ALIGNMENT
                        ws_seq.cell(row=out_row, column=len(headers) + 2, value=gsize).alignment = CELL_ALIGNMENT
                        ws_seq.cell(row=out_row, column=len(headers) + 3,
                                    value=str(n_int)).alignment = CELL_ALIGNMENT
                        ws_seq.cell(row=out_row, column=len(headers) + 3).number_format = '@'
                        # 高亮
                        for j in range(len(headers)):
                            ws_seq.cell(row=out_row, column=j + 1).fill = fill
                        out_row += 1
                self._log(f"  连号明细: {out_row - 2} 行")
            elif do_seq:
                ws_seq = out_wb.create_sheet("连号明细")
                ws_seq.cell(row=1, column=1, value="未发现连号发票").alignment = CELL_ALIGNMENT

            # 保存
            out_wb.save(out_path)
            out_wb.close()

            self._log("=" * 60)
            dup_count = len(dup_groups) if do_dup else 0
            seq_count = len(seq_groups) if do_seq else 0
            self._log(f"检测完成！重复 {dup_count} 组，连号 {seq_count} 组")
            self._log(f"结果已保存: {os.path.basename(out_path)}")
            self._log("=" * 60)

            self._reset_btn("⚡ 开始检测")

            # 结果弹窗
            msg = f"检测完成！\n\n"
            if do_dup:
                msg += f"重复检测：发现 {dup_count} 组重复\n"
            if do_seq:
                msg += f"连号检测：发现 {seq_count} 组连号（≥{threshold} 张）\n"
            msg += f"\n结果已保存至:\n{out_path}"
            self.frame.after(0, lambda m=msg: messagebox.showinfo("检测完成", m))

        except Exception as e:
            self._log(f"[ERROR] {e}")
            import traceback
            self._log(traceback.format_exc())
            self._reset_btn("⚡ 开始检测")
            err_msg = str(e)
            self.frame.after(0, lambda m=err_msg: messagebox.showerror("检测失败", f"发生错误:\n{m}"))

    def _style_header_cell(self, cell):
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT
        cell.border = HEADER_BORDER

    def _reset_btn(self, text):
        self.frame.after(0, lambda: self.btn_start.configure(state=tk.NORMAL, text=text))
