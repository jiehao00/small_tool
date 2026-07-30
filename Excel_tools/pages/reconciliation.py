# -*- coding: utf-8 -*-
"""
财务对账页面模块

功能：汇总表（A）vs 明细表（B）→ 确认汇总金额是否与明细加总一致
  - 选择主键列（匹配用）和对比列（数值列）
  - 明细表按主键分组，对比列求和
  - 与汇总表逐行比对，不一致的单元格标黄
  - 输出带差异标注的汇总表
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

# 表头样式
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))

# 差异高亮
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def create_page(parent, app):
    return ReconciliationPage(parent, app).frame


class ReconciliationPage:
    """财务对账页面"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self._a_path = None
        self._b_path = None
        self._a_headers = []
        self._b_headers = []

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== 构建 UI ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        tk.Label(self.frame, text="财务对账（汇总表 vs 明细累加对比）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
            padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        self._file_row(file_frame, "汇总表（A 文件）", "entry_file_a", self._browse_file_a)
        self._file_row(file_frame, "明细表（B 文件）", "entry_file_b", self._browse_file_b)

        # ---- 主键列配置 ----
        key_frame = ttk.LabelFrame(self.frame, text=" 主键列配置（用于匹配汇总表与明细表）",
            style="Card.TLabelframe", padding=(16, 10))
        key_frame.pack(fill=tk.X, pady=(0, 12))

        row_key_a = tk.Frame(key_frame, bg=bg)
        row_key_a.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row_key_a, text="汇总主键列", font=("Microsoft YaHei", 10),
            bg=bg, width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_key_a = ttk.Entry(row_key_a, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_key_a.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_key_a, text="🏷️ 选列", style="Secondary.TButton",
            command=lambda: self._show_column_picker(self.entry_key_a, self._a_headers)
        ).pack(side=tk.LEFT, padx=(0, 12))
        self.lbl_a_info = tk.Label(row_key_a, text="请先选择文件", font=("Microsoft YaHei", 8),
            bg=bg, fg="#95a5a6")
        self.lbl_a_info.pack(side=tk.LEFT)

        row_key_b = tk.Frame(key_frame, bg=bg)
        row_key_b.pack(fill=tk.X)
        tk.Label(row_key_b, text="明细主键列", font=("Microsoft YaHei", 10),
            bg=bg, width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_key_b = ttk.Entry(row_key_b, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_key_b.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_key_b, text="🏷️ 选列", style="Secondary.TButton",
            command=lambda: self._show_column_picker(self.entry_key_b, self._b_headers)
        ).pack(side=tk.LEFT, padx=(0, 12))
        self.lbl_b_info = tk.Label(row_key_b, text="请先选择文件", font=("Microsoft YaHei", 8),
            bg=bg, fg="#95a5a6")
        self.lbl_b_info.pack(side=tk.LEFT)

        # ---- 对比列配置 ----
        cmp_frame = ttk.LabelFrame(self.frame, text=" 对比列配置（要核对的数值列，两侧一一对应）",
            style="Card.TLabelframe", padding=(16, 10))
        cmp_frame.pack(fill=tk.X, pady=(0, 12))

        row_cmp_a = tk.Frame(cmp_frame, bg=bg)
        row_cmp_a.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row_cmp_a, text="汇总对比列", font=("Microsoft YaHei", 10),
            bg=bg, width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_cmp_a = ttk.Entry(row_cmp_a, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_cmp_a.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_cmp_a, text="🏷️ 选列", style="Secondary.TButton",
            command=lambda: self._show_column_picker(self.entry_cmp_a, self._a_headers)
        ).pack(side=tk.LEFT)

        row_cmp_b = tk.Frame(cmp_frame, bg=bg)
        row_cmp_b.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row_cmp_b, text="明细对比列", font=("Microsoft YaHei", 10),
            bg=bg, width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_cmp_b = ttk.Entry(row_cmp_b, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_cmp_b.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_cmp_b, text="🏷️ 选列", style="Secondary.TButton",
            command=lambda: self._show_column_picker(self.entry_cmp_b, self._b_headers)
        ).pack(side=tk.LEFT)

        tk.Label(cmp_frame,
            text="提示：汇总对比列的第1列与明细对比列的第1列对比，第2列与第2列对比…两侧列数须一致。"
                 "明细表按主键分组后对比列自动求和。",
            font=("Microsoft YaHei", 9), bg=bg, fg="#95a5a6",
            wraplength=600, justify=tk.LEFT
        ).pack(anchor=tk.W, fill=tk.X, pady=(6, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
            padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=16)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始对账",
            style="Action.TButton", width=22,
            command=self._start)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
            font=("Microsoft YaHei", 11, "bold"), bg=bg, fg="#2c3e50",
            padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== UI 辅助 ====================
    def _file_row(self, parent, label, attr, cmd):
        row = tk.Frame(parent, bg=self.app.COLOR_CONTENT_BG)
        row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row, text=label, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=22, anchor=tk.E
        ).pack(side=tk.LEFT, padx=(0, 8))
        entry = ttk.Entry(row, style="Readonly.TEntry", state="readonly")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="📂 浏览", style="Secondary.TButton",
            command=cmd).pack(side=tk.LEFT)
        setattr(self, attr, entry)

    def _set_entry_text(self, entry, text):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, text)
        entry.configure(state="readonly")

    # ==================== 弹窗选列 ====================
    def _show_column_picker(self, entry, cols):
        if not cols:
            messagebox.showwarning("提示", "请先选择对应的文件以加载列名")
            return

        existing = [c.strip() for c in entry.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择列名")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()

        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
            font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
        ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
            selectmode=tk.EXTENDED,
            bg="white", fg="#2c3e50", relief="flat",
            selectbackground="#3498db", selectforeground="white",
            exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in cols:
            lb.insert(tk.END, col)

        sel_order = []
        for i, col in enumerate(cols):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            sel_order[:] = [i for i in sel_order if i not in removed]
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            entry.delete(0, tk.END)
            entry.insert(0, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
            width=8, command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10),
            width=8, bg="#3498db", fg="white", relief="flat",
            activebackground="#2980b9", activeforeground="white",
            command=on_ok).pack(side=tk.RIGHT)

        popup.wait_window()

    # ==================== 浏览 & 加载列名 ====================
    def _browse_file_a(self):
        path = filedialog.askopenfilename(
            title="选择汇总表（A 文件）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._a_path = path
            self._set_entry_text(self.entry_file_a, path)
            self._log(f"已选择汇总表: {os.path.basename(path)}")
            self._load_columns(path, "A")

    def _browse_file_b(self):
        path = filedialog.askopenfilename(
            title="选择明细表（B 文件）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._b_path = path
            self._set_entry_text(self.entry_file_b, path)
            self._log(f"已选择明细表: {os.path.basename(path)}")
            self._load_columns(path, "B")

    def _load_columns(self, path, tag):
        try:
            wb = load_workbook(path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                headers = [str(c) for c in row if c is not None]
                if tag == "A":
                    self._a_headers = headers
                    self.lbl_a_info.config(text=f"共 {len(headers)} 列", fg="#27ae60")
                    self.entry_key_a.delete(0, tk.END)
                    self.entry_cmp_a.delete(0, tk.END)
                else:
                    self._b_headers = headers
                    self.lbl_b_info.config(text=f"共 {len(headers)} 列", fg="#27ae60")
                    self.entry_key_b.delete(0, tk.END)
                    self.entry_cmp_b.delete(0, tk.END)
                self._log(f"{tag} 文件已加载 {len(headers)} 列")
        except Exception as e:
            self._log(f"[ERROR] 加载 {tag} 文件列名失败: {e}")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="对账结果.xlsx"
        )
        if path:
            self._set_entry_text(self.entry_out, path)
            self._log(f"已设置输出路径: {path}")

    # ==================== 日志 ====================
    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.app.log(message)

    # ==================== 启动 ====================
    def _start(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        file_a = self._a_path or ""
        file_b = self._b_path or ""
        out_path = self.entry_out.get().strip()
        key_a = self.entry_key_a.get().strip()
        key_b = self.entry_key_b.get().strip()
        cmp_a = self.entry_cmp_a.get().strip()
        cmp_b = self.entry_cmp_b.get().strip()

        if not all([file_a, file_b, out_path, key_a, key_b, cmp_a, cmp_b]):
            messagebox.showwarning("参数缺失", "请填写所有必填项！")
            return
        if not os.path.isfile(file_a):
            messagebox.showerror("文件错误", f"汇总表不存在:\n{file_a}"); return
        if not os.path.isfile(file_b):
            messagebox.showerror("文件错误", f"明细表不存在:\n{file_b}"); return

        key_a_cols = [c.strip() for c in key_a.split(",") if c.strip()]
        key_b_cols = [c.strip() for c in key_b.split(",") if c.strip()]
        cmp_a_cols = [c.strip() for c in cmp_a.split(",") if c.strip()]
        cmp_b_cols = [c.strip() for c in cmp_b.split(",") if c.strip()]

        if len(key_a_cols) != len(key_b_cols):
            messagebox.showwarning("列数不一致",
                f"主键列数量不一致：汇总 {len(key_a_cols)} vs 明细 {len(key_b_cols)}"); return
        if len(cmp_a_cols) != len(cmp_b_cols):
            messagebox.showwarning("列数不一致",
                f"对比列数量不一致：汇总 {len(cmp_a_cols)} vs 明细 {len(cmp_b_cols)}"); return

        self.btn_start.configure(state="disabled", text="处理中...")
        t = threading.Thread(
            target=self._do_reconciliation,
            args=(file_a, file_b, key_a_cols, key_b_cols, cmp_a_cols, cmp_b_cols, out_path),
            daemon=True)
        t.start()

    # ==================== 核心逻辑 ====================
    def _do_reconciliation(self, file_a, file_b, key_a_cols, key_b_cols,
                            cmp_a_cols, cmp_b_cols, out_path):
        try:
            self._log("=" * 50)
            self._log("开始财务对账")

            self._log(f"主键列: 汇总[{key_a_cols}] <-> 明细[{key_b_cols}]")
            self._log(f"对比列: 汇总[{cmp_a_cols}] <-> 明细[{cmp_b_cols}]")

            # 加载工作簿
            self._log("加载汇总表...")
            wb_a = load_workbook(file_a)
            ws_a = wb_a.active
            self._log(f"汇总表: {ws_a.title}，{ws_a.max_row} 行 × {ws_a.max_column} 列")

            self._log("加载明细表...")
            wb_b = load_workbook(file_b)
            ws_b = wb_b.active
            self._log(f"明细表: {ws_b.title}，{ws_b.max_row} 行 × {ws_b.max_column} 列")

            # 表头映射
            headers_a = {cell.value: idx for idx, cell in enumerate(ws_a[1], start=1) if cell.value}
            headers_b = {cell.value: idx for idx, cell in enumerate(ws_b[1], start=1) if cell.value}

            for c in key_a_cols + cmp_a_cols:
                if c not in headers_a:
                    raise ValueError(f"汇总表中不存在列名: '{c}'")
            for c in key_b_cols + cmp_b_cols:
                if c not in headers_b:
                    raise ValueError(f"明细表中不存在列名: '{c}'")

            # ----- 步骤1：明细表按主键分组，对比列求和 -----
            self._log("明细表按主键分组 & 对比列求和...")
            detail_sums = {}
            for row_b in ws_b.iter_rows(min_row=2, values_only=False):
                key = tuple(
                    str(row_b[headers_b[c] - 1].value or "").strip()
                    for c in key_b_cols
                )
                if key not in detail_sums:
                    detail_sums[key] = [0.0] * len(cmp_b_cols)
                for i, c in enumerate(cmp_b_cols):
                    val = self._to_float(row_b[headers_b[c] - 1].value)
                    detail_sums[key][i] += val
            self._log(f"明细分组完成，共 {len(detail_sums)} 个唯一主键")

            # ----- 步骤2：遍历汇总表，逐行比对 -----
            self._log("逐行比对汇总表与明细汇总...")
            match_count = 0
            diff_count = 0
            no_detail_count = 0

            for row_idx, row_a in enumerate(ws_a.iter_rows(min_row=2, values_only=False), start=2):
                key = tuple(
                    str(row_a[headers_a[c] - 1].value or "").strip()
                    for c in key_a_cols
                )
                if key not in detail_sums:
                    no_detail_count += 1
                    continue

                detail_vals = detail_sums[key]
                has_diff = False
                for i, c in enumerate(cmp_a_cols):
                    summary_val = self._to_float(row_a[headers_a[c] - 1].value)
                    if abs(summary_val - detail_vals[i]) > 0.001:
                        has_diff = True
                        break
                if has_diff:
                    diff_count += 1
                    # 整行标黄
                    for cell in row_a:
                        cell.fill = YELLOW_FILL
                else:
                    match_count += 1

            # ----- 步骤3：检查明细有但汇总没有的 -----
            summary_keys = set()
            for row_a in ws_a.iter_rows(min_row=2, values_only=False):
                key = tuple(
                    str(row_a[headers_a[c] - 1].value or "").strip()
                    for c in key_a_cols
                )
                summary_keys.add(key)
            detail_only = set(detail_sums.keys()) - summary_keys

            self._log(f"对账结果：一致 {match_count} 行 / 不一致 {diff_count} 行（已标黄）")
            self._log(f"汇总无对应明细: {no_detail_count} 行")
            if detail_only:
                self._log(f"明细有但汇总无: {len(detail_only)} 条（仅记录，不输出）")

            # ----- 保存 -----
            wb_a.save(out_path)
            self._log(f"结果已保存: {out_path}")
            self._log("=" * 50)
            self._log("对账完成！")

            summary = (
                f"财务对账完成！\n\n"
                f"一致: {match_count} 行\n"
                f"不一致: {diff_count} 行（已标黄色高亮）\n"
                f"汇总无对应明细: {no_detail_count} 行\n"
                f"\n明细有但汇总无: {len(detail_only)} 条\n\n"
                f"已保存至:\n{out_path}"
            )
            messagebox.showinfo("对账完成", summary)

        except Exception as e:
            self._log(f"[错误] {e}")
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(
                state="normal", text="⚡ 开始对账"))

    # ==================== 工具方法 ====================
    @staticmethod
    def _to_float(val):
        """将单元格值转为 float，无法转换返回 0"""
        if val is None:
            return 0.0
        try:
            return float(val)
        except (ValueError, TypeError):
            return 0.0
