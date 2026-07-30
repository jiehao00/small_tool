# -*- coding: utf-8 -*-
"""
数据拆分（按分隔符拆分列→多行）页面模块

功能：对一个 Excel 文件，将选中列的单元格按分隔符拆分为多个值，
      每个值生成一行新数据，其余列内容原样复制。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")


def create_page(parent, app):
    """工厂函数：构建并返回数据拆分页面 Frame"""
    return _DataSplitPage(parent, app).frame


class _DataSplitPage:
    """数据拆分页面构建 & 业务逻辑"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.file_path = None     # 源文件路径
        self._col_headers = []   # 列名缓存

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # --- 页面标题 ---
        tk.Label(
            self.frame, text="★ 数据拆分（按分隔符拆分列 → 多行）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
                                     padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        row_file = tk.Frame(file_frame, bg=bg)
        row_file.pack(fill=tk.X)
        tk.Label(row_file, text="源 Excel 文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_file = ttk.Entry(row_file, style="Readonly.TEntry", state="readonly")
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_file, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_file).pack(side=tk.LEFT)

        # ---- 拆分配置 ----
        config_frame = ttk.LabelFrame(self.frame, text=" 拆分配置 ", style="Card.TLabelframe",
                                       padding=(16, 10))
        config_frame.pack(fill=tk.X, pady=(0, 12))

        # 第一行：拆分列（单选）
        col_frame = tk.Frame(config_frame, bg=bg)
        col_frame.pack(fill=tk.X, pady=(0, 10))

        col_row = tk.Frame(col_frame, bg=bg)
        col_row.pack(fill=tk.X)
        tk.Label(col_row, text="拆分列（单选）", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_col = ttk.Entry(col_row, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_col.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(col_row, text="🏷️ 选列", style="Secondary.TButton",
                   command=self._show_col_picker).pack(side=tk.LEFT)
        tk.Label(col_frame, text="提示：选择一个拆分列，该列内容将按分隔符拆为多行",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        # 第二行：分隔符
        sep_frame = tk.Frame(config_frame, bg=bg)
        sep_frame.pack(fill=tk.X)

        sep_row = tk.Frame(sep_frame, bg=bg)
        sep_row.pack(fill=tk.X)
        tk.Label(sep_row, text="分隔符", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_delimiter = ttk.Entry(sep_row, style="Normal.TEntry", width=28)
        self.entry_delimiter.insert(0, ",")
        self.entry_delimiter.pack(side=tk.LEFT)
        tk.Label(sep_frame, text="提示：常用分隔符包括逗号 , | 空格 | 换行 \\n | 分号 ; | 自定义任意字符",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=14)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始拆分",
                                     style="Action.TButton", width=22,
                                     command=self._start_split)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
                                   font=("Microsoft YaHei", 11, "bold"),
                                   bg=bg, fg="#2c3e50",
                                   padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== 交互 ====================
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择要拆分的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self.file_path = path
        self._set_readonly_value(self.entry_file, os.path.basename(path))
        self._load_columns()

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="拆分结果.xlsx"
        )
        if path:
            self._set_readonly_value(self.entry_out, path)

    def _load_columns(self):
        """加载选中文件的列名到缓存"""
        self._col_headers = []
        if not self.file_path:
            return
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                for name in row:
                    self._col_headers.append(str(name) if name is not None else "")
                self.log(f"[INFO] 已加载 {len(self._col_headers)} 个列名")
        except Exception as e:
            self.log(f"[ERROR] 读取列名失败: {e}")

    def _set_readonly_value(self, entry: ttk.Entry, value: str):
        """安全地给只读 Entry 赋值"""
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.configure(state="readonly")

    # ==================== 弹窗选列 ====================
    def _show_col_picker(self):
        """弹出窗口选择拆分列"""
        if not self._col_headers:
            messagebox.showwarning("提示", "请先选择源 Excel 文件以加载列名")
            return

        existing = self.entry_col.get().strip()

        popup = tk.Toplevel(self.parent)
        popup.title("选择拆分列")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()
        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="点击选择要拆分的列",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.SINGLE,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in self._col_headers:
            lb.insert(tk.END, col)

        if existing in self._col_headers:
            idx = self._col_headers.index(existing)
            lb.selection_set(idx)
            lb.see(idx)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            sel = lb.curselection()
            if sel:
                self._set_readonly_value(self.entry_col, lb.get(sel[0]))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, bg="#ecf0f1", relief="flat", cursor="hand2",
                  command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10, "bold"),
                  width=8, bg="#3498db", fg="white", relief="flat", cursor="hand2",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

        popup.bind("<Return>", lambda e: on_ok())

    # ==================== 日志 ====================
    def log(self, msg):
        """向日志区域追加一行文字"""
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ==================== 拆分入口 ====================
    def _start_split(self):
        """校验用户输入，启动后台线程执行拆分"""
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("依赖缺失", "缺少 openpyxl 库，请先安装：pip install openpyxl")
            return
        if not self.file_path:
            messagebox.showwarning("提示", "请先选择要拆分的 Excel 文件")
            return

        # 获取选中列
        split_col = self.entry_col.get().strip()
        if not split_col:
            messagebox.showwarning("提示", "请选择要拆分的列")
            return

        # 获取分隔符
        delimiter = self.entry_delimiter.get()
        if not delimiter:
            messagebox.showwarning("提示", "请输入分隔符")
            return
        # 处理转义字符
        delimiter = delimiter.replace("\\n", "\n").replace("\\t", "\t")

        # 获取输出配置
        out_path = self.entry_out.get().strip()
        if not out_path:
            messagebox.showwarning("提示", "请设置输出路径")
            return

        self.btn_start.configure(state=tk.DISABLED, text="拆分中，请稍候...")
        threading.Thread(
            target=self._do_split,
            args=(split_col, delimiter, out_path),
            daemon=True
        ).start()

    # ==================== 拆分逻辑 ====================
    def _do_split(self, split_col, delimiter, out_path):
        try:
            self.log(f"\n{'='*50}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 开始数据拆分...")
            self.log(f"  文件: {os.path.basename(self.file_path)}")
            self.log(f"  拆分列: {split_col}")
            self.log(f"  分隔符: '{delimiter}'")

            # ---- 读取文件 ----
            wb = load_workbook(self.file_path, read_only=True)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not all_rows:
                self.log("[ERROR] 文件为空")
                self._reset_btn("⚡ 开始拆分")
                return

            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = all_rows[1:]

            if split_col not in headers:
                self.log(f"[ERROR] 列 '{split_col}' 不存在，可用列: {', '.join(headers)}")
                self._reset_btn("⚡ 开始拆分")
                return

            self.log(f"  原始行数: {len(data_rows)}")
            self.log(f"  列数: {len(headers)}")

            col_idx = headers.index(split_col)

            # ---- 执行拆分 ----
            output_rows = []   # [(row_data, is_split), ...]
            split_count = 0
            total_original = len(data_rows)

            for row_data in data_rows:
                cell_value = row_data[col_idx]
                # 将单元格值转为字符串
                cell_str = str(cell_value) if cell_value is not None else ""
                parts = cell_str.split(delimiter) if cell_str else [""]

                # 去掉每个 part 两端空白（常见需求）
                parts = [p.strip() for p in parts]

                if len(parts) > 1:
                    # 拆分为多行
                    for i, part in enumerate(parts):
                        new_row = list(row_data)
                        new_row[col_idx] = part
                        output_rows.append((new_row, True))
                    split_count += 1
                else:
                    # 无需拆分，原样保留
                    output_rows.append((list(row_data), False))

            self.log(f"  拆分后总行数: {len(output_rows)}")
            self.log(f"  被拆分的行数: {split_count}")

            # ---- 写入输出文件 ----
            self._write_output(out_path, headers, output_rows)

            self.log(f"\n{'='*50}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 拆分完成！")
            self.log(f"  输出文件: {os.path.basename(out_path)}")
            self.log(f"  原始行数: {total_original} → 拆分后: {len(output_rows)}")
            self.log(f"  其中 {split_count} 行被拆分，拆分出的新行以黄底标记")
            self.log(f"{'='*50}")
            self._reset_btn("⚡ 开始拆分")

        except Exception as e:
            self.log(f"[ERROR] {e}")
            import traceback
            self.log(traceback.format_exc())
            self._reset_btn("⚡ 开始拆分")

    def _write_output(self, out_path, headers, output_rows):
        """将拆分结果写入 Excel，拆分出的行标黄"""
        self.log(f"  写入结果到 {os.path.basename(out_path)}...")

        out_wb = Workbook()
        out_ws = out_wb.active

        # 写入表头
        for j, h in enumerate(headers, 1):
            cell = out_ws.cell(row=1, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        # 写入数据行
        for i, (row_data, is_split) in enumerate(output_rows, 2):
            for j, val in enumerate(row_data, 1):
                cell = out_ws.cell(row=i, column=j, value=val)
                cell.alignment = CELL_ALIGNMENT
                if is_split:
                    cell.fill = YELLOW_FILL

        out_wb.save(out_path)
        self.log(f"  写入完成: {len(output_rows)} 行数据")

    def _reset_btn(self, text):
        """恢复按钮状态（需在主线程调用）"""
        self.frame.after(0, lambda: self.btn_start.configure(state=tk.NORMAL, text=text))
