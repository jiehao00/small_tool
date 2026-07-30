# -*- coding: utf-8 -*-
"""
多文件追加合并（纵向追加）页面模块

功能：选择多个 Excel 文件，按列名对齐后纵向追加合并为一个文件。
支持列对齐策略：
  - 取交集：仅保留所有文件都有的列
  - 取并集：保留所有列，缺失值留空
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

from .widgets import RoundedCheckbox

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")


def create_page(parent, app):
    """工厂函数：构建并返回多文件追加合并页面 Frame"""
    return _DataAppendPage(parent, app).frame


class _DataAppendPage:
    """多文件追加合并页面"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.file_paths = []        # 已选择的文件路径列表

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # --- 页面标题 ---
        tk.Label(
            self.frame, text="★ 多文件追加合并（纵向追加）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
                                     padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        btn_row = tk.Frame(file_frame, bg=bg)
        btn_row.pack(fill=tk.X)
        ttk.Button(btn_row, text="📂 添加文件", style="Secondary.TButton",
                   command=self._add_files).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="❌ 移除选中", style="Secondary.TButton",
                   command=self._remove_selected).pack(side=tk.LEFT, padx=(0, 8))
        ttk.Button(btn_row, text="🗑 清空列表", style="Secondary.TButton",
                   command=self._clear_all).pack(side=tk.LEFT)

        # 文件列表
        list_container = tk.Frame(file_frame, bg=bg)
        list_container.pack(fill=tk.BOTH, expand=True, pady=(8, 0))

        self.file_listbox = tk.Listbox(
            list_container, font=("Microsoft YaHei", 9), height=6,
            bg="white", fg="#2c3e50", relief="flat",
            selectbackground="#3498db", selectforeground="white",
            selectmode=tk.EXTENDED, exportselection=False
        )
        list_scroll = tk.Scrollbar(list_container, orient=tk.VERTICAL)
        self.file_listbox.config(yscrollcommand=list_scroll.set)
        list_scroll.config(command=self.file_listbox.yview)
        self.file_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        list_scroll.pack(side=tk.RIGHT, fill=tk.Y)

        # 拖拽排序支持
        self._drag_data = {"index": None, "widget": None}
        self.file_listbox.bind("<ButtonPress-1>", self._on_drag_start)
        self.file_listbox.bind("<B1-Motion>", self._on_drag_motion)
        self.file_listbox.bind("<ButtonRelease-1>", self._on_drag_end)

        tk.Label(file_frame, text="提示：文件将按列表顺序从上到下追加，可通过拖拽或移除/重新添加调整顺序",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6"
                 ).pack(anchor=tk.W, pady=(2, 0))

        # ---- 合并选项 ----
        opt_frame = ttk.LabelFrame(self.frame, text=" 合并选项 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        opt_frame.pack(fill=tk.X, pady=(0, 12))

        # 列对齐策略
        self.align_var = tk.StringVar(value="union")
        align_row = tk.Frame(opt_frame, bg=bg)
        align_row.pack(fill=tk.X)
        tk.Label(align_row, text="列对齐策略", font=("Microsoft YaHei", 10, "bold"),
                 bg=bg, fg="#34495e", width=14, anchor=tk.E
                 ).pack(side=tk.LEFT, padx=(0, 12))

        # 分段控件 —— Canvas 绘制，风格统一
        self._seg_opts = [
            ("取并集（保留所有列，缺失值留空）", "union"),
            ("取交集（仅保留共有列）", "intersect"),
        ]
        seg_h = 34
        seg_c = tk.Canvas(align_row, height=seg_h, bg=bg, highlightthickness=0)
        seg_c.pack(side=tk.LEFT)

        SEL_BG = "#3498db"; SEL_FG = "white"
        BAR_BG = "#dfe3e8";  IDL_FG = "#555"
        PAD_X = 4; RADIUS = 6

        def _round_rect(canvas, x1, y1, x2, y2, r, **kw):
            """画圆角矩形"""
            points = (x1 + r, y1, x2 - r, y1, x2, y1, x2, y1 + r,
                      x2, y2 - r, x2, y2, x2 - r, y2,
                      x1 + r, y2, x1, y2, x1, y2 - r,
                      x1, y1 + r, x1, y1)
            return canvas.create_polygon(points, smooth=True, **kw)

        def _seg_redraw():
            seg_c.delete("all")
            # 测量两段宽度
            widths = []
            for text, _ in self._seg_opts:
                tw = seg_c.create_text(0, 0, text=text,
                       font=("Microsoft YaHei", 10), anchor="w")
                bbox = seg_c.bbox(tw)
                widths.append(bbox[2] - bbox[0])
                seg_c.delete(tw)
            total_w = sum(widths) + PAD_X * 4 + 4
            seg_c.configure(width=total_w)

            # 背景条
            _round_rect(seg_c, 0, 0, total_w, seg_h, r=RADIUS, fill=BAR_BG, outline="")

            # 两个分段
            cur_val = self.align_var.get()
            cx = PAD_X
            for i, (text, val) in enumerate(self._seg_opts):
                seg_w = widths[i] + PAD_X * 2
                ex = cx + seg_w
                is_sel = (val == cur_val)
                if is_sel:
                    _round_rect(seg_c, cx, 2, ex, seg_h - 2, r=RADIUS,
                                fill=SEL_BG, outline="")
                seg_c.create_text(cx + seg_w / 2, seg_h / 2, text=text,
                                  font=("Microsoft YaHei", 10),
                                  fill=SEL_FG if is_sel else IDL_FG,
                                  anchor="center", tags=f"s_{val}")
                cx = ex

        def _on_seg_click(event):
            for _, val in self._seg_opts:
                found = seg_c.find_withtag(f"s_{val}")
                if found:
                    bbox = seg_c.bbox(found[0])
                    if bbox and bbox[0] <= event.x <= bbox[2] and bbox[1] <= event.y <= bbox[3]:
                        self.align_var.set(val)
                        return

        seg_c.bind("<Button-1>", _on_seg_click)
        seg_c._redraw = _seg_redraw  # store for external resize
        self._seg_canvas = seg_c

        def _update_seg(*args):
            if hasattr(self, '_seg_canvas') and self._seg_canvas.winfo_exists():
                self._seg_canvas._redraw()

        _update_seg()
        self.align_var.trace_add("write", _update_seg)
        seg_c.after(100, _update_seg)  # 延迟一次确保字体测量准确

        # 首行表头
        header_row = tk.Frame(opt_frame, bg=bg)
        header_row.pack(fill=tk.X, pady=(10, 0))
        tk.Label(header_row, text="", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.header_var = tk.BooleanVar(value=True)
        hdr_cb = tk.Frame(header_row, bg=bg)
        hdr_cb.pack(side=tk.LEFT)
        RoundedCheckbox(hdr_cb, variable=self.header_var, bg=bg).pack(side=tk.LEFT)
        tk.Label(hdr_cb, text=" 每个文件的第一行是列名（仅用于对齐，不作为数据追加）",
                 font=("Microsoft YaHei", 10), bg=bg,
                 fg="#2c3e50").pack(side=tk.LEFT)

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=14)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始合并",
                                     style="Action.TButton", width=22,
                                     command=self._start_append)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
                                   font=("Microsoft YaHei", 11, "bold"),
                                   bg=bg, fg="#2c3e50",
                                   padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== 文件列表面板交互 ====================
    def _add_files(self):
        paths = filedialog.askopenfilenames(
            title="选择要合并的 Excel 文件（可多选）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")]
        )
        if not paths:
            return
        for p in paths:
            if p not in self.file_paths:
                self.file_paths.append(p)
                self.file_listbox.insert(tk.END, os.path.basename(p))
        self.log(f"已添加 {len(paths)} 个文件，当前共 {len(self.file_paths)} 个")

    def _remove_selected(self):
        sel = sorted(self.file_listbox.curselection(), reverse=True)
        if not sel:
            messagebox.showwarning("提示", "请先在列表中选择要移除的文件")
            return
        for i in sel:
            self.file_listbox.delete(i)
            del self.file_paths[i]
        self.log(f"已移除 {len(sel)} 个文件，当前共 {len(self.file_paths)} 个")

    def _clear_all(self):
        self.file_listbox.delete(0, tk.END)
        self.file_paths.clear()
        self.log("文件列表已清空")

    # ==================== 拖拽排序 ====================
    def _on_drag_start(self, event):
        """记录拖拽起始位置"""
        idx = self.file_listbox.nearest(event.y)
        if idx >= 0:
            self._drag_data["index"] = idx
            self._drag_data["widget"] = self.file_listbox
            # 高亮当前拖拽项
            self.file_listbox.selection_clear(0, tk.END)
            self.file_listbox.selection_set(idx)

    def _on_drag_motion(self, event):
        """拖拽移动时，将当前项插入到目标位置"""
        if self._drag_data["widget"] is None:
            return
        target = self.file_listbox.nearest(event.y)
        src = self._drag_data["index"]
        if target < 0 or target == src:
            return

        # 在 Listbox 中移动
        item_text = self.file_listbox.get(src)
        self.file_listbox.delete(src)
        self.file_listbox.insert(target, item_text)

        # 同步 file_paths 列表
        item_path = self.file_paths.pop(src)
        self.file_paths.insert(target, item_path)

        # 更新拖拽索引到新位置
        self._drag_data["index"] = target
        self.file_listbox.selection_clear(0, tk.END)
        self.file_listbox.selection_set(target)

    def _on_drag_end(self, event):
        """拖拽结束，清除拖拽状态"""
        self._drag_data["index"] = None
        self._drag_data["widget"] = None

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="合并结果.xlsx"
        )
        if path:
            self._set_readonly_value(self.entry_out, path)

    def _set_readonly_value(self, entry: ttk.Entry, value: str):
        """安全地给只读 Entry 赋值"""
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.configure(state="readonly")

    # ==================== 日志 ====================
    def log(self, msg):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ==================== 合并入口 ====================
    def _start_append(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("依赖缺失", "缺少 openpyxl 库，请先安装：pip install openpyxl")
            return
        if len(self.file_paths) < 2:
            messagebox.showwarning("提示", "请至少添加 2 个文件")
            return

        out_path = self.entry_out.get().strip()
        if not out_path:
            messagebox.showwarning("提示", "请设置输出路径")
            return

        align_strategy = self.align_var.get()
        has_header = self.header_var.get()

        self.btn_start.configure(state=tk.DISABLED, text="合并中，请稍候...")
        threading.Thread(
            target=self._do_append,
            args=(out_path, align_strategy, has_header),
            daemon=True
        ).start()

    # ==================== 合并逻辑 ====================
    def _do_append(self, out_path, align_strategy, has_header):
        try:
            self.log(f"\n{'='*60}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 开始多文件追加合并...")
            self.log(f"  文件总数: {len(self.file_paths)}")
            self.log(f"  列对齐策略: {'取并集' if align_strategy == 'union' else '取交集'}")
            self.log(f"  首行为表头: {'是' if has_header else '否'}")

            # ---- 读取所有文件的列名和数据 ----
            all_headers = []     # 每个文件的列名列表
            all_data = []        # 每个文件的(rows, start_row, end_row) 元组
            total_rows = 0
            final_headers = None

            for idx, fpath in enumerate(self.file_paths):
                fname = os.path.basename(fpath)
                self.log(f"  [{idx+1}/{len(self.file_paths)}] 读取: {fname}")

                wb = load_workbook(fpath, read_only=True)
                ws = wb.active
                raw_rows = list(ws.iter_rows(values_only=True))
                wb.close()

                if not raw_rows:
                    self.log(f"    [警告] 文件为空，跳过")
                    all_headers.append([])
                    all_data.append([])
                    continue

                if has_header:
                    headers = [str(h) if h is not None else f"Column{j+1}"
                               for j, h in enumerate(raw_rows[0])]
                    data_rows = raw_rows[1:]
                    start_row = 2  # 数据从第2行开始（1-indexed）
                else:
                    headers = [f"Column{j+1}" for j in range(len(raw_rows[0]))]
                    data_rows = raw_rows
                    start_row = 1

                all_headers.append(headers)
                all_data.append((data_rows, start_row, fname))
                total_rows += len(data_rows)
                self.log(f"    {len(headers)} 列 × {len(data_rows)} 行")

            # ---- 计算最终列顺序 ----
            if align_strategy == "intersect":
                # 取交集：所有文件都有的列（保持第一个文件的列顺序）
                if not all_headers or not all_headers[0]:
                    self.log("[ERROR] 没有有效数据")
                    self._reset_btn("⚡ 开始合并")
                    return
                common = set(all_headers[0])
                for h in all_headers[1:]:
                    if h:  # 跳过空文件
                        common = common & set(h)
                # 按第一个文件的顺序排列
                final_headers = [h for h in all_headers[0] if h in common]
                if not final_headers:
                    self.log("[ERROR] 所有文件没有共同的列，无法取交集合并")
                    self.log("  提示：请切换为「取并集」策略重试")
                    self._reset_btn("⚡ 开始合并")
                    return
                self.log(f"  交集列数: {len(final_headers)}")
            else:
                # 取并集：收集所有列，按第一个文件的顺序，再追加新列
                seen = set()
                final_headers = []
                for h_list in all_headers:
                    if not h_list:
                        continue
                    for h in h_list:
                        if h not in seen:
                            seen.add(h)
                            final_headers.append(h)
                self.log(f"  并集列数: {len(final_headers)}")

            # ---- 写入输出文件 ----
            self.log(f"\n  写入结果到 {os.path.basename(out_path)}...")
            self._write_output(out_path, final_headers, all_headers, all_data)

            self.log(f"\n{'='*60}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 合并完成！")
            self.log(f"  输出文件: {os.path.basename(out_path)}")
            self.log(f"  合并文件数: {len(self.file_paths)}")
            self.log(f"  总数据行数: {total_rows}")
            self.log(f"  最终列数: {len(final_headers)}")
            self.log(f"{'='*60}")
            self._reset_btn("⚡ 开始合并")

        except Exception as e:
            self.log(f"[ERROR] {e}")
            import traceback
            self.log(traceback.format_exc())
            self._reset_btn("⚡ 开始合并")

    def _write_output(self, out_path, final_headers, all_headers, all_data):
        """写入合并后的 Excel，每个文件的数据来源写入备注列"""
        out_wb = Workbook()
        out_ws = out_wb.active

        # 追加一列标注数据来源
        source_col_name = "_来源文件"
        output_headers = final_headers + [source_col_name]

        # 写表头
        for j, h in enumerate(output_headers, 1):
            cell = out_ws.cell(row=1, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        # 写数据
        out_row = 2
        for file_idx, (headers, (data_rows, start_row, fname)) in enumerate(zip(all_headers, all_data)):
            if not headers:
                continue
            # 构建列索引映射：输出列 → 源文件中的列索引
            col_map = {}
            for out_j, col_name in enumerate(final_headers):
                if col_name in headers:
                    col_map[out_j] = headers.index(col_name)
                else:
                    col_map[out_j] = -1  # 该列不存在

            for row_data in data_rows:
                for out_j in range(len(final_headers)):
                    src_idx = col_map[out_j]
                    val = row_data[src_idx] if src_idx >= 0 and src_idx < len(row_data) else ""
                    cell = out_ws.cell(row=out_row, column=out_j + 1, value=val)
                    cell.alignment = CELL_ALIGNMENT

                # 来源列
                src_cell = out_ws.cell(row=out_row, column=len(output_headers), value=fname)
                src_cell.alignment = CELL_ALIGNMENT

                out_row += 1

        out_wb.save(out_path)
        self.log(f"  写入完成: {out_row - 2} 行数据")

    def _reset_btn(self, text):
        """恢复按钮状态（需在主线程调用）"""
        self.frame.after(0, lambda: self.btn_start.configure(state=tk.NORMAL, text=text))
