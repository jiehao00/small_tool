# -*- coding: utf-8 -*-
"""
按列值拆分文件页面模块

功能：选择一个 Excel 文件，指定拆分列，按该列不同取值将数据拆分
      为多个独立的 Excel/PNG/PDF 文件，每个文件只包含对应取值的行。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import re
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

from .widgets import RoundedCheckbox

# ---- 可选库：PNG / PDF 导出 ----
try:
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    import matplotlib.font_manager as fm
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False

try:
    from fpdf import FPDF
    FPDF2_AVAILABLE = True
except ImportError:
    FPDF2_AVAILABLE = False

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")


# ---------- 字体工具 ----------

def _find_chinese_font():
    """查找系统中可用的中文字体"""
    # 优先直接查 Windows 字体目录
    search_paths = [
        r"C:\Windows\Fonts\msyh.ttc",
        r"C:\Windows\Fonts\simhei.ttf",
        r"C:\Windows\Fonts\msyhbd.ttc",
        r"C:\Windows\Fonts\simsun.ttc",
    ]
    for p in search_paths:
        if os.path.exists(p):
            return p
    # 通过 matplotlib 系统字体列表查找
    if MATPLOTLIB_AVAILABLE:
        candidates = ["msyh", "simhei", "simsun", "microsoft yahei"]
        for fp in fm.findSystemFonts():
            low = os.path.basename(fp).lower().replace("_", "").replace("-", "")
            for c in candidates:
                if c.replace(" ", "") in low:
                    return fp
    return None


_CHINESE_FONT_PATH = _find_chinese_font()

# matplotlib 全局字体设置
if MATPLOTLIB_AVAILABLE and _CHINESE_FONT_PATH:
    fm.fontManager.addfont(_CHINESE_FONT_PATH)
    prop = fm.FontProperties(fname=_CHINESE_FONT_PATH)
    plt.rcParams["font.family"] = prop.get_name()


def _table_to_png(headers, rows, out_path):
    if not MATPLOTLIB_AVAILABLE:
        raise RuntimeError("matplotlib 未安装")
    ncols = len(headers)
    nrows = len(rows)
    col_width = max(1.8, 12.0 / max(ncols, 1))
    row_height = 0.4
    fig_w = col_width * ncols + 1.5
    fig_h = max(2.0, row_height * (nrows + 1) + 1.2)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.axis("off")
    font_prop = fm.FontProperties(fname=_CHINESE_FONT_PATH, size=10) if _CHINESE_FONT_PATH else None

    cell_text = [[str(v) if v is not None else "" for v in row] for row in rows]
    table = ax.table(cellText=cell_text, colLabels=headers, loc="center", cellLoc="center")
    table.auto_set_font_size(False)
    table.set_fontsize(8)
    table.scale(1.0, 1.5)

    for key, cell in table.get_celld().items():
        r, _ = key
        if r == 0:
            cell.set_facecolor("#3498db")
            cell.set_text_props(color="white", fontweight="bold")
        else:
            cell.set_facecolor("#f8f9fa" if r % 2 == 0 else "white")
        cell.set_edgecolor("#bdc3c7")
        cell.set_linewidth(0.3)
        if font_prop:
            cell.get_text().set_fontproperties(font_prop)

    plt.tight_layout(pad=0.5)
    plt.savefig(out_path, dpi=150, bbox_inches="tight", facecolor="white")
    plt.close(fig)


def _table_to_pdf(headers, rows, out_path, title_text):
    """PDF 表格列数过多时整体换行：表头+数据成对向下排列"""
    if not FPDF2_AVAILABLE:
        raise RuntimeError("fpdf2 未安装")
    font_path = None
    for p in [r"C:\Windows\Fonts\simhei.ttf", r"C:\Windows\Fonts\msyh.ttc"]:
        if os.path.exists(p):
            font_path = p
            break
    pdf = FPDF(orientation="L", unit="mm", format="A4")
    if font_path:
        pdf.add_font("cjk", "", font_path)
    else:
        pdf.add_font("cjk", "", r"C:\Windows\Fonts\simhei.ttf")

    ncols = len(headers)
    available_w = pdf.w - 2 * pdf.l_margin
    line_h = 6
    min_col_w = 25                       # 单列最小宽度
    page_bottom = pdf.h - pdf.b_margin

    # 每行能放多少列（列太多就换行）
    cols_per_line = int(available_w / min_col_w)
    cols_per_line = max(1, min(cols_per_line, ncols))
    col_w = available_w / cols_per_line

    def _render_cells(cells, is_header=False, fill_color=None):
        """渲染一行单元格，文字自动换行，行高自适应"""
        font_size = 9 if is_header else 8
        pdf.set_font("cjk", "", font_size)

        # 预计算每列折行后的行数
        cell_lines = []
        max_lines = 1
        for item in cells:
            txt = str(item) if item is not None else ""
            lines = pdf.multi_cell(col_w, line_h, txt, split_only=True)
            cell_lines.append(lines)
            max_lines = max(max_lines, len(lines))

        row_height = max_lines * line_h

        # 页面溢出检查
        if pdf.get_y() + row_height > page_bottom:
            pdf.add_page()

        x_start = pdf.l_margin
        y_start = pdf.get_y()

        for i, txt_lines in enumerate(cell_lines):
            x_pos = x_start + i * col_w

            # 背景 & 边框
            if is_header:
                pdf.set_fill_color(52, 152, 219)
                pdf.set_text_color(255, 255, 255)
            else:
                pdf.set_text_color(0, 0, 0)
                if fill_color:
                    pdf.set_fill_color(*fill_color)
                else:
                    pdf.set_fill_color(255, 255, 255)
            pdf.rect(x_pos, y_start, col_w, row_height, style="DF")

            # 文字垂直居中
            text_y_start = y_start + (row_height - len(txt_lines) * line_h) / 2
            pdf.set_font("cjk", "", font_size)
            for j, line in enumerate(txt_lines):
                pdf.set_xy(x_pos + 0.5, text_y_start + j * line_h + 0.5)
                pdf.cell(col_w - 1, line_h, line, align="C")

        pdf.set_xy(x_start, y_start + row_height)

    # ===== 开始渲染 =====
    pdf.add_page()
    if title_text:
        pdf.set_font("cjk", "", 10)
        pdf.cell(0, 8, title_text, new_x="LMARGIN", new_y="NEXT", align="C")
        pdf.ln(3)

    for row_idx, row in enumerate(rows):
        # 每条数据按列块分组：表头 + 数据 成对向下排列
        for col_start in range(0, ncols, cols_per_line):
            col_end = min(col_start + cols_per_line, ncols)
            chunk_headers = headers[col_start:col_end]
            chunk_values = [row[i] if i < len(row) else "" for i in range(col_start, col_end)]

            _render_cells(chunk_headers, is_header=True)
            _render_cells(chunk_values,
                          fill_color=(248, 249, 250) if row_idx % 2 == 0 else (255, 255, 255))

        # 行与行之间留小间隙
        pdf.ln(3)

    pdf.output(out_path)


# ---------- 页面 ----------

def create_page(parent, app):
    """工厂函数：构建并返回按列值拆分文件页面 Frame"""
    return _ColumnSplitPage(parent, app).frame


class _ColumnSplitPage:
    """按列值拆分文件页面"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.file_path = None          # 源文件路径
        self.out_dir = None            # 输出目录
        self._col_headers = []         # 列名缓存

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # --- 页面标题 ---
        tk.Label(
            self.frame, text="★ 按列值拆分文件（支持多列组合拆分）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
                                     padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        row_file = tk.Frame(file_frame, bg=bg)
        row_file.pack(fill=tk.X)
        tk.Label(row_file, text="源 Excel 文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_file = ttk.Entry(row_file, style="Readonly.TEntry", state="readonly")
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_file, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_file).pack(side=tk.LEFT)

        # ---- 拆分依据列 ----
        col_frame = ttk.LabelFrame(self.frame, text=" 拆分依据列 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        col_frame.pack(fill=tk.X, pady=(0, 12))

        # 第一行：拆分依据列
        key_col = tk.Frame(col_frame, bg=bg)
        key_col.pack(fill=tk.X)

        col_row = tk.Frame(key_col, bg=bg)
        col_row.pack(fill=tk.X)
        tk.Label(col_row, text="拆分依据列",
                 font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_cols = ttk.Entry(col_row, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_cols.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(col_row, text="🏷️ 选列", style="Secondary.TButton",
                   command=self._show_col_picker).pack(side=tk.LEFT)
        tk.Label(key_col, text="提示：可按多列组合值分组；选择文件后点击按钮选取拆分列，支持 Ctrl 多选",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        # 输出目录
        dir_col = tk.Frame(out_frame, bg=bg)
        dir_col.pack(fill=tk.X, pady=(0, 10))

        dir_row = tk.Frame(dir_col, bg=bg)
        dir_row.pack(fill=tk.X)
        tk.Label(dir_row, text="输出目录", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_dir = ttk.Entry(dir_row, style="Readonly.TEntry", state="readonly", width=32)
        self.entry_dir.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(dir_row, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_dir).pack(side=tk.LEFT)
        tk.Label(dir_col,
                 text="提示：每个唯一组合值将生成对应文件，如：结果_北京.xlsx / 结果_上海.png（取决于所选格式）",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        name_row = tk.Frame(out_frame, bg=bg)
        name_row.pack(fill=tk.X)
        tk.Label(name_row, text="文件名前缀", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_prefix = ttk.Entry(name_row, style="Normal.TEntry")
        self.entry_prefix.insert(0, "拆分结果_")
        self.entry_prefix.pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Label(name_row, text="生成格式：前缀 + 组合值 + .xlsx / .png / .pdf",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6"
                 ).pack(side=tk.LEFT, padx=(8, 0))

        # ---- 输出格式选择（无框平铺）----
        fmt_row = tk.Frame(out_frame, bg=bg)
        fmt_row.pack(fill=tk.X, pady=(8, 0))

        tk.Label(fmt_row, text="输出格式", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))

        # 复选框容器
        fmt_inner = tk.Frame(fmt_row, bg=bg)
        fmt_inner.pack(side=tk.LEFT, fill=tk.X, expand=True)

        self.fmt_excel = tk.BooleanVar(value=True)
        self.fmt_png = tk.BooleanVar(value=False)
        self.fmt_pdf = tk.BooleanVar(value=False)

        cb1 = tk.Frame(fmt_inner, bg=bg)
        cb1.pack(side=tk.LEFT, padx=(0, 16))
        RoundedCheckbox(cb1, variable=self.fmt_excel, bg=bg).pack(side=tk.LEFT)
        tk.Label(cb1, text=" Excel (.xlsx)", font=("Microsoft YaHei", 10),
                 bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        cb2 = tk.Frame(fmt_inner, bg=bg)
        cb2.pack(side=tk.LEFT, padx=(0, 16))
        RoundedCheckbox(cb2, variable=self.fmt_png, bg=bg).pack(side=tk.LEFT)
        tk.Label(cb2, text=" PNG (.png)", font=("Microsoft YaHei", 10),
                 bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        cb3 = tk.Frame(fmt_inner, bg=bg)
        cb3.pack(side=tk.LEFT)
        RoundedCheckbox(cb3, variable=self.fmt_pdf, bg=bg).pack(side=tk.LEFT)
        tk.Label(cb3, text=" PDF (.pdf)", font=("Microsoft YaHei", 10),
                 bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        def toggle_all():
            new_val = not (self.fmt_excel.get() and self.fmt_png.get() and self.fmt_pdf.get())
            self.fmt_excel.set(new_val)
            self.fmt_png.set(new_val)
            self.fmt_pdf.set(new_val)

        # 全选按钮放在最右侧，使用与浏览按钮一致的样式
        def _update_toggle_text(*args):
            all_selected = self.fmt_excel.get() and self.fmt_png.get() and self.fmt_pdf.get()
            self.btn_toggle.config(text="✗ 全不选" if all_selected else "✓ 全选")

        self.btn_toggle = tk.Button(fmt_inner, text="✓ 全选",
                                     font=("Microsoft YaHei", 9),
                                     bg="#3498db", fg="white", relief="flat",
                                     activebackground="#2980b9", activeforeground="white",
                                     cursor="hand2", padx=8, pady=2,
                                     command=toggle_all)
        self.btn_toggle.pack(side=tk.RIGHT)

        self.fmt_excel.trace_add("write", _update_toggle_text)
        self.fmt_png.trace_add("write", _update_toggle_text)
        self.fmt_pdf.trace_add("write", _update_toggle_text)
        _update_toggle_text()

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=14)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始拆分",
                                     style="Action.TButton", width=22,
                                     command=self._start_split)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
                                   font=("Microsoft YaHei", 11, "bold"),
                                   bg=bg, fg="#2c3e50",
                                   padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== 交互 ====================
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择要拆分的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self.file_path = path
        self._set_readonly_value(self.entry_file, os.path.basename(path))
        self._load_columns()

    def _browse_dir(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self.out_dir = path
            self._set_readonly_value(self.entry_dir, path)

    def _load_columns(self):
        """加载选中文件的列名到缓存"""
        self._col_headers = []
        if not self.file_path:
            return
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                for name in row:
                    self._col_headers.append(str(name) if name is not None else "")
                self.log(f"[INFO] 已加载 {len(self._col_headers)} 个列名")
        except Exception as e:
            self.log(f"[ERROR] 读取列名失败: {e}")

    def _set_readonly_value(self, entry: ttk.Entry, value: str):
        """安全地给只读 Entry 赋值"""
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.configure(state="readonly")

    # ==================== 弹窗选列 ====================
    def _show_col_picker(self):
        """弹出窗口选择拆分依据列"""
        if not self._col_headers:
            messagebox.showwarning("提示", "请先选择源 Excel 文件以加载列名")
            return

        existing = [c.strip() for c in self.entry_cols.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择拆分依据列")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()
        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.EXTENDED,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in self._col_headers:
            lb.insert(tk.END, col)

        sel_order = []
        for i, col in enumerate(self._col_headers):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            sel_order[:] = [i for i in sel_order if i not in removed]
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            self._set_readonly_value(self.entry_cols, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, bg="#ecf0f1", relief="flat", cursor="hand2",
                  command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10, "bold"),
                  width=8, bg="#3498db", fg="white", relief="flat", cursor="hand2",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

        popup.bind("<Return>", lambda e: on_ok())

    # ==================== 日志 ====================
    def log(self, msg):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ==================== 拆分入口 ====================
    def _start_split(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("依赖缺失", "缺少 openpyxl 库，请先安装：pip install openpyxl")
            return
        if not self.file_path:
            messagebox.showwarning("提示", "请先选择要拆分的 Excel 文件")
            return

        split_cols = [c.strip() for c in self.entry_cols.get().strip().split(",") if c.strip()]
        if not split_cols:
            messagebox.showwarning("提示", "请选择拆分依据列（可按住 Ctrl 多选）")
            return

        if not self.out_dir:
            messagebox.showwarning("提示", "请选择输出目录")
            return

        prefix = self.entry_prefix.get().strip()
        if not prefix:
            prefix = "拆分结果_"

        use_excel = self.fmt_excel.get()
        use_png = self.fmt_png.get()
        use_pdf = self.fmt_pdf.get()
        if not (use_excel or use_png or use_pdf):
            messagebox.showwarning("提示", "请至少选择一种输出格式")
            return

        self.btn_start.configure(state=tk.DISABLED, text="拆分中，请稍候...")
        threading.Thread(
            target=self._do_split,
            args=(split_cols, prefix, use_excel, use_png, use_pdf),
            daemon=True
        ).start()

    # ==================== 拆分逻辑 ====================
    def _do_split(self, split_cols, prefix, use_excel, use_png, use_pdf):
        try:
            self.log(f"\n{'='*60}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 开始按列值拆分文件...")
            self.log(f"  文件: {os.path.basename(self.file_path)}")
            self.log(f"  拆分列: {', '.join(split_cols)}")
            self.log(f"  输出目录: {self.out_dir}")
            fmt_list = []
            if use_excel: fmt_list.append("Excel")
            if use_png: fmt_list.append("PNG")
            if use_pdf: fmt_list.append("PDF")
            self.log(f"  输出格式: {', '.join(fmt_list)}")

            # ---- 读取文件 ----
            wb = load_workbook(self.file_path)
            ws = wb.active
            all_rows = list(ws.iter_rows(values_only=True))
            wb.close()

            if not all_rows:
                self.log("[ERROR] 文件为空")
                self._reset_btn("⚡ 开始拆分")
                return

            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = all_rows[1:]

            # 校验所有拆分列
            missing = [c for c in split_cols if c not in headers]
            if missing:
                self.log(f"[ERROR] 列 {', '.join(missing)} 不存在，可用列: {', '.join(headers)}")
                self._reset_btn("⚡ 开始拆分")
                return

            col_indices = [headers.index(c) for c in split_cols]
            self.log(f"  总列数: {len(headers)}")
            self.log(f"  总行数: {len(data_rows)}")

            # ---- 按多列组合值分组 ----
            groups = {}  # (val1, val2, ...) → [rows]
            for row_data in data_rows:
                key = tuple(
                    str(row_data[i]) if row_data[i] is not None else "(空值)"
                    for i in col_indices
                )
                if key not in groups:
                    groups[key] = []
                groups[key].append(row_data)

            self.log(f"  唯一取值数量: {len(groups)}")

            # ---- 创建子目录 ----
            try:
                os.makedirs(self.out_dir, exist_ok=True)
                if use_excel:
                    os.makedirs(os.path.join(self.out_dir, "excel"), exist_ok=True)
                if use_png:
                    os.makedirs(os.path.join(self.out_dir, "png"), exist_ok=True)
                if use_pdf:
                    os.makedirs(os.path.join(self.out_dir, "pdf"), exist_ok=True)
            except Exception as e:
                self.log(f"[ERROR] 创建输出目录失败: {e}")
                self._reset_btn("⚡ 开始拆分")
                return

            # ---- 处理非法字符函数 ----
            def safe_filename(name):
                s = str(name)
                s = re.sub(r'[\\/*?:"<>|]', '_', s)
                s = s.strip().strip(".")
                return s or "unnamed"

            written_count = 0

            for value, rows in groups.items():
                # 多列组合值用下划线拼接
                safe_val = safe_filename("_".join(value) if isinstance(value, tuple) else value)

                # --- Excel ---
                if use_excel:
                    try:
                        out_path = os.path.join(self.out_dir, "excel", f"{prefix}{safe_val}.xlsx")
                        self.log(f"  [Excel] 生成: {os.path.basename(out_path)} ({len(rows)} 行)")
                        self._save_excel(headers, rows, out_path)
                        written_count += 1
                    except Exception as e:
                        self.log(f"  [Excel] 生成失败 [{safe_val}]: {e}")

                # --- PNG ---
                if use_png:
                    try:
                        if not MATPLOTLIB_AVAILABLE:
                            self.log(f"  [PNG] 跳过 [{safe_val}]: matplotlib 未安装")
                        else:
                            out_path = os.path.join(self.out_dir, "png", f"{prefix}{safe_val}.png")
                            self.log(f"  [PNG] 生成: {os.path.basename(out_path)} ({len(rows)} 行)")
                            disp_r = [[r[i] if i < len(r) else "" for i in range(len(headers))]
                                      for r in rows]
                            _table_to_png(headers, disp_r, out_path)
                            written_count += 1
                    except Exception as e:
                        self.log(f"  [PNG] 生成失败 [{safe_val}]: {e}")

                # --- PDF ---
                if use_pdf:
                    try:
                        if not FPDF2_AVAILABLE:
                            self.log(f"  [PDF] 跳过 [{safe_val}]: fpdf2 未安装")
                        else:
                            out_path = os.path.join(self.out_dir, "pdf", f"{prefix}{safe_val}.pdf")
                            self.log(f"  [PDF] 生成: {os.path.basename(out_path)} ({len(rows)} 行)")
                            disp_r = [[r[i] if i < len(r) else "" for i in range(len(headers))]
                                      for r in rows]
                            _table_to_pdf(headers, disp_r, out_path, title_text="")
                            written_count += 1
                    except Exception as e:
                        self.log(f"  [PDF] 生成失败 [{safe_val}]: {e}")

            self.log(f"\n{'='*60}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 拆分完成！")
            self.log(f"  输出目录: {self.out_dir}")
            self.log(f"  拆分依据列: {', '.join(split_cols)}")
            self.log(f"  生成文件数: {written_count}")
            self.log(f"  唯一取值组合数: {len(groups)}")
            self.log(f"{'='*60}")
            self._reset_btn("⚡ 开始拆分")

        except Exception as e:
            self.log(f"[ERROR] {e}")
            import traceback
            self.log(traceback.format_exc())
            self._reset_btn("⚡ 开始拆分")

    def _save_excel(self, headers, rows, out_path):
        """保存分组数据为新 Excel 文件"""
        out_wb = Workbook()
        out_ws = out_wb.active

        for j, h in enumerate(headers, 1):
            cell = out_ws.cell(row=1, column=j, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER

        for i, row_data in enumerate(rows, 2):
            for j, val in enumerate(row_data, 1):
                cell = out_ws.cell(row=i, column=j, value=val)
                cell.alignment = CELL_ALIGNMENT

        out_wb.save(out_path)

    def _reset_btn(self, text):
        self.frame.after(0, lambda: self.btn_start.configure(state=tk.NORMAL, text=text))
