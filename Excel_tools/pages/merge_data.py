# -*- coding: utf-8 -*-
"""
填充数据（Excel 内容填充）页面模块

功能：当 A 文件的匹配列与 B 文件的匹配列相匹配，
      且 A 文件的内容填充列为空时，
      将 B 文件的内容列填充到 A 文件中，输出新文件。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

# 表头样式（供所有输出 Excel 共用）
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="bdc3c7")
)


def create_page(parent, app):
    """工厂函数：构建并返回填充数据页面 Frame"""
    builder = MergeDataBuilder(parent, app)
    return builder.build()


class MergeDataBuilder:
    """填充数据页面构建 & 业务逻辑"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app       # ExcelToolApp 实例，提供 log()、配色等
        self.cols_a = []     # A 文件列名（选择文件后自动加载）
        self.cols_b = []     # B 文件列名（选择文件后自动加载）

    # --------------- UI 组件引用 ---------------
    # 这些会在 build() 中创建
    entry_file_a = None
    entry_file_b = None
    entry_match_a = None
    entry_match_b = None
    entry_fill_a = None
    entry_fill_b = None
    entry_out = None
    btn_start = None
    log_text = None

    # --------------- 构建 UI ---------------
    def build(self):
        bg = self.app.COLOR_CONTENT_BG

        frame = tk.Frame(self.parent, bg=bg)

        # 页面标题
        tk.Label(
            frame, text="★ Excel 内容填充（填充数据）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        # 细分隔线
        tk.Frame(frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(frame, text=" 文件选择 ", style="Card.TLabelframe",
            padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        self._file_row(file_frame, "A 文件（需填充的表）", "entry_file_a", self._browse_file_a)
        self._file_row(file_frame, "B 文件（数据源表）", "entry_file_b", self._browse_file_b)

        # ---- 列配置 ----
        col_frame = ttk.LabelFrame(frame, text=" 列匹配配置 ", style="Card.TLabelframe",
            padding=(16, 10))
        col_frame.pack(fill=tk.X, pady=(0, 12))

        self._pair_row(col_frame, "A 匹配列", "entry_match_a", "B 匹配列", "entry_match_b",
                       picker_cmd1=lambda: self._show_column_picker(self.entry_match_a, self.cols_a),
                       picker_cmd2=lambda: self._show_column_picker(self.entry_match_b, self.cols_b))
        self._pair_row(col_frame, "A 填充列", "entry_fill_a", "B 内容列",   "entry_fill_b",
                       picker_cmd1=lambda: self._show_column_picker(self.entry_fill_a, self.cols_a),
                       picker_cmd2=lambda: self._show_column_picker(self.entry_fill_b, self.cols_b))

        tk.Label(col_frame,
            text="提示：列名需与 Excel 表头（第一行）完全一致，支持多个列名用英文逗号分隔",
            font=("Microsoft YaHei", 9), bg=bg, fg="#95a5a6"
        ).pack(anchor=tk.W, pady=(6, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(frame, text=" 输出设置 ", style="Card.TLabelframe",
            padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(frame, bg=bg)
        btn_frame.pack(pady=16)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始转换",
            style="Action.TButton", width=22,
            command=self._start_merge)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(frame, text=" 处理日志 ",
            font=("Microsoft YaHei", 11, "bold"), bg=bg, fg="#2c3e50",
            padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        return frame

    # --------------- UI 辅助 ---------------
    def _file_row(self, parent, label, attr, cmd):
        row = tk.Frame(parent, bg=self.app.COLOR_CONTENT_BG)
        row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row, text=label, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=22, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        entry = ttk.Entry(row, style="Readonly.TEntry", state="readonly")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="📂 浏览", style="Secondary.TButton", command=cmd).pack(side=tk.LEFT)
        setattr(self, attr, entry)

    def _pair_row(self, parent, label1, attr1, label2, attr2,
                  picker_cmd1=None, picker_cmd2=None):
        row = tk.Frame(parent, bg=self.app.COLOR_CONTENT_BG)
        row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row, text=label1, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 6))
        e1 = ttk.Entry(row, style="Readonly.TEntry", state="readonly", width=16)
        e1.pack(side=tk.LEFT, padx=(0, 2))
        if picker_cmd1:
            ttk.Button(row, text="🏷️ 选列", style="Secondary.TButton", command=picker_cmd1).pack(side=tk.LEFT, padx=(0, 12))
        else:
            tk.Frame(row, width=6, bg=self.app.COLOR_CONTENT_BG).pack(side=tk.LEFT, padx=(0, 12))
        tk.Label(row, text=label2, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 6))
        e2 = ttk.Entry(row, style="Readonly.TEntry", state="readonly", width=16)
        e2.pack(side=tk.LEFT, padx=(0, 2))
        if picker_cmd2:
            ttk.Button(row, text="🏷️ 选列", style="Secondary.TButton", command=picker_cmd2).pack(side=tk.LEFT)
        setattr(self, attr1, e1)
        setattr(self, attr2, e2)

    def _set_entry_path(self, entry, path):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, path)
        entry.configure(state="readonly")

    # --------------- 浏览对话框 ---------------
    def _browse_file_a(self):
        path = filedialog.askopenfilename(
            title="选择 A 文件（需填充的表）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_file_a, path)
            self._load_columns(path, 'a')
            self.app.log(f"已选择 A 文件: {path}")

    def _browse_file_b(self):
        path = filedialog.askopenfilename(
            title="选择 B 文件（数据源表）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_file_b, path)
            self._load_columns(path, 'b')
            self.app.log(f"已选择 B 文件: {path}")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")],
            initialfile="已填充.xlsx"
        )
        if path:
            self._set_entry_path(self.entry_out, path)
            self.app.log(f"已设置输出路径: {path}")

    def _load_columns(self, filepath, which):
        """读取文件第一行列名，存入 self.cols_a 或 self.cols_b"""
        try:
            wb = load_workbook(filepath, read_only=True, data_only=True)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            cols = [str(c) for c in row if c is not None]
            if which == 'a':
                self.cols_a = cols
            else:
                self.cols_b = cols
            self._log(f"[INFO] 已加载 {which.upper()} 文件列名，共 {len(cols)} 列")
        except Exception as e:
            self._log(f"[ERROR] 读取 {which.upper()} 文件列名失败: {e}")

    def _show_column_picker(self, entry, cols):
        """弹出列名选择窗口，选中后自动将列名（逗号分隔）填入 entry"""
        if not cols:
            messagebox.showwarning("提示", "请先选择对应的文件以加载列名")
            return

        existing = [c.strip() for c in entry.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择列名")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()

        # 居中
        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.EXTENDED,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in cols:
            lb.insert(tk.END, col)

        # 预选中已有列名，跟踪选择顺序
        sel_order = []
        for i, col in enumerate(cols):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        # 绑定事件跟踪点击顺序（单选/多选）
        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            # 移除被取消的项
            sel_order[:] = [i for i in sel_order if i not in removed]
            # 新增的项追加到末尾，保留点击顺序
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            entry.delete(0, tk.END)
            entry.insert(0, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10),
                  width=8, bg="#3498db", fg="white", relief="flat",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

        popup.wait_window()

    # --------------- 日志（委托给主 app，写入本页日志控件） ---------------
    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        # 同时写入主 app 日志（如果有的话）
        self.app.log(message)

    # --------------- 核心处理 ---------------
    def _start_merge(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        file_a = self.entry_file_a.get().strip()
        file_b = self.entry_file_b.get().strip()
        out_path = self.entry_out.get().strip()
        match_a = self.entry_match_a.get().strip()
        match_b = self.entry_match_b.get().strip()
        fill_a = self.entry_fill_a.get().strip()
        fill_b = self.entry_fill_b.get().strip()

        if not all([file_a, file_b, out_path, match_a, match_b, fill_a, fill_b]):
            messagebox.showwarning("参数缺失", "请填写所有必填项！")
            return
        if not os.path.isfile(file_a):
            messagebox.showerror("文件错误", f"A 文件不存在:\n{file_a}"); return
        if not os.path.isfile(file_b):
            messagebox.showerror("文件错误", f"B 文件不存在:\n{file_b}"); return

        self.btn_start.configure(state="disabled", text="处理中...")
        t = threading.Thread(
            target=self._do_merge,
            args=(file_a, file_b, match_a, match_b, fill_a, fill_b, out_path),
            daemon=True)
        t.start()

    def _do_merge(self, file_a, file_b, match_a, match_b, fill_a, fill_b, out_path):
        try:
            self._log("=" * 50)
            self._log("开始处理 Excel 内容填充")

            match_a_cols = [c.strip() for c in match_a.split(",") if c.strip()]
            match_b_cols = [c.strip() for c in match_b.split(",") if c.strip()]
            fill_a_cols  = [c.strip() for c in fill_a.split(",") if c.strip()]
            fill_b_cols  = [c.strip() for c in fill_b.split(",") if c.strip()]

            if len(match_a_cols) != len(match_b_cols):
                raise ValueError(f"匹配列数量不一致：A 有 {len(match_a_cols)} 列，B 有 {len(match_b_cols)} 列")
            if len(fill_a_cols) != len(fill_b_cols):
                raise ValueError(f"填充列数量不一致：A 有 {len(fill_a_cols)} 列，B 有 {len(fill_b_cols)} 列")

            self._log(f"A 匹配列: {match_a_cols}")
            self._log(f"B 匹配列: {match_b_cols}")
            self._log(f"A 填充列: {fill_a_cols}")
            self._log(f"B 内容列: {fill_b_cols}")

            # 加载工作簿
            self._log("加载 A 文件...")
            wb_a = load_workbook(file_a)
            ws_a = wb_a.active
            self._log(f"A 文件: {ws_a.title}，{ws_a.max_row} 行 × {ws_a.max_column} 列")

            self._log("加载 B 文件...")
            wb_b = load_workbook(file_b)
            ws_b = wb_b.active
            self._log(f"B 文件: {ws_b.title}，{ws_b.max_row} 行 × {ws_b.max_column} 列")

            # 表头映射
            headers_a = {cell.value: idx for idx, cell in enumerate(ws_a[1], start=1) if cell.value}
            headers_b = {cell.value: idx for idx, cell in enumerate(ws_b[1], start=1) if cell.value}

            for c in match_a_cols + fill_a_cols:
                if c not in headers_a:
                    raise ValueError(f"A 文件中不存在列名: '{c}'")
            for c in match_b_cols + fill_b_cols:
                if c not in headers_b:
                    raise ValueError(f"B 文件中不存在列名: '{c}'")

            # B 文件索引（支持重复 key → 存为列表）
            self._log("构建 B 文件查找索引...")
            b_lookup = {}
            dup_count = 0
            for row_b in ws_b.iter_rows(min_row=2, values_only=False):
                key = tuple(str(row_b[headers_b[c] - 1].value or "").strip() for c in match_b_cols)
                if key in b_lookup:
                    b_lookup[key].append(row_b)
                    dup_count += 1
                else:
                    b_lookup[key] = [row_b]
            if dup_count > 0:
                self._log(f"索引完成，共 {len(b_lookup)} 条唯一 key（含 {dup_count} 条重复记录）")
            else:
                self._log(f"索引完成，共 {len(b_lookup)} 条唯一记录")

            # 遍历填充
            self._log("开始遍历 A 文件并填充...")
            fill_count = skip_count = no_match_count = 0
            merged_count = 0  # 因重复 key 被拼接的次数

            for row_a in ws_a.iter_rows(min_row=2, values_only=False):
                key = tuple(str(row_a[headers_a[c] - 1].value or "").strip() for c in match_a_cols)
                rows_b = b_lookup.get(key)
                if rows_b is None:
                    no_match_count += 1
                    continue
                for f_a, f_b in zip(fill_a_cols, fill_b_cols):
                    cell_a = row_a[headers_a[f_a] - 1]
                    if cell_a.value is None or str(cell_a.value).strip() == "":
                        # 从所有匹配的 B 行中收集非空值并去重
                        raw_values = []
                        for row_b in rows_b:
                            v = row_b[headers_b[f_b] - 1].value
                            if v is not None and str(v).strip() != "":
                                raw_values.append(v)
                        if not raw_values:
                            continue
                        # 去重（保持首次出现顺序）
                        seen = set()
                        unique_vals = []
                        for v in raw_values:
                            vid = str(v).strip()
                            if vid not in seen:
                                seen.add(vid)
                                unique_vals.append(v)
                        if len(unique_vals) == 1:
                            cell_a.value = unique_vals[0]
                        else:
                            cell_a.value = "; ".join(str(v) for v in unique_vals)
                            merged_count += 1
                        fill_count += 1
                    else:
                        skip_count += 1

            if merged_count > 0:
                self._log(f"填充 {fill_count} 个（其中 {merged_count} 个因重复 key 被拼接），跳过 {skip_count} 个（非空），未匹配 {no_match_count} 行")
            else:
                self._log(f"填充 {fill_count} 个，跳过 {skip_count} 个（非空），未匹配 {no_match_count} 行")

            # 保存
            wb_a.save(out_path)
            self._log(f"结果已保存: {out_path}")
            self._log("=" * 50)
            self._log("全部完成！")

            messagebox.showinfo("处理完成", f"Excel 内容填充已完成！\n\n结果已保存至:\n{out_path}")

        except Exception as e:
            self._log(f"[错误] {e}")
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(state="normal", text="⚡ 开始转换"))

    @staticmethod
    def _style_header(ws):
        """美化 sheet 第一行表头：深色底 + 白色粗体 + 居中 + 底边线"""
        for cell in ws[1]:
            if cell.value is None:
                continue
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER
        # 适当加高表头行
        ws.row_dimensions[1].height = 24
