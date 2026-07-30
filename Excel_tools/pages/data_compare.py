# -*- coding: utf-8 -*-
"""
Excel 数据对比页面模块

功能：输入两个 Excel 和关联列，输出一个包含三个 Sheet 的新 Excel：
  Sheet 1 — 合并结果：两表匹配行的列合并（匹配列去重）
  Sheet 2 — 仅 A 有：A 中存在但 B 中不存在的行
  Sheet 3 — 仅 B 有：B 中存在但 A 中不存在的行
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 表头样式
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(
    bottom=Side(style="thin", color="bdc3c7")
)


def create_page(parent, app):
    return DataComparePage(parent, app).frame


class DataComparePage:
    """数据对比页面 构建 & 业务逻辑"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self._a_path = None
        self._b_path = None
        self._a_headers = []   # A 文件列名（选文件后自动加载）
        self._b_headers = []   # B 文件列名（选文件后自动加载）

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== 构建 UI ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # 标题
        tk.Label(self.frame, text="数据对比（交集差集对比）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
            padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        self._file_row(file_frame, "A 文件（Excel A）", "entry_file_a", self._browse_file_a)
        self._file_row(file_frame, "B 文件（Excel B）", "entry_file_b", self._browse_file_b)

        # ---- 关联列配置 ----
        col_frame = ttk.LabelFrame(self.frame, text=" 关联列配置（点击 📋 选列，Ctrl+点击多选）",
                                     style="Card.TLabelframe", padding=(16, 10))
        col_frame.pack(fill=tk.X, pady=(0, 12))

        # A 关联列
        row_a = tk.Frame(col_frame, bg=bg)
        row_a.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row_a, text="A 关联列", font=("Microsoft YaHei", 10),
                 bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_match_a = ttk.Entry(row_a, style="Normal.TEntry", width=28)
        self.entry_match_a.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_a, text="📋", width=3,
            command=lambda: self._show_column_picker(self.entry_match_a, self._a_headers)
        ).pack(side=tk.LEFT, padx=(0, 20))
        self.lbl_a_info = tk.Label(row_a, text="请先选择文件", font=("Microsoft YaHei", 8),
                                    bg=bg, fg="#95a5a6")
        self.lbl_a_info.pack(side=tk.LEFT)

        # B 关联列
        row_b = tk.Frame(col_frame, bg=bg)
        row_b.pack(fill=tk.X)
        tk.Label(row_b, text="B 关联列", font=("Microsoft YaHei", 10),
                 bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_match_b = ttk.Entry(row_b, style="Normal.TEntry", width=28)
        self.entry_match_b.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(row_b, text="📋", width=3,
            command=lambda: self._show_column_picker(self.entry_match_b, self._b_headers)
        ).pack(side=tk.LEFT, padx=(0, 20))
        self.lbl_b_info = tk.Label(row_b, text="请先选择文件", font=("Microsoft YaHei", 8),
                                    bg=bg, fg="#95a5a6")
        self.lbl_b_info.pack(side=tk.LEFT)

        # 底部提示
        tk.Label(col_frame,
            text="提示：点 📋 按钮挑选列，Ctrl+点击多选，按选择顺序输出。两侧列数须一致（第1个↔第1个、第2个↔第2个…）。",
            font=("Microsoft YaHei", 9), bg=bg, fg="#95a5a6"
        ).pack(anchor=tk.W, pady=(8, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
            padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=16)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始对比",
            style="Action.TButton", width=22,
            command=self._start_compare)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
            font=("Microsoft YaHei", 11, "bold"), bg=bg, fg="#2c3e50",
            padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== UI 辅助方法 ====================
    def _file_row(self, parent, label, attr, cmd):
        row = tk.Frame(parent, bg=self.app.COLOR_CONTENT_BG)
        row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row, text=label, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=22, anchor=tk.E
            ).pack(side=tk.LEFT, padx=(0, 8))
        entry = ttk.Entry(row, style="Readonly.TEntry", state="readonly")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="📂 浏览", style="Secondary.TButton",
            command=cmd).pack(side=tk.LEFT)
        setattr(self, attr, entry)

    def _set_entry_text(self, entry, text):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, text)
        entry.configure(state="readonly")

    # ==================== 弹窗选列 ====================
    def _show_column_picker(self, entry, cols):
        """弹出列名选择窗口，选中后自动将列名（逗号分隔）填入 entry"""
        if not cols:
            messagebox.showwarning("提示", "请先选择对应的文件以加载列名")
            return

        existing = [c.strip() for c in entry.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择关联列")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()

        # 居中
        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.EXTENDED,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in cols:
            lb.insert(tk.END, col)

        # 预选中已有列名，跟踪选择顺序
        sel_order = []
        for i, col in enumerate(cols):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        # 绑定事件跟踪点击顺序
        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            sel_order[:] = [i for i in sel_order if i not in removed]
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            entry.delete(0, tk.END)
            entry.insert(0, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10),
                  width=8, bg="#3498db", fg="white", relief="flat",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

        popup.wait_window()

    # ==================== 浏览对话框 & 加载列名 ====================
    def _browse_file_a(self):
        path = filedialog.askopenfilename(
            title="选择 A 文件（Excel A）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._a_path = path
            self._set_entry_text(self.entry_file_a, path)
            self._log(f"已选择 A 文件: {os.path.basename(path)}")
            self._load_columns(path, "A")

    def _browse_file_b(self):
        path = filedialog.askopenfilename(
            title="选择 B 文件（Excel B）",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._b_path = path
            self._set_entry_text(self.entry_file_b, path)
            self._log(f"已选择 B 文件: {os.path.basename(path)}")
            self._load_columns(path, "B")

    def _load_columns(self, path, tag):
        """读取文件第一行列名并存储"""
        try:
            wb = load_workbook(path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                headers = [str(c) for c in row if c is not None]
                if tag == "A":
                    self._a_headers = headers
                    self.lbl_a_info.config(text=f"共 {len(headers)} 列", fg="#27ae60")
                    self.entry_match_a.delete(0, tk.END)
                else:
                    self._b_headers = headers
                    self.lbl_b_info.config(text=f"共 {len(headers)} 列", fg="#27ae60")
                    self.entry_match_b.delete(0, tk.END)
                self._log(f"{tag} 文件已加载 {len(headers)} 列")
            else:
                self._log(f"{tag} 文件未读取到列名")
        except Exception as e:
            self._log(f"[ERROR] 加载 {tag} 文件列名失败: {e}")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="对比结果.xlsx"
        )
        if path:
            self._set_entry_text(self.entry_out, path)
            self._log(f"已设置输出路径: {path}")

    # ==================== 日志 ====================
    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.app.log(message)

    # ==================== 启动处理 ====================
    def _start_compare(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        file_a = self._a_path or ""
        file_b = self._b_path or ""
        out_path = self.entry_out.get().strip()

        # 从 Entry 中解析列名，保留用户选择顺序
        sel_a = [c.strip() for c in self.entry_match_a.get().strip().split(",") if c.strip()]
        sel_b = [c.strip() for c in self.entry_match_b.get().strip().split(",") if c.strip()]

        if not file_a:
            messagebox.showwarning("参数缺失", "请先选择 A 文件！"); return
        if not file_b:
            messagebox.showwarning("参数缺失", "请先选择 B 文件！"); return
        if not out_path:
            messagebox.showwarning("参数缺失", "请选择输出路径！"); return
        if not sel_a:
            messagebox.showwarning("参数缺失", "请点击 📋 选择 A 文件的关联列！"); return
        if not sel_b:
            messagebox.showwarning("参数缺失", "请点击 📋 选择 B 文件的关联列！"); return
        if len(sel_a) != len(sel_b):
            messagebox.showwarning("列数不一致",
                f"两侧关联列数量不一致：A 选了 {len(sel_a)} 列，B 选了 {len(sel_b)} 列\n请调整使两侧数量一致。")
            return
        if not os.path.isfile(file_a):
            messagebox.showerror("文件错误", f"A 文件不存在:\n{file_a}"); return
        if not os.path.isfile(file_b):
            messagebox.showerror("文件错误", f"B 文件不存在:\n{file_b}"); return

        match_a = ",".join(sel_a)
        match_b = ",".join(sel_b)

        self.btn_start.configure(state="disabled", text="处理中...")
        t = threading.Thread(
            target=self._do_compare,
            args=(file_a, file_b, match_a, match_b, out_path),
            daemon=True)
        t.start()

    # ==================== 核心处理逻辑 ====================
    def _do_compare(self, file_a, file_b, match_a, match_b, out_path):
        try:
            self._log("=" * 50)
            self._log("开始 Excel 数据对比")

            match_a_cols = [c.strip() for c in match_a.split(",") if c.strip()]
            match_b_cols = [c.strip() for c in match_b.split(",") if c.strip()]

            if len(match_a_cols) != len(match_b_cols):
                raise ValueError(f"关联列数量不一致：A 有 {len(match_a_cols)} 列，B 有 {len(match_b_cols)} 列")

            self._log(f"A 关联列: {match_a_cols}")
            self._log(f"B 关联列: {match_b_cols}")

            # 加载工作簿
            self._log("加载 A 文件...")
            wb_a = load_workbook(file_a)
            ws_a = wb_a.active
            self._log(f"A 文件: {ws_a.title}，{ws_a.max_row} 行 × {ws_a.max_column} 列")

            self._log("加载 B 文件...")
            wb_b = load_workbook(file_b)
            ws_b = wb_b.active
            self._log(f"B 文件: {ws_b.title}，{ws_b.max_row} 行 × {ws_b.max_column} 列")

            # 读取表头及列索引
            headers_a, col_idx_a = self._read_headers(ws_a)
            headers_b, col_idx_b = self._read_headers(ws_b)
            self._log(f"A 表头: {list(headers_a)}")
            self._log(f"B 表头: {list(headers_b)}")

            # 校验关联列存在
            for c in match_a_cols:
                if c not in col_idx_a:
                    raise ValueError(f"A 文件中不存在列名: '{c}'")
            for c in match_b_cols:
                if c not in col_idx_b:
                    raise ValueError(f"B 文件中不存在列名: '{c}'")

            # 读取 A 所有数据行
            self._log("读取 A 文件数据...")
            rows_a = []
            for row in ws_a.iter_rows(min_row=2, values_only=True):
                rows_a.append(row)
            self._log(f"A 文件数据行数: {len(rows_a)}")

            # 读取 B 所有数据行 & 构建索引
            self._log("读取 B 文件数据 & 构建索引...")
            rows_b = []
            b_index = {}
            for idx, row in enumerate(ws_b.iter_rows(min_row=2, values_only=True)):
                rows_b.append(row)
                key = self._make_key(row, match_b_cols, col_idx_b)
                if key:
                    b_index.setdefault(key, []).append(idx)
            self._log(f"B 文件数据行数: {len(rows_b)}，唯一键数: {len(b_index)}")

            # 分类
            matched_a_indices = []
            matched_b_indices = set()
            unmatched_a_indices = []

            for a_idx, row_a in enumerate(rows_a):
                key = self._make_key(row_a, match_a_cols, col_idx_a)
                b_matches = b_index.get(key, [])
                if b_matches:
                    b_idx = b_matches[0]
                    matched_a_indices.append((a_idx, b_idx))
                    matched_b_indices.add(b_idx)
                else:
                    unmatched_a_indices.append(a_idx)

            unmatched_b_indices = [i for i in range(len(rows_b)) if i not in matched_b_indices]

            self._log(f"匹配行数: {len(matched_a_indices)}")
            self._log(f"仅 A 有: {len(unmatched_a_indices)} 行")
            self._log(f"仅 B 有: {len(unmatched_b_indices)} 行")

            # 构建输出工作簿
            out_wb = Workbook()

            # Sheet 1: 合并结果
            ws1 = out_wb.active
            ws1.title = "合并结果"

            extra_b_headers = [h for h in headers_b if h not in headers_a]
            extra_b_indices = [col_idx_b[h] for h in extra_b_headers]

            merged_headers = list(headers_a) + extra_b_headers
            self._log(f"合并结果表头: {merged_headers}")
            ws1.append(merged_headers)

            for a_idx, b_idx in matched_a_indices:
                row_a = rows_a[a_idx]
                row_b = rows_b[b_idx]
                merged_row = list(row_a) + [row_b[bi] for bi in extra_b_indices]
                ws1.append(merged_row)

            # Sheet 2: 仅 A 有
            ws2 = out_wb.create_sheet("仅A有")
            ws2.append(list(headers_a))
            for a_idx in unmatched_a_indices:
                ws2.append(list(rows_a[a_idx]))

            # Sheet 3: 仅 B 有
            ws3 = out_wb.create_sheet("仅B有")
            ws3.append(list(headers_b))
            for b_idx in unmatched_b_indices:
                ws3.append(list(rows_b[b_idx]))

            # 美化表头 + 自动调整列宽
            for ws in [ws1, ws2, ws3]:
                self._style_header(ws)
                self._auto_col_width(ws)

            # 保存
            out_wb.save(out_path)
            self._log(f"结果已保存: {out_path}")
            self._log("=" * 50)
            self._log("全部完成！")

            summary = (
                f"数据对比完成！\n\n"
                f"匹配行数: {len(matched_a_indices)}\n"
                f"仅 A 有: {len(unmatched_a_indices)} 行\n"
                f"仅 B 有: {len(unmatched_b_indices)} 行\n\n"
                f"已保存至:\n{out_path}"
            )
            messagebox.showinfo("对比完成", summary)

        except Exception as e:
            self._log(f"[错误] {e}")
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(
                state="normal", text="⚡ 开始对比"))

    # ==================== 工具方法 ====================
    @staticmethod
    def _read_headers(ws):
        headers = []
        col_idx = {}
        for cell in ws[1]:
            if cell.value is not None:
                name = str(cell.value).strip()
                headers.append(name)
                col_idx[name] = cell.column - 1
        return headers, col_idx

    @staticmethod
    def _make_key(row, match_cols, col_idx):
        parts = []
        for c in match_cols:
            val = row[col_idx[c]] if col_idx[c] < len(row) else ""
            parts.append(str(val or "").strip())
        return tuple(parts)

    @staticmethod
    def _auto_col_width(ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = col_cells[0].column_letter
            for cell in col_cells:
                if cell.value:
                    val = str(cell.value)
                    length = sum(2 if ord(ch) > 127 else 1 for ch in val)
                    if length > max_len:
                        max_len = length
            ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    @staticmethod
    def _style_header(ws):
        for cell in ws[1]:
            if cell.value is None:
                continue
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = HEADER_ALIGNMENT
            cell.border = HEADER_BORDER
        ws.row_dimensions[1].height = 24
