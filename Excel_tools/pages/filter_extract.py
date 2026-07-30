# -*- coding: utf-8 -*-
"""
条件筛选提取页面模块

功能：按自定义多条件（AND/OR 组合）从 Excel 中筛选数据，输出指定列
  - 支持 =, ≠, >, <, >=, <=, 包含, 不包含, 为空, 不为空 多种运算符
  - 动态添加/删除条件行，每行可选择 AND/OR 连接方式
  - 可选择输出特定列
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# 运算符列表
OPERATORS = ["=", "≠", ">", "<", ">=", "<=", "包含", "不包含", "为空", "不为空"]
# 不需要输入值的运算符
NO_VALUE_OPS = {"为空", "不为空"}
# 连接符
CONNECTORS = ["AND", "OR"]


def create_page(parent, app):
    return FilterExtractPage(parent, app).frame


class FilterExtractPage:
    """条件筛选提取页面"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self._file_path = None
        self._headers = []
        self._condition_rows = []  # [(cond_frame, conn_var, col_var, op_var, val_entry), ...]
        self._output_col_vars = []  # [(col_name, tk.BooleanVar), ...]

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== 构建 UI ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # 标题（固定）
        tk.Label(self.frame, text="条件筛选提取（按条件过滤数据）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 处理日志（先 pack，固定底部） ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
            font=("Microsoft YaHei", 11, "bold"), bg=bg, fg="#2c3e50",
            padx=10, pady=10, relief="solid", bd=1, height=180)
        log_frame.pack(fill=tk.X, side=tk.BOTTOM, pady=(5, 0))
        log_frame.pack_propagate(False)  # 固定高度不被内部控件挤压

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        # ---- 可滚动内容区（后 pack，填充剩余空间） ----
        self._scroll_canvas = tk.Canvas(self.frame, bg=bg, highlightthickness=0)
        scrollbar = ttk.Scrollbar(self.frame, orient=tk.VERTICAL, command=self._scroll_canvas.yview)
        self._scroll_canvas.configure(yscrollcommand=scrollbar.set)

        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self._scroll_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self._content_frame = tk.Frame(self._scroll_canvas, bg=bg)
        self._content_id = self._scroll_canvas.create_window((0, 0),
            window=self._content_frame, anchor=tk.NW)

        def _on_canvas_configure(event):
            self._scroll_canvas.itemconfig(self._content_id, width=event.width)

        self._scroll_canvas.bind("<Configure>", _on_canvas_configure)

        def _on_frame_configure(event):
            self._scroll_canvas.configure(scrollregion=self._scroll_canvas.bbox("all"))

        self._content_frame.bind("<Configure>", _on_frame_configure)

        # 鼠标滚轮支持
        def _on_mousewheel(event):
            self._scroll_canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        self._scroll_canvas.bind("<Enter>", lambda e: self._scroll_canvas.bind_all("<MouseWheel>", _on_mousewheel))
        self._scroll_canvas.bind("<Leave>", lambda e: self._scroll_canvas.unbind_all("<MouseWheel>"))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self._content_frame, text=" 文件选择 ",
            style="Card.TLabelframe", padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        row_file = tk.Frame(file_frame, bg=bg)
        row_file.pack(fill=tk.X)
        tk.Label(row_file, text="待筛选文件", font=("Microsoft YaHei", 10),
            bg=bg, width=12, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_file = ttk.Entry(row_file, style="Readonly.TEntry", state="readonly")
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_file, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_file).pack(side=tk.LEFT)

        self.lbl_file_info = tk.Label(row_file, text="", font=("Microsoft YaHei", 8),
            bg=bg, fg="#95a5a6", width=18, anchor=tk.W)
        self.lbl_file_info.pack(side=tk.LEFT, padx=(10, 0))

        # ---- 筛选条件配置 ----
        cond_frame = ttk.LabelFrame(self._content_frame, text=" 筛选条件（AND/OR 可切换）",
            style="Card.TLabelframe", padding=(16, 10))
        cond_frame.pack(fill=tk.X, pady=(0, 12))

        self.cond_container = tk.Frame(cond_frame, bg=bg)
        self.cond_container.pack(fill=tk.X)

        # 提示和添加按钮
        btn_row = tk.Frame(cond_frame, bg=bg)
        btn_row.pack(fill=tk.X, pady=(8, 0))
        ttk.Button(btn_row, text="＋ 添加条件", style="Secondary.TButton",
            command=self._add_condition_row).pack(side=tk.LEFT)

        tk.Label(btn_row,
            text="提示：每个 OR 开启一组新条件，组内 AND（交集），组间 OR（并集）",
            font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6"
        ).pack(side=tk.LEFT, padx=(16, 0))

        self._cond_hint = tk.Label(self.cond_container, text="请先选择文件，加载列名后可添加筛选条件",
            font=("Microsoft YaHei", 9), bg=bg, fg="#95a5a6")
        self._cond_hint.pack(pady=(8, 4))

        # ---- 输出列选择 ----
        out_col_frame = ttk.LabelFrame(self._content_frame, text=" 输出列（勾选要输出的列，默认全部输出）",
            style="Card.TLabelframe", padding=(16, 10))
        out_col_frame.pack(fill=tk.X, pady=(0, 12))

        # 全选/全不选按钮
        col_ctrl = tk.Frame(out_col_frame, bg=bg)
        col_ctrl.pack(fill=tk.X, pady=(0, 6))
        ttk.Button(col_ctrl, text="全选", style="Secondary.TButton",
            command=self._select_all_cols, width=8).pack(side=tk.LEFT, padx=(0, 6))
        ttk.Button(col_ctrl, text="全不选", style="Secondary.TButton",
            command=self._deselect_all_cols, width=8).pack(side=tk.LEFT)

        self.out_col_hint = tk.Label(out_col_frame, text="请先选择文件加载列名",
            font=("Microsoft YaHei", 9), bg=bg, fg="#95a5a6")
        self.out_col_hint.pack(anchor=tk.W, pady=(4, 0))

        # 列复选框容器
        self.out_col_container = tk.Frame(out_col_frame, bg=bg)
        self.out_col_container.pack(fill=tk.X)

        # ---- 输出文件设置 ----
        out_frame = ttk.LabelFrame(self._content_frame, text=" 输出文件 ",
            style="Card.TLabelframe", padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出路径", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self._content_frame, bg=bg)
        btn_frame.pack(pady=(16, 8))
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始筛选",
            style="Action.TButton", width=22,
            command=self._start)
        self.btn_start.pack()

    # ==================== 条件行管理 ====================
    def _add_condition_row(self):
        """动态添加一行筛选条件"""
        bg = self.app.COLOR_CONTENT_BG

        if not self._headers:
            return

        # 添加条件后隐藏提示
        self._cond_hint.pack_forget()

        row_idx = len(self._condition_rows)
        row_frame = tk.Frame(self.cond_container, bg=bg)
        row_frame.pack(fill=tk.X, pady=(0, 6))

        # AND/OR 切换（第一行不显示，仅占位）
        if row_idx == 0:
            conn_var = tk.StringVar(value="")
            tk.Label(row_frame, text="", width=5, bg=bg).pack(side=tk.LEFT, padx=(0, 6))
        else:
            conn_var = tk.StringVar(value="AND")
            conn_cb = ttk.Combobox(row_frame, textvariable=conn_var,
                values=CONNECTORS, state="readonly", width=5,
                font=("Microsoft YaHei", 9, "bold"),
                style="CondConn.TCombobox")
            conn_cb.pack(side=tk.LEFT, padx=(0, 6))

        # 列名下拉
        col_var = tk.StringVar(value=self._headers[0])
        col_dd = ttk.Combobox(row_frame, textvariable=col_var,
            values=self._headers, state="readonly", width=14,
            font=("Microsoft YaHei", 10), style="CondCol.TCombobox")
        col_dd.pack(side=tk.LEFT, padx=(0, 8))

        # 运算符下拉
        op_var = tk.StringVar(value="=")
        op_dd = ttk.Combobox(row_frame, textvariable=op_var,
            values=OPERATORS, state="readonly", width=8,
            font=("Microsoft YaHei", 10), style="CondOp.TCombobox")
        op_dd.pack(side=tk.LEFT, padx=(0, 8))

        # 值输入框
        val_entry = ttk.Entry(row_frame, style="CondEntry.TEntry", width=22)
        val_entry.pack(side=tk.LEFT, padx=(0, 8))

        def _on_op_change(*args):
            if op_var.get() in NO_VALUE_OPS:
                val_entry.pack_forget()
                val_entry.delete(0, tk.END)
            else:
                children = row_frame.pack_slaves()
                del_btn = None
                for c in children:
                    if isinstance(c, ttk.Button):
                        del_btn = c
                        break
                if del_btn:
                    val_entry.pack(side=tk.LEFT, padx=(0, 8), before=del_btn)
                else:
                    val_entry.pack(side=tk.LEFT, padx=(0, 8))

        op_var.trace("w", _on_op_change)

        # 删除按钮
        del_btn = ttk.Button(row_frame, text="✕",
            style="Secondary.TButton", width=3,
            command=lambda rf=row_frame: self._remove_condition_row(rf))
        del_btn.pack(side=tk.LEFT)

        self._condition_rows.append((row_frame, conn_var, col_var, op_var, val_entry))

    def _remove_condition_row(self, row_frame):
        """删除一行条件，并刷新第一行的连接符显示"""
        self._condition_rows = [
            r for r in self._condition_rows if r[0] != row_frame
        ]
        row_frame.destroy()
        # 重建所有条件行的连接符（第一行可能已不是第一行了）
        self._refresh_connectors()
        # 所有条件都删完时恢复提示
        if not self._condition_rows:
            self._cond_hint.pack(pady=(8, 4))

    def _refresh_connectors(self):
        """刷新每行的 AND/OR 显示：第一行空白，其余行 AND/OR"""
        for idx, (rf, conn_var, _, _, _) in enumerate(self._condition_rows):
            # 找到并删除连接符控件（row_frame 第一个子控件）
            slaves = rf.pack_slaves()
            if slaves and isinstance(slaves[0], (tk.Label, ttk.Combobox)):
                txt = slaves[0].cget("text") if isinstance(slaves[0], tk.Label) else ""
                # 第一行的空 label 或 AND/OR combobox
                if txt in ("", "AND", "OR") or isinstance(slaves[0], ttk.Combobox):
                    slaves[0].destroy()

            # 剩下的第一个控件是列名下拉，新连接符插到它前面
            anchor = rf.pack_slaves()[0] if rf.pack_slaves() else None
            bg = self.app.COLOR_CONTENT_BG

            if idx == 0:
                conn_var.set("")
                lbl = tk.Label(rf, text="", width=5, bg=bg)
                if anchor:
                    lbl.pack(side=tk.LEFT, padx=(0, 6), before=anchor)
                else:
                    lbl.pack(side=tk.LEFT, padx=(0, 6))
            else:
                conn_var.set("AND")
                cb = ttk.Combobox(rf, textvariable=conn_var,
                    values=CONNECTORS, state="readonly", width=5,
                    font=("Microsoft YaHei", 9, "bold"),
                    style="CondConn.TCombobox")
                if anchor:
                    cb.pack(side=tk.LEFT, padx=(0, 6), before=anchor)
                else:
                    cb.pack(side=tk.LEFT, padx=(0, 6))

    # ==================== 输出列管理 ====================
    def _build_output_col_checkboxes(self):
        """根据加载的列名构建复选框"""
        # 清空旧内容
        for w in self.out_col_container.winfo_children():
            w.destroy()
        self._output_col_vars.clear()

        self.out_col_hint.pack_forget()

        bg = self.app.COLOR_CONTENT_BG
        # 显示提示行
        tk.Label(self.out_col_container, text="选择输出列（默认全选）：",
            font=("Microsoft YaHei", 8), bg=bg, fg="#7f8c8d"
        ).pack(anchor=tk.W, pady=(2, 4))

        # 每行放 2 个复选框，给长列名足够空间
        row_frame = None
        for i, h in enumerate(self._headers):
            if i % 2 == 0:
                row_frame = tk.Frame(self.out_col_container, bg=bg)
                row_frame.pack(fill=tk.X, pady=(0, 4))
            var = tk.BooleanVar(value=True)
            cb = tk.Checkbutton(row_frame, text=h, variable=var,
                font=("Microsoft YaHei", 9), bg=bg,
                selectcolor=bg, activebackground=bg,
                anchor=tk.W, justify=tk.LEFT,
                wraplength=360, width=46)
            cb.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
            self._output_col_vars.append((h, var))

    def _select_all_cols(self):
        for _, var in self._output_col_vars:
            var.set(True)

    def _deselect_all_cols(self):
        for _, var in self._output_col_vars:
            var.set(False)

    def _get_selected_cols(self):
        """返回选中的列名列表"""
        if not self._output_col_vars:
            return self._headers  # 还没加载时返回全部
        selected = [name for name, var in self._output_col_vars if var.get()]
        return selected if selected else self._headers

    # ==================== 浏览 & 加载 ====================
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择待筛选的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._file_path = path
            self._set_entry_text(self.entry_file, path)
            self._log(f"已选择文件: {os.path.basename(path)}")
            self._load_columns(path)

    def _load_columns(self, path):
        try:
            wb = load_workbook(path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                self._headers = [str(c) for c in row if c is not None]
                self._cond_hint.pack_forget()
                self.lbl_file_info.config(
                    text=f"共 {len(self._headers)} 列", fg="#27ae60")
                self._log(f"已加载 {len(self._headers)} 列")

                # 清空已有条件
                for rf, _, _, _, _ in self._condition_rows:
                    rf.destroy()
                self._condition_rows.clear()
                # 不自动创建条件行，显示提示
                self._cond_hint.config(
                    text="点击「＋ 添加条件」设置筛选规则；不添加条件则输出全部行")
                self._cond_hint.pack(pady=(8, 4))

                # 重建输出列复选框
                self._build_output_col_checkboxes()
        except Exception as e:
            self._log(f"[ERROR] 加载列名失败: {e}")

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="筛选结果.xlsx"
        )
        if path:
            self._set_entry_text(self.entry_out, path)
            self._log(f"已设置输出路径: {path}")

    def _set_entry_text(self, entry, text):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, text)
        entry.configure(state="readonly")

    # ==================== 日志 ====================
    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.app.log(message)

    # ==================== 启动 ====================
    def _start(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        file_path = self._file_path or ""
        out_path = self.entry_out.get().strip()

        if not file_path:
            messagebox.showwarning("参数缺失", "请选择待筛选的 Excel 文件！")
            return
        if not out_path:
            messagebox.showwarning("参数缺失", "请设置输出文件路径！")
            return
        if not os.path.isfile(file_path):
            messagebox.showerror("文件错误", f"文件不存在:\n{file_path}")
            return

        # 收集条件和连接符 → condition_groups: [[(col, op, val), ...], ...]
        condition_groups = []
        current_group = []
        for row_frame, conn_var, col_var, op_var, val_entry in self._condition_rows:
            col = col_var.get().strip()
            op = op_var.get().strip()
            val = val_entry.get().strip()
            if not col:
                continue
            if op not in NO_VALUE_OPS and not val:
                continue

            conn = conn_var.get().strip()
            if conn == "OR" and current_group:
                # OR 分隔，当前 group 结束，开始新 group
                condition_groups.append(current_group)
                current_group = []
            current_group.append((col, op, val))

        if current_group:
            condition_groups.append(current_group)

        if not condition_groups:
            condition_groups = []  # 空条件 = 输出全部行

        # 输出列
        selected_cols = self._get_selected_cols()

        self.btn_start.configure(state="disabled", text="处理中...")
        t = threading.Thread(
            target=self._do_filter,
            args=(file_path, condition_groups, selected_cols, out_path),
            daemon=True)
        t.start()

    # ==================== 核心逻辑 ====================
    def _do_filter(self, file_path, condition_groups, selected_cols, out_path):
        try:
            self._log("=" * 50)
            if condition_groups:
                self._log("开始条件筛选")
            else:
                self._log("开始提取（无筛选条件，输出全部行）")

            for i, group in enumerate(condition_groups):
                connector = "OR" if i > 0 else ""
                prefix = f"  {connector} " if connector else "  "
                for j, (col, op, val) in enumerate(group):
                    joiner = "AND" if j > 0 else ""
                    p2 = f"    {joiner} " if joiner else ""
                    self._log(f"{prefix}{p2}[{col}] {op} {val if op not in NO_VALUE_OPS else ''}")
                    prefix = ""  # 只第一行显示 OR

            self._log(f"输出列: {', '.join(selected_cols)}")

            # 加载文件
            self._log("加载文件...")
            wb = load_workbook(file_path)
            ws = wb.active
            total_rows = ws.max_row - 1
            self._log(f"文件: {ws.title}，{total_rows} 行数据")

            # 表头映射
            all_headers_raw = [cell.value for cell in ws[1]]
            headers = {cell.value: idx for idx, cell in enumerate(ws[1], start=1)
                       if cell.value}

            # 验证条件列和输出列
            for group in condition_groups:
                for col, op, val in group:
                    if col not in headers:
                        raise ValueError(f"文件中不存在列名: '{col}'")
            for col in selected_cols:
                if col not in headers:
                    raise ValueError(f"文件中不存在列名: '{col}'")

            # 输出列索引
            out_col_indexes = [headers[c] for c in selected_cols]

            # 逐行匹配
            self._log("逐行筛选...")
            matched_data = []
            # 表头行：只输出选中的列
            header_row = [all_headers_raw[idx - 1] for idx in out_col_indexes]
            matched_data.append(header_row)

            matched_count = 0
            for row in ws.iter_rows(min_row=2, values_only=False):
                if self._check_row_groups(row, headers, condition_groups):
                    matched_data.append([row[idx - 1].value for idx in out_col_indexes])
                    matched_count += 1

            wb.close()

            # 输出结果
            self._log(f"筛选完成: {matched_count} / {total_rows} 行匹配")

            if matched_count == 0:
                self._log("没有匹配的数据行")
                messagebox.showinfo("筛选完成",
                    f"没有匹配的数据行！\n"
                    f"原始数据: {total_rows} 行\n"
                    f"匹配数据: 0 行")
                self.app.root.after(0, lambda: self.btn_start.configure(
                    state="normal", text="⚡ 开始筛选"))
                return

            self._log("写入输出文件...")

            from openpyxl import Workbook
            wb_out = Workbook()
            ws_out = wb_out.active
            ws_out.title = "筛选结果"

            for row_data in matched_data:
                ws_out.append(row_data)

            wb_out.save(out_path)
            self._log(f"结果已保存: {out_path}")
            self._log("=" * 50)
            self._log("筛选完成！")

            messagebox.showinfo("筛选完成",
                f"条件筛选完成！\n\n"
                f"原始数据: {total_rows} 行\n"
                f"匹配数据: {matched_count} 行\n"
                f"筛选比例: {matched_count * 100 // total_rows}%\n"
                f"输出列数: {len(selected_cols)}\n\n"
                f"已保存至:\n{out_path}")

        except Exception as e:
            self._log(f"[错误] {e}")
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(
                state="normal", text="⚡ 开始筛选"))

    # ==================== 行匹配逻辑 ====================
    @staticmethod
    def _check_row_groups(row, headers, condition_groups):
        """AND/OR 分组匹配：组内 AND，组间 OR。空条件 = 全部匹配"""
        if not condition_groups:
            return True
        for group in condition_groups:
            if FilterExtractPage._check_group(row, headers, group):
                return True
        return False

    @staticmethod
    def _normalize(value):
        """标准化单元格值：datetime → 'YYYY-MM-DD'，其他 → str"""
        from datetime import datetime
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        return str(value).strip() if value is not None else ""

    @staticmethod
    def _try_date_compare(cell_str, op, val):
        """尝试将双方解析为日期比较；无法解析则返回 None"""
        from datetime import datetime
        formats = ["%Y-%m-%d", "%Y/%m/%d", "%Y%m%d",
                   "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]
        for fmt in formats:
            try:
                cell_dt = datetime.strptime(cell_str[:len(fmt.replace('%H:%M:%S', ''))], fmt.replace('%H:%M:%S', ''))
                val_dt = datetime.strptime(val[:len(fmt.replace('%H:%M:%S', ''))], fmt.replace('%H:%M:%S', ''))
                return cell_dt, val_dt
            except (ValueError, IndexError):
                continue
        return None

    @staticmethod
    def _check_group(row, headers, conditions):
        """组内所有条件都满足（AND）"""
        for col, op, val in conditions:
            cell_value = row[headers[col] - 1].value
            cell_str = FilterExtractPage._normalize(cell_value)

            # 日期比较优先
            date_pair = FilterExtractPage._try_date_compare(cell_str, op, val)
            if date_pair and op in (">", "<", ">=", "<=", "=", "≠"):
                cell_dt, val_dt = date_pair
                if op == "=":
                    if cell_dt != val_dt:
                        return False
                elif op == "≠":
                    if cell_dt == val_dt:
                        return False
                elif op == ">":
                    if cell_dt <= val_dt:
                        return False
                elif op == "<":
                    if cell_dt >= val_dt:
                        return False
                elif op == ">=":
                    if cell_dt < val_dt:
                        return False
                elif op == "<=":
                    if cell_dt > val_dt:
                        return False
                continue  # 日期比较完成，跳过后续

            if op == "=":
                if cell_str != val:
                    return False
            elif op == "≠":
                if cell_str == val:
                    return False
            elif op == ">":
                try:
                    if float(cell_str) <= float(val):
                        return False
                except ValueError:
                    if cell_str <= val:
                        return False
            elif op == "<":
                try:
                    if float(cell_str) >= float(val):
                        return False
                except ValueError:
                    if cell_str >= val:
                        return False
            elif op == ">=":
                try:
                    if float(cell_str) < float(val):
                        return False
                except ValueError:
                    if cell_str < val:
                        return False
            elif op == "<=":
                try:
                    if float(cell_str) > float(val):
                        return False
                except ValueError:
                    if cell_str > val:
                        return False
            elif op == "包含":
                if val not in cell_str:
                    return False
            elif op == "不包含":
                if val in cell_str:
                    return False
            elif op == "为空":
                if cell_str != "":
                    return False
            elif op == "不为空":
                if cell_str == "":
                    return False
        return True
