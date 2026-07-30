# -*- coding: utf-8 -*-
"""
数据合并（单表按主键去重合并 + 聚合统计）页面模块

功能：对一个 Excel 文件，按主键列分组，
      对数值列执行用户选择的聚合操作（求和/计数/平均值/最大/最小），输出新文件。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
from datetime import datetime
from collections import defaultdict

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    PatternFill = Font = Alignment = Border = Side = None

from .widgets import RoundedCheckbox

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)
HEADER_BORDER = Border(bottom=Side(style="thin", color="bdc3c7"))
CELL_ALIGNMENT = Alignment(horizontal="left", vertical="center")
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

AGG_OPTS = [
    ("求和",   "sum"),
    ("计数",   "count"),
    ("平均值", "avg"),
    ("最大值", "max"),
    ("最小值", "min"),
]


def create_page(parent, app):
    """工厂函数：构建并返回数据合并页面 Frame"""
    return _DataMergePage(parent, app).frame


class _DataMergePage:
    """数据合并页面构建 & 业务逻辑"""

    def __init__(self, parent, app):
        self.parent = parent
        self.app = app
        self.file_path = None     # 源文件路径
        self._key_headers = []   # 列名缓存

        self.frame = tk.Frame(parent, bg=app.COLOR_CONTENT_BG)
        self._build_ui()

    # ==================== UI 构建 ====================
    def _build_ui(self):
        bg = self.app.COLOR_CONTENT_BG

        # --- 页面标题 ---
        tk.Label(
            self.frame, text="★ 数据合并（单表按主键去重 + 聚合统计）",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 文件选择 ----
        file_frame = ttk.LabelFrame(self.frame, text=" 文件选择 ", style="Card.TLabelframe",
                                     padding=(16, 10))
        file_frame.pack(fill=tk.X, pady=(0, 12))

        row_file = tk.Frame(file_frame, bg=bg)
        row_file.pack(fill=tk.X)
        tk.Label(row_file, text="源 Excel 文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_file = ttk.Entry(row_file, style="Readonly.TEntry", state="readonly")
        self.entry_file.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_file, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_file).pack(side=tk.LEFT)

        # ---- 主键与聚合选择 ----
        config_frame = ttk.LabelFrame(self.frame, text=" 合并配置 ", style="Card.TLabelframe",
                                       padding=(16, 10))
        config_frame.pack(fill=tk.X, pady=(0, 12))

        # 第一行：主键列
        key_frame = tk.Frame(config_frame, bg=bg)
        key_frame.pack(fill=tk.X, pady=(0, 10))

        key_row = tk.Frame(key_frame, bg=bg)
        key_row.pack(fill=tk.X)
        tk.Label(key_row, text="主键列（可多选）", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_keys = ttk.Entry(key_row, style="Readonly.TEntry", state="readonly", width=28)
        self.entry_keys.pack(side=tk.LEFT, padx=(0, 4))
        ttk.Button(key_row, text="🏷️ 选列", style="Secondary.TButton",
                   command=self._show_key_picker).pack(side=tk.LEFT)
        tk.Label(key_frame, text="提示：选择文件后点击按钮选取主键列，支持 Ctrl 多选",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        # 第二行：聚合方式
        agg_frame = tk.Frame(config_frame, bg=bg)
        agg_frame.pack(fill=tk.X)

        # 标签 + 全选按钮 同行
        agg_title_row = tk.Frame(agg_frame, bg=bg)
        agg_title_row.pack(anchor=tk.W, fill=tk.X)
        tk.Label(agg_title_row, text="数值列聚合方式", font=("Microsoft YaHei", 10, "bold"),
                 bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        self.agg_vars = {}
        agg_content = tk.Frame(agg_frame, bg=bg)
        agg_content.pack(anchor=tk.W, pady=(4, 0))

        for i, (label, key) in enumerate(AGG_OPTS):
            var = tk.BooleanVar(value=(label == "求和"))
            self.agg_vars[label] = var
            cb_frame = tk.Frame(agg_content, bg=bg)
            cb_frame.grid(row=0, column=i, sticky=tk.W, padx=(0, 24), pady=2)
            RoundedCheckbox(cb_frame, variable=var, bg=bg).pack(side=tk.LEFT)
            tk.Label(cb_frame, text=" " + label, font=("Microsoft YaHei", 10),
                     bg=bg, fg="#2c3e50").pack(side=tk.LEFT)

        # 全选/全不选按钮（放在标题行右侧）
        def _toggle_all_agg():
            all_on = all(v.get() for v in self.agg_vars.values())
            for v in self.agg_vars.values():
                v.set(not all_on)

        def _update_agg_toggle_text(*args):
            all_on = all(v.get() for v in self.agg_vars.values())
            self.btn_agg_toggle.config(text="✗ 全不选" if all_on else "✓ 全选")

        self.btn_agg_toggle = tk.Button(agg_title_row, text="✓ 全选",
                                         font=("Microsoft YaHei", 9),
                                         bg="#3498db", fg="white", relief="flat",
                                         activebackground="#2980b9", activeforeground="white",
                                         cursor="hand2", padx=8, pady=2,
                                         command=_toggle_all_agg)
        self.btn_agg_toggle.pack(side=tk.RIGHT, padx=(8, 0))

        for v in self.agg_vars.values():
            v.trace_add("write", _update_agg_toggle_text)
        _update_agg_toggle_text()

        tk.Label(agg_frame, text="提示：相同主键的行，数值列按勾选方式汇总，非数值列取第一行",
                 font=("Microsoft YaHei", 8), bg=bg, fg="#95a5a6",
                 wraplength=700, justify=tk.LEFT, anchor=tk.W
                 ).pack(fill=tk.X, pady=(6, 0))

        # ---- 输出设置 ----
        out_frame = ttk.LabelFrame(self.frame, text=" 输出设置 ", style="Card.TLabelframe",
                                    padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        row_out = tk.Frame(out_frame, bg=bg)
        row_out.pack(fill=tk.X)
        tk.Label(row_out, text="输出文件", font=("Microsoft YaHei", 10),
                 bg=bg, width=14, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_out = ttk.Entry(row_out, style="Readonly.TEntry", state="readonly")
        self.entry_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row_out, text="📂 浏览", style="Secondary.TButton",
                   command=self._browse_out).pack(side=tk.LEFT)

        # ---- 开始按钮 ----
        btn_frame = tk.Frame(self.frame, bg=bg)
        btn_frame.pack(pady=14)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始合并",
                                     style="Action.TButton", width=22,
                                     command=self._start_merge)
        self.btn_start.pack()

        # ---- 处理日志 ----
        log_frame = tk.LabelFrame(self.frame, text=" 处理日志 ",
                                   font=("Microsoft YaHei", 11, "bold"),
                                   bg=bg, fg="#2c3e50",
                                   padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

    # ==================== 交互 ====================
    def _browse_file(self):
        path = filedialog.askopenfilename(
            title="选择要合并的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")]
        )
        if not path:
            return
        self.file_path = path
        self._set_readonly_value(self.entry_file, os.path.basename(path))
        self._load_columns()

    def _browse_out(self):
        path = filedialog.asksaveasfilename(
            title="设置输出路径",
            defaultextension=".xlsx",
            filetypes=[("Excel 文件", "*.xlsx"), ("所有文件", "*.*")],
            initialfile="合并结果.xlsx"
        )
        if path:
            self._set_readonly_value(self.entry_out, path)

    def _set_readonly_value(self, entry, value):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, value)
        entry.configure(state="readonly")

    def _load_columns(self):
        self._key_headers = []
        if not self.file_path:
            return
        try:
            wb = load_workbook(self.file_path)
            ws = wb.active
            row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            wb.close()
            if row:
                for col_name in row:
                    if col_name is not None:
                        self._key_headers.append(str(col_name))
                self.log(f"[INFO] 已加载 {len(self._key_headers)} 个列名")
        except Exception as e:
            self.log(f"[ERROR] 读取列名失败: {e}")

    # ==================== 日志 ====================
    def log(self, msg):
        self.log_text.configure(state=tk.NORMAL)
        self.log_text.insert(tk.END, msg + "\n")
        self.log_text.see(tk.END)
        self.log_text.configure(state=tk.DISABLED)

    # ==================== 合并逻辑 ====================
    def _get_selected_keys(self):
        return [c.strip() for c in self.entry_keys.get().strip().split(",") if c.strip()]

    # ==================== 弹窗选列 ====================
    def _show_key_picker(self):
        """弹出窗口选择主键列"""
        if not self._key_headers:
            messagebox.showwarning("提示", "请先选择源 Excel 文件以加载列名")
            return

        existing = [c.strip() for c in self.entry_keys.get().strip().split(",") if c.strip()]

        popup = tk.Toplevel(self.parent)
        popup.title("选择主键列")
        popup.geometry("300x350")
        popup.resizable(False, False)
        popup.transient(self.parent)
        popup.grab_set()
        popup.update_idletasks()
        x = self.parent.winfo_rootx() + (self.parent.winfo_width() - 300) // 2
        y = self.parent.winfo_rooty() + (self.parent.winfo_height() - 350) // 2
        popup.geometry(f"+{x}+{y}")

        bg = self.app.COLOR_CONTENT_BG
        f = tk.Frame(popup, bg=bg, padx=12, pady=12)
        f.pack(fill=tk.BOTH, expand=True)

        tk.Label(f, text="Ctrl+点击多选，按选择顺序输出",
                 font=("Microsoft YaHei", 9), bg=bg, fg="#7f8c8d"
                 ).pack(anchor=tk.W)

        list_frame = tk.Frame(f, bg=bg)
        list_frame.pack(fill=tk.BOTH, expand=True, pady=(8, 12))

        lb = tk.Listbox(list_frame, font=("Microsoft YaHei", 10),
                        selectmode=tk.EXTENDED,
                        bg="white", fg="#2c3e50", relief="flat",
                        selectbackground="#3498db", selectforeground="white",
                        exportselection=False)
        sb = tk.Scrollbar(list_frame, orient=tk.VERTICAL)
        lb.config(yscrollcommand=sb.set)
        sb.config(command=lb.yview)
        lb.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        sb.pack(side=tk.RIGHT, fill=tk.Y)

        for col in self._key_headers:
            lb.insert(tk.END, col)

        sel_order = []
        for i, col in enumerate(self._key_headers):
            if col in existing:
                lb.selection_set(i)
                sel_order.append(i)

        def on_select_change(event):
            old_set = set(sel_order)
            new_set = set(lb.curselection())
            removed = old_set - new_set
            added = new_set - old_set
            sel_order[:] = [i for i in sel_order if i not in removed]
            for i in sorted(added):
                if i not in sel_order:
                    sel_order.append(i)

        lb.bind("<<ListboxSelect>>", on_select_change)

        btn_row = tk.Frame(f, bg=bg)
        btn_row.pack(fill=tk.X)

        def on_ok():
            selected = [lb.get(i) for i in sel_order]
            self._set_readonly_value(self.entry_keys, ", ".join(selected))
            popup.destroy()

        tk.Button(btn_row, text="取消", font=("Microsoft YaHei", 10),
                  width=8, bg="#ecf0f1", relief="flat", cursor="hand2",
                  command=popup.destroy).pack(side=tk.RIGHT, padx=(8, 0))
        tk.Button(btn_row, text="确定", font=("Microsoft YaHei", 10, "bold"),
                  width=8, bg="#3498db", fg="white", relief="flat", cursor="hand2",
                  activebackground="#2980b9", activeforeground="white",
                  command=on_ok).pack(side=tk.RIGHT)

        popup.bind("<Return>", lambda e: on_ok())

    def _get_agg_ops(self):
        return [op for label, op in AGG_OPTS if self.agg_vars[label].get()]

    def _start_merge(self):
        if not self.file_path:
            messagebox.showwarning("提示", "请先选择一个 Excel 文件")
            return

        keys = self._get_selected_keys()
        if not keys:
            messagebox.showwarning("提示", "请至少选择一个主键列")
            return

        ops = self._get_agg_ops()
        if not ops:
            messagebox.showwarning("提示", "请至少选择一种聚合方式")
            return

        out_path = self.entry_out.get().strip()
        if not out_path:
            messagebox.showwarning("提示", "请设置输出路径")
            return

        self.btn_start.configure(state=tk.DISABLED, text="合并中，请稍候...")
        threading.Thread(target=self._do_merge, args=(keys, ops, out_path), daemon=True).start()

    def _do_merge(self, keys, ops, out_path):
        try:
            self.log(f"\n{'='*50}")
            self.log(f"[{datetime.now().strftime('%H:%M:%S')}] 开始数据合并...")
            self.log(f"  文件: {os.path.basename(self.file_path)}")
            self.log(f"  主键列: {', '.join(keys)}")
            self.log(f"  聚合方式: {', '.join(self._get_agg_ops())}")

            # ---- 读取文件 ----
            wb = load_workbook(self.file_path, read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            try:
                headers = next(rows_iter)
            except StopIteration:
                self.log("[ERROR] 文件为空")
                wb.close()
                self.app.root.after(0, self._done)
                return

            headers = [str(h) if h is not None else f"Column_{j}" for j, h in enumerate(headers)]

            # 验证主键列
            for key in keys:
                if key not in headers:
                    self.log(f"[ERROR] 文件中不存在主键列「{key}」")
                    wb.close()
                    self.app.root.after(0, self._done)
                    return

            # 读取数据
            all_rows = []
            numeric_cols = set()
            non_numeric_cols = set()

            for row in rows_iter:
                if row is None:
                    continue
                row_dict = {}
                for j, val in enumerate(row):
                    if j >= len(headers):
                        break
                    col = headers[j]
                    row_dict[col] = val

                for col, val in row_dict.items():
                    if col in keys:
                        continue
                    if col in numeric_cols or col in non_numeric_cols:
                        continue
                    if isinstance(val, (int, float)) and not isinstance(val, bool):
                        numeric_cols.add(col)
                    elif val is not None:
                        non_numeric_cols.add(col)

                all_rows.append(row_dict)

            wb.close()
            self.log(f"  共读取 {len(all_rows)} 行数据")
            self.log(f"  数值列（将聚合）: {sorted(numeric_cols)}")
            self.log(f"  非数值列（取首行）: {sorted(non_numeric_cols)}")

            # ---- 按主键分组 ----
            groups = defaultdict(list)
            for row_dict in all_rows:
                key_tuple = tuple(row_dict.get(k) for k in keys)
                groups[key_tuple].append(row_dict)

            self.log(f"  按主键分组后共 {len(groups)} 组")

            if len(groups) == len(all_rows):
                self.log("[INFO] 没有发现重复主键，无需合并，将原样输出")

            # ---- 计算聚合列名映射 ----
            agg_col_map = {}
            for col in numeric_cols:
                agg_col_map[col] = {}
                for op in ops:
                    op_label = next(l for l, o in AGG_OPTS if o == op)
                    agg_col_map[col][op] = f"{col}_{op_label}"

            # ---- 构建输出 ----
            all_header_cols = []
            for h in headers:
                if h not in numeric_cols:
                    all_header_cols.append(h)

            output_headers = list(all_header_cols)
            for col in numeric_cols:
                for op in ops:
                    output_headers.append(agg_col_map[col][op])

            result_rows = []
            aggregated_row_indices = set()  # 记录哪些结果行是汇总行（原始数据中存在重复主键）
            for idx, (key_tuple, rows) in enumerate(groups.items()):
                if len(rows) > 1:
                    aggregated_row_indices.add(idx)  # 多行汇总 → 标记黄底
                result_row = {}

                # 非数值列 + 主键列：取第一行
                for col in all_header_cols:
                    if col in keys:
                        idx = keys.index(col)
                        result_row[col] = key_tuple[idx] if idx < len(key_tuple) else ""
                    else:
                        # 非数值非主键列：取第一行
                        for r in rows:
                            v = r.get(col)
                            if v is not None:
                                result_row[col] = v
                                break
                        else:
                            result_row[col] = ""

                # 数值列做聚合
                for col in numeric_cols:
                    values = [r.get(col) for r in rows if r.get(col) is not None]
                    for op in ops:
                        col_name = agg_col_map[col][op]
                        filtered = [v for v in values if isinstance(v, (int, float))]
                        if op == "sum":
                            result_row[col_name] = sum(filtered)
                        elif op == "count":
                            result_row[col_name] = len(values)
                        elif op == "avg":
                            result_row[col_name] = sum(filtered) / len(filtered) if filtered else 0
                        elif op == "max":
                            result_row[col_name] = max(filtered) if filtered else None
                        elif op == "min":
                            result_row[col_name] = min(filtered) if filtered else None

                result_rows.append(result_row)

            self.log(f"  其中汇总行（重复主键合并→标黄）: {len(aggregated_row_indices)} 行")

            # ---- 写入结果 ----
            self.log(f"  写入结果到 {out_path}...")
            out_wb = Workbook()
            out_ws = out_wb.active
            out_ws.title = "合并结果"

            for j, header in enumerate(output_headers, 1):
                cell = out_ws.cell(row=1, column=j, value=header)
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = HEADER_ALIGNMENT
                cell.border = HEADER_BORDER

            for i, row_data in enumerate(result_rows, 2):
                is_aggregated = (i - 2) in aggregated_row_indices
                for j, header in enumerate(output_headers, 1):
                    val = row_data.get(header, "")
                    cell = out_ws.cell(row=i, column=j, value=val)
                    cell.alignment = CELL_ALIGNMENT
                    if is_aggregated:
                        cell.fill = YELLOW_FILL

            # 自适应列宽
            for col_cells in out_ws.columns:
                col_letter = col_cells[0].column_letter
                max_len = 0
                for cell in col_cells:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                out_ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

            out_wb.save(out_path)
            out_wb.close()

            self.log(f"[OK] 合并完成！共输出 {len(result_rows)} 行，保存至:\n  {out_path}")

        except Exception as e:
            self.log(f"[ERROR] 合并失败: {e}")
            import traceback
            self.log(traceback.format_exc())

        self.app.root.after(0, self._done)

    def _done(self):
        self.btn_start.configure(state=tk.NORMAL, text="⚡ 开始合并")
