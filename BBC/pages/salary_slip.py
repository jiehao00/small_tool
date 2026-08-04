# -*- coding: utf-8 -*-
"""
工资条生成页面
根据汇总表和明细表，按人员拆分生成单独的工资条 Excel 文件。
"""

import os
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
from datetime import datetime

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    import xlrd
    XLRD_AVAILABLE = True
except ImportError:
    XLRD_AVAILABLE = False


def create_page(parent, app):
    """创建工资条生成页面，返回 Frame"""
    builder = SalarySlipBuilder(parent, app)
    return builder.frame


class SalarySlipBuilder:
    """工资条生成页面"""

    def __init__(self, parent, app):
        self.app = app
        self.bg = app.COLOR_CONTENT_BG
        self.frame = tk.Frame(parent, bg=self.bg)
        self._build_ui()

    # --------------- UI 构建 ---------------
    def _build_ui(self):
        # 标题
        title_row = tk.Frame(self.frame, bg=self.bg)
        title_row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(title_row, text="★ 工资条生成", font=("Microsoft YaHei", 16, "bold"),
                 bg=self.bg, fg="#2c3e50").pack(side=tk.LEFT)
        self.lbl_status = tk.Label(title_row, text="",
                                   font=("Microsoft YaHei", 9),
                                   bg=self.bg, fg="#95a5a6")
        self.lbl_status.pack(side=tk.LEFT, padx=(12, 0))

        # 分隔线
        tk.Frame(self.frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ---- 参数区域 ----
        param_frame = tk.Frame(self.frame, bg=self.bg)
        param_frame.pack(fill=tk.X, pady=(0, 10))

        # 汇总表文件
        self._file_row(param_frame, "汇总表文件", "entry_summary", self._browse_summary)

        # 明细表文件
        self._file_row(param_frame, "明细表文件", "entry_detail", self._browse_detail)

        # 参数输入行（汇总对比列、明细对比列、明细剔除列、明细剔除值）
        param_row1 = tk.Frame(param_frame, bg=self.bg)
        param_row1.pack(fill=tk.X, pady=(0, 8))
        self._param_input(param_row1, "汇总对比列", "entry_summary_key", "工号", 25)
        self._param_input(param_row1, "明细对比列", "entry_detail_key", "业务员代码", 25)

        param_row2 = tk.Frame(param_frame, bg=self.bg)
        param_row2.pack(fill=tk.X, pady=(0, 8))
        self._param_input(param_row2, "明细剔除列", "entry_filter_col", "实发佣金", 25)
        self._param_input(param_row2, "剔除列值", "entry_filter_val", "0", 10)

        # 输出目录
        self._file_row(param_frame, "输出目录", "entry_dir_out", self._browse_dir_out)

        # ---- 操作按钮（居中） ----
        btn_row = tk.Frame(self.frame, bg=self.bg)
        btn_row.pack(fill=tk.X, pady=(0, 10))
        self.btn_start = ttk.Button(btn_row, text="⚡ 开始生成",
                                    style="Primary.TButton",
                                    command=self._start_process)
        self.btn_start.pack(anchor=tk.CENTER)

        # ---- 日志区域 ----
        log_label = tk.Frame(self.frame, bg=self.bg)
        log_label.pack(fill=tk.X)
        tk.Label(log_label, text="处理日志", font=("Microsoft YaHei", 10, "bold"),
                 bg=self.bg, fg="#2c3e50").pack(anchor=tk.W, pady=(0, 4))

        log_frame = tk.Frame(self.frame, bg="white", padx=2, pady=2)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 8))
        self.log_text = tk.Text(log_frame, font=("Consolas", 9),
                                bg="#fafbfc", fg="#2c3e50",
                                wrap=tk.WORD, state=tk.DISABLED,
                                relief=tk.FLAT, padx=8, pady=6)
        scroll = ttk.Scrollbar(log_frame, orient=tk.VERTICAL,
                               command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scroll.set)
        self.log_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scroll.pack(side=tk.RIGHT, fill=tk.Y)

    # --------------- UI 辅助 ---------------
    def _file_row(self, parent, label, attr, cmd):
        row = tk.Frame(parent, bg=self.bg)
        row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(row, text=label, font=("Microsoft YaHei", 10),
                 bg=self.bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        entry = ttk.Entry(row, style="Readonly.TEntry", state="readonly")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="📂 浏览", style="Secondary.TButton",
                   command=cmd).pack(side=tk.LEFT)
        setattr(self, attr, entry)

    def _param_input(self, parent, label, attr, default, width):
        """创建标签 + 可编辑输入框"""
        tk.Label(parent, text=label, font=("Microsoft YaHei", 10),
                 bg=self.bg, width=10, anchor=tk.E).pack(
                     side=tk.LEFT, padx=(0, 6))
        entry = ttk.Entry(parent, width=width, font=("Microsoft YaHei", 10))
        entry.insert(0, default)
        entry.pack(side=tk.LEFT, padx=(0, 14))
        setattr(self, attr, entry)

    def _set_entry_path(self, entry, path):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, path)
        entry.configure(state="readonly")

    # --------------- 浏览对话框 ---------------
    def _browse_summary(self):
        path = filedialog.askopenfilename(
            title="选择汇总表 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_summary, path)
            self._log(f"已选择汇总表: {path}")

    def _browse_detail(self):
        path = filedialog.askopenfilename(
            title="选择明细表 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_detail, path)
            self._log(f"已选择明细表: {path}")

    def _browse_dir_out(self):
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self._set_entry_path(self.entry_dir_out, path)
            self._log(f"已设置输出目录: {path}")

    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.app.log(message)

    # --------------- 开始处理 ---------------
    def _start_process(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        summary_path = self.entry_summary.get().strip()
        detail_path = self.entry_detail.get().strip()
        out_dir = self.entry_dir_out.get().strip()

        summary_key = self.entry_summary_key.get().strip()
        detail_key = self.entry_detail_key.get().strip()
        filter_col = self.entry_filter_col.get().strip()
        filter_val = self.entry_filter_val.get().strip()

        if not summary_path:
            messagebox.showwarning("参数缺失", "请选择汇总表文件！")
            return
        if not detail_path:
            messagebox.showwarning("参数缺失", "请选择明细表文件！")
            return
        if not out_dir:
            messagebox.showwarning("参数缺失", "请选择输出目录！")
            return
        if not summary_key or not detail_key:
            messagebox.showwarning("参数缺失", "请填写汇总对比列和明细对比列！")
            return
        if not filter_col:
            messagebox.showwarning("参数缺失", "请填写明细剔除列！")
            return

        if not os.path.isfile(summary_path):
            messagebox.showerror("文件错误", f"汇总表文件不存在:\n{summary_path}")
            return
        if not os.path.isfile(detail_path):
            messagebox.showerror("文件错误", f"明细表文件不存在:\n{detail_path}")
            return

        self.btn_start.configure(state="disabled", text="处理中...")
        t = threading.Thread(
            target=self._do_process,
            args=(summary_path, detail_path, out_dir,
                  summary_key, detail_key, filter_col, filter_val),
            daemon=True)
        t.start()

    def _do_process(self, summary_path, detail_path, out_dir,
                    summary_key, detail_key, filter_col, filter_val):
        """核心处理逻辑（参考 combine_salary.py 完整复刻）"""
        try:
            import pandas as pd
            import re

            self._log("=" * 50)
            self._log("开始生成工资条...")

            # ============ 辅助函数 ============
            def find_col(df, keywords):
                """在 DataFrame 列中查找包含指定关键词的列名"""
                for col in df.columns:
                    col_str = str(col).strip()
                    for kw in keywords:
                        if kw in col_str:
                            return col
                return None

            def find_col_index(df_raw, keywords, search_rows=3):
                """在无表头 DataFrame 中查找包含关键词的列索引"""
                for r in range(min(search_rows, len(df_raw))):
                    for c in range(df_raw.shape[1]):
                        val = str(df_raw.iloc[r, c]).strip()
                        for kw in keywords:
                            if kw in val:
                                return c
                return None

            def safe_cell_value(val):
                """将长数字转为文本，避免 Excel 显示为科学计数法"""
                if pd.isna(val):
                    return val
                if isinstance(val, float):
                    s = str(val)
                    if 'e' in s.lower() or ('e' not in s.lower() and len(s.replace('.', '')) >= 12):
                        try:
                            return str(int(val))
                        except (ValueError, OverflowError):
                            return val
                if isinstance(val, int):
                    s = str(val)
                    if len(s) >= 12:
                        return s
                return val

            def safe_filename(name):
                return re.sub(r'[\\/*?:"<>|]', '', str(name))

            def is_id_column(col_name):
                """判断列名是否为 ID 类列（需要文本格式避免科学计数法）"""
                col_str = str(col_name)
                for kw in ['代码', '工号', '序号', 'ID', '编号', '卡号', '账号',
                           '保单号', '单号', '合同号', '凭证号', '流水号',
                           '手机', '电话', '身份证']:
                    if kw in col_str:
                        return True
                if col_str.endswith('号') or col_str.endswith('码'):
                    return True
                if summary_key in col_str or detail_key in col_str:
                    return True
                return False

            # ============ 1. 读取明细数据（所有 sheet） ============
            self._log("正在读取明细数据...")
            detail_xl = pd.ExcelFile(detail_path)
            detail_sheet_names = detail_xl.sheet_names
            self._log(f"  明细共有 {len(detail_sheet_names)} 个sheet: {', '.join(detail_sheet_names)}")

            detail_sheets = []  # [{name, df, cols, biz_col}]
            for sname in detail_sheet_names:
                df = pd.read_excel(detail_path, sheet_name=sname)
                # 修复 Unnamed 首列
                if str(df.columns[0]).startswith('Unnamed'):
                    df.rename(columns={df.columns[0]: '序号'}, inplace=True)
                cols = df.columns.tolist()
                # 查找用户指定的对比列
                biz_col = find_col(df, [detail_key])
                if biz_col is None:
                    self._log(f'  警告: sheet "{sname}" 中未找到"{detail_key}"列，跳过')
                    continue
                # 剔除指定值的行
                if filter_col and filter_val:
                    f_col = find_col(df, [filter_col])
                    if f_col is not None:
                        before = len(df)
                        try:
                            filter_num = float(filter_val)
                            df[f_col] = pd.to_numeric(df[f_col], errors='coerce')
                            df = df[~(df[f_col].isna() | (df[f_col] == filter_num))]
                        except ValueError:
                            df[f_col] = df[f_col].astype(str).str.strip()
                            df = df[df[f_col] != filter_val]
                        after = len(df)
                        self._log(f'  {sname}: 剔除"{filter_col}"={filter_val}的数据，{before}行 → {after}行')
                    else:
                        self._log(f'  警告: sheet "{sname}" 中未找到剔除列"{filter_col}"，跳过过滤')
                # 创建字符串版业务员代码用于匹配
                df['_biz_code_str'] = df[biz_col].apply(
                    lambda x: str(int(float(x))) if pd.notna(x) and not isinstance(x, str) else str(x)
                )
                detail_sheets.append({'name': sname, 'df': df, 'cols': cols, 'biz_col': biz_col})
                self._log(f"  {sname}: {len(df)} 行, {len(cols)} 列")
            if not detail_sheets:
                raise ValueError(f'明细文件中未找到任何包含"{detail_key}"列的sheet')

            # ============ 2. 解析汇总数据（所有 sheet） ============
            def parse_summary_sheet(sheet_name):
                df_raw = pd.read_excel(summary_path, sheet_name=sheet_name, header=None)
                # 动态查找对比列和姓名列
                key_col = find_col_index(df_raw, [summary_key])
                name_col = find_col_index(df_raw, ['姓名'])
                if key_col is None or name_col is None:
                    self._log(f'  警告: sheet "{sheet_name}" 中未找到"{summary_key}"或姓名列，跳过')
                    return None
                main_headers = [str(df_raw.iloc[0, c]) if pd.notna(df_raw.iloc[0, c]) else '' for c in range(df_raw.shape[1])]
                sub_headers = [str(df_raw.iloc[1, c]) if pd.notna(df_raw.iloc[1, c]) else '' for c in range(df_raw.shape[1])]
                persons = []
                # 每4行一组（一人一区块）
                for idx in range(2, len(df_raw), 4):
                    if idx >= len(df_raw):
                        break
                    row = df_raw.iloc[idx]
                    key_val = row.iloc[key_col]
                    if pd.notna(key_val) and str(key_val) != summary_key:
                        data = [row.iloc[c] for c in range(len(row))]
                        kv = str(int(float(key_val))) if not isinstance(key_val, str) else str(key_val)
                        xm = str(row.iloc[name_col])
                        persons.append({
                            'gonghao': kv, 'name': xm, 'data': data,
                            'sheet': sheet_name, 'main_headers': main_headers, 'sub_headers': sub_headers
                        })
                return {'name': sheet_name, 'main_headers': main_headers, 'sub_headers': sub_headers, 'persons': persons}

            self._log("正在解析汇总数据...")
            sum_xl = pd.ExcelFile(summary_path)
            sum_sheet_names = sum_xl.sheet_names
            self._log(f"  汇总共有 {len(sum_sheet_names)} 个sheet: {', '.join(sum_sheet_names)}")
            all_persons = []
            sum_max_cols = 0
            for sname in sum_sheet_names:
                result = parse_summary_sheet(sname)
                if result is None:
                    continue
                all_persons.extend(result['persons'])
                sum_max_cols = max(sum_max_cols, len(result['main_headers']))
                self._log(f"  {sname}: {len(result['persons'])} 人")
            self._log(f"共计 {len(all_persons)} 人")
            if not all_persons:
                raise ValueError('汇总文件中未找到任何人员数据')

            # ============ 3. 匹配明细 ============
            self._log("正在匹配明细...")
            for p in all_persons:
                p['details'] = []
                for ds in detail_sheets:
                    matched = ds['df'][ds['df']['_biz_code_str'] == p['gonghao']].drop(columns=['_biz_code_str'])
                    p['details'].append({
                        'label': ds['name'],
                        'df': matched,
                        'cols': ds['cols']
                    })
                parts = []
                for d in p['details']:
                    if len(d['df']) > 0:
                        parts.append(f'{d["label"]}{len(d["df"])}条')
                detail_info = ', '.join(parts) if parts else '无明细'
                self._log(f"  {summary_key} {p['gonghao']} ({p['name']}): {detail_info}")

            # ============ 4. 样式定义 ============
            header_font = Font(name='微软雅黑', bold=True, size=10)
            header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
            sub_header_font = Font(name='微软雅黑', bold=True, size=9)
            sub_header_fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
            detail_palette = [
                PatternFill(start_color='FCE4D6', end_color='FCE4D6', fill_type='solid'),
                PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid'),
                PatternFill(start_color='E4DFEC', end_color='E4DFEC', fill_type='solid'),
                PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
                PatternFill(start_color='D9EAD3', end_color='D9EAD3', fill_type='solid'),
                PatternFill(start_color='FADBD8', end_color='FADBD8', fill_type='solid'),
            ]
            detail_fills = {ds['name']: detail_palette[i % len(detail_palette)]
                            for i, ds in enumerate(detail_sheets)}
            data_font = Font(name='微软雅黑', size=9)
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

            def write_detail_section(ws, start_row, df, cols, label):
                """写入一个明细 sheet 的数据区块"""
                if len(df) == 0:
                    return start_row
                r = start_row + 5  # 明细区块前空 5 行
                fill = detail_fills.get(label, detail_palette[0])
                # 明细表头
                for col_idx, col_name in enumerate(cols):
                    cell = ws.cell(row=r, column=col_idx + 1)
                    cell.value = col_name
                    cell.font = header_font
                    cell.fill = fill
                    cell.alignment = center_align
                    cell.border = thin_border
                r += 1
                # 明细数据
                for _, detail_row in df.iterrows():
                    for col_idx, col_name in enumerate(cols):
                        cell = ws.cell(row=r, column=col_idx + 1)
                        val = detail_row[col_name]
                        if pd.notna(val):
                            cell.value = safe_cell_value(val)
                        if is_id_column(col_name) and pd.notna(val):
                            cell.number_format = '@'
                        cell.font = data_font
                        cell.alignment = center_align
                        cell.border = thin_border
                    r += 1
                return r

            # 计算最大列数
            detail_max_cols = max((len(ds['cols']) for ds in detail_sheets), default=0)
            max_cols = max(sum_max_cols, detail_max_cols)

            # ============ 5. 生成输出文件 ============
            self._log("正在生成输出文件...")
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception as e:
                self._log(f"[错误] 创建输出目录失败: {e}")
                return

            generated_count = 0
            for pi, p in enumerate(all_persons):
                seq = pi + 1
                gh = p['gonghao']
                name = p['name']
                main_headers = p['main_headers']
                sub_headers = p['sub_headers']
                person_data = p['data']

                safe_name = safe_filename(name)
                out_filename = f'{seq}_{safe_name}_{gh}.xlsx'
                out_path = os.path.join(out_dir, out_filename)

                wb_out = Workbook()
                ws = wb_out.active
                ws.title = f'{safe_name}_{gh}'

                current_row = 1
                # 一级表头
                for col_idx in range(len(main_headers)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = main_headers[col_idx].replace('\n', '')
                    if val:
                        cell.value = val
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

                # 二级表头
                for col_idx in range(len(sub_headers)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = sub_headers[col_idx].replace('\n', '')
                    if val:
                        cell.value = val
                        cell.font = sub_header_font
                        cell.fill = sub_header_fill
                        cell.alignment = center_align
                        cell.border = thin_border
                    else:
                        cell.border = thin_border
                current_row += 1

                # 汇总数据行
                for col_idx in range(len(person_data)):
                    cell = ws.cell(row=current_row, column=col_idx + 1)
                    val = person_data[col_idx]
                    if pd.notna(val):
                        cell.value = safe_cell_value(val)
                    header_name = main_headers[col_idx] if col_idx < len(main_headers) else ''
                    if is_id_column(header_name) and pd.notna(val):
                        cell.number_format = '@'
                    cell.font = data_font
                    cell.alignment = center_align
                    cell.border = thin_border
                current_row += 1

                # 遍历所有明细 sheet
                for d in p['details']:
                    current_row = write_detail_section(ws, current_row, d['df'], d['cols'], d['label'])

                # 调整列宽
                for col_idx in range(1, max_cols + 1):
                    ws.column_dimensions[get_column_letter(col_idx)].width = 14

                wb_out.save(out_path)
                wb_out.close()
                generated_count += 1
                self._log(f"  [{seq}/{len(all_persons)}] {out_filename}")

            self._log(f"共生成 {generated_count} 个工资条文件")
            self._log("=" * 50)
            self._log("全部完成！")

            msg = f"处理完成！\n\n共生成 {generated_count} 个工资条文件\n\n保存目录:\n{out_dir}"
            if generated_count == 0:
                msg += "\n\n（没有找到匹配的数据，请检查参数设置）"
            messagebox.showinfo("处理完成", msg)

        except Exception as e:
            import traceback
            self._log(f"[错误] {e}")
            self._log(traceback.format_exc())
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(
                state="normal", text="⚡ 开始生成"))

    # --------------- 工具方法 ---------------
    @staticmethod
    def _find_col(headers, target):
        """在表头列表中查找精确匹配的列索引（1-based），找不到返回 None"""
        target = target.strip()
        for idx, h in enumerate(headers):
            if h.strip() == target:
                return idx + 1
        return None

    def _read_excel(self, path, label):
        """统一读取 .xls 或 .xlsx，返回 workbook-like 对象（有 .active 和 .close()）"""
        ext = os.path.splitext(path)[1].lower()
        if ext in ('.xls',) and XLRD_AVAILABLE:
            return _XlrdBook(path)
        elif ext in ('.xlsx', '.xlsm', '.xltx', '.xltm', '.xls'):
            if not OPENPYXL_AVAILABLE:
                raise RuntimeError("需要安装 openpyxl 来读取 .xlsx 文件")
            try:
                return load_workbook(path, data_only=True)
            except Exception:
                if XLRD_AVAILABLE:
                    self._log(f"[提示] openpyxl 无法打开，尝试用 xlrd 读取...")
                    return _XlrdBook(path)
                raise
        else:
            raise RuntimeError(f"不支持的文件格式: {ext}")

    @staticmethod
    def _safe_filename(name):
        """将字符串转为安全的文件名"""
        invalid_chars = '<>:"/\\|?*'
        for ch in invalid_chars:
            name = name.replace(ch, '_')
        name = name.strip().strip('.')
        if not name:
            name = "未命名"
        return name


# ─────────────────  xlrd 适配器（兼容 .xls 旧格式） ─────────────────
class _XlrdCell:
    """模拟 openpyxl 的 Cell 对象"""
    __slots__ = ('_sheet', '_row', '_col')

    def __init__(self, sheet, row_1based, col_1based):
        self._sheet = sheet
        self._row = row_1based - 1   # xlrd 是 0-based
        self._col = col_1based - 1

    @property
    def value(self):
        val = self._sheet.cell_value(self._row, self._col)
        if isinstance(val, str) and val == '':
            return None
        return val


class _XlrdSheet:
    """模拟 openpyxl 的 Worksheet 对象"""

    def __init__(self, sheet):
        self._sheet = sheet

    @property
    def title(self):
        return self._sheet.name

    @property
    def max_row(self):
        return self._sheet.nrows

    @property
    def max_column(self):
        return self._sheet.ncols

    def cell(self, row, col):
        return _XlrdCell(self._sheet, row, col)


class _XlrdBook:
    """模拟 openpyxl 的 Workbook，提供 .active 和 .close()"""

    def __init__(self, path):
        self._book = xlrd.open_workbook(path)
        self.active = _XlrdSheet(self._book.sheet_by_index(0))

    def close(self):
        pass  # xlrd 无需显式关闭
