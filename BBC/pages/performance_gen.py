# -*- coding: utf-8 -*-
"""
业绩生成页面模块

功能：导入一个 Excel 文档，处理后生成另一个文档。
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import threading
import os
import tempfile
import shutil
from datetime import datetime

from collections import defaultdict
from datetime import datetime, timedelta
import re

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

# ---- 可选库：PNG / PDF 导出（暂不使用 PIL） ----
try:
    import win32com.client
    _WIN32COM_OK = True
except ImportError:
    _WIN32COM_OK = False

try:
    import fitz  # PyMuPDF — PDF 转 PNG，无需 poppler
    _PYMUPDF_OK = True
except ImportError:
    _PYMUPDF_OK = False


from .widgets import RoundedCheckbox


def create_page(parent, app):
    """工厂函数：构建并返回业绩生成页面 Frame"""
    builder = PerformanceBuilder(parent, app)
    return builder.build()


class PerformanceBuilder:
    """业绩生成页面构建 & 业务逻辑"""

    # PyInstaller 打包后资源路径解析
    def __init__(self, parent, app):
        self.parent = parent
        self.app = app       # ExcelToolApp 实例，提供 log()、配色等

    # --------------- UI 组件引用 ---------------
    entry_file_in = None
    entry_file_personnel = None
    lbl_file_info = None
    lbl_personnel_info = None
    entry_dir_out = None
    entry_prefix = None
    btn_start = None
    log_text = None
    fmt_excel_var = None
    fmt_png_var = None
    fmt_pdf_var = None
    cb_excel = None
    cb_png = None
    cb_pdf = None
    btn_toggle_all = None
    filter_col_var = None
    filter_val_var = None
    cbo_filter_col = None
    cbo_filter_val = None
    filter_enable_var = None
    chk_filter_enable = None
    report_title_var = None
    entry_report_title = None
    _filter_file_path = None

    # --------------- 构建 UI ---------------
    def build(self):
        bg = self.app.COLOR_CONTENT_BG

        frame = tk.Frame(self.parent, bg=bg)

        # 页面标题
        tk.Label(
            frame, text="★ 业绩生成",
            font=("Microsoft YaHei", 16, "bold"), bg=bg, fg="#2c3e50"
        ).pack(anchor=tk.W, pady=(0, 12))

        # 细分隔线
        tk.Frame(frame, height=1, bg="#e8ecf0").pack(fill=tk.X, pady=(0, 14))

        # ========== 输入文件 ==========
        in_frame = ttk.LabelFrame(frame, text=" 输入文件 ", style="Card.TLabelframe",
            padding=(16, 10))
        in_frame.pack(fill=tk.X, pady=(0, 12))

        self._file_row(in_frame, "mis报表数据", "entry_file_in", self._browse_file_in)

        # 文件信息区域
        self.lbl_file_info = tk.Label(
            in_frame, text="", font=("Microsoft YaHei", 9),
            bg=bg, fg="#95a5a6", anchor=tk.W
        )
        self.lbl_file_info.pack(fill=tk.X, pady=(6, 0))

        # 基础数据文件（必传）
        self._file_row(in_frame, "基础数据", "entry_file_personnel", self._browse_file_personnel)

        self.lbl_personnel_info = tk.Label(
            in_frame, text="", font=("Microsoft YaHei", 9),
            bg=bg, fg="#95a5a6", anchor=tk.W
        )
        self.lbl_personnel_info.pack(fill=tk.X, pady=(6, 0))

        # 排除条件（可选）
        filter_frame = tk.Frame(in_frame, bg=bg)
        filter_frame.pack(fill=tk.X, pady=(10, 0))
        tk.Label(filter_frame, text="排除条件", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))

        self.filter_enable_var = tk.BooleanVar(value=True)
        self.chk_filter_enable = RoundedCheckbox(filter_frame,
            variable=self.filter_enable_var, bg=bg)
        self.chk_filter_enable.pack(side=tk.LEFT, padx=(0, 6))
        self.filter_enable_var.trace_add("write",
            lambda *a: self._on_filter_enable_changed())
        tk.Label(filter_frame, text="启用", font=("Microsoft YaHei", 10),
            bg=bg, fg="#2c3e50").pack(side=tk.LEFT, padx=(0, 8))

        tk.Label(filter_frame, text="当列", font=("Microsoft YaHei", 9),
            bg=bg).pack(side=tk.LEFT, padx=(0, 4))
        self.filter_col_var = tk.StringVar(value="业务类型")
        self.cbo_filter_col = ttk.Combobox(filter_frame, textvariable=self.filter_col_var,
            values=["业务类型"], state="readonly", width=14, font=("Microsoft YaHei", 10),
            style="CondCol.TCombobox")
        self.cbo_filter_col.pack(side=tk.LEFT, padx=(0, 4))
        self.cbo_filter_col.bind("<<ComboboxSelected>>", self._on_filter_col_changed)

        tk.Label(filter_frame, text="的值 =", font=("Microsoft YaHei", 9),
            bg=bg).pack(side=tk.LEFT, padx=(0, 4))
        self.filter_val_var = tk.StringVar(value="续期")
        self.cbo_filter_val = ttk.Combobox(filter_frame, textvariable=self.filter_val_var,
            values=["续期"], state="normal", width=12, font=("Microsoft YaHei", 10),
            style="CondCol.TCombobox")
        self.cbo_filter_val.pack(side=tk.LEFT, padx=(0, 4))

        tk.Label(filter_frame, text="时，跳过该行", font=("Microsoft YaHei", 9),
            bg=bg, fg="#95a5a6").pack(side=tk.LEFT)

        # 报表标题
        title_row = tk.Frame(in_frame, bg=bg)
        title_row.pack(fill=tk.X, pady=(10, 0))
        tk.Label(title_row, text="报表标题", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.report_title_var = tk.StringVar(value="BBC业绩报表")
        self.entry_report_title = ttk.Entry(title_row, style="Normal.TEntry",
            textvariable=self.report_title_var, font=("Microsoft YaHei", 10))
        self.entry_report_title.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # ========== 输出设置 ==========
        out_frame = ttk.LabelFrame(frame, text=" 输出设置 ", style="Card.TLabelframe",
            padding=(16, 10))
        out_frame.pack(fill=tk.X, pady=(0, 12))

        # 输出目录
        dir_row = tk.Frame(out_frame, bg=bg)
        dir_row.pack(fill=tk.X, pady=(0, 10))
        tk.Label(dir_row, text="输出目录", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_dir_out = ttk.Entry(dir_row, style="Readonly.TEntry", state="readonly")
        self.entry_dir_out.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 4))
        ttk.Button(dir_row, text="📂 浏览", style="Secondary.TButton",
            command=self._browse_dir_out).pack(side=tk.LEFT)

        # 文件名前缀
        name_row = tk.Frame(out_frame, bg=bg)
        name_row.pack(fill=tk.X, pady=(0, 8))
        tk.Label(name_row, text="文件名前缀", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        self.entry_prefix = ttk.Entry(name_row, style="Normal.TEntry")
        self.entry_prefix.insert(0, "BBC业绩_")
        self.entry_prefix.pack(side=tk.LEFT, fill=tk.X, expand=True)

        # 输出格式选择（复选框，可多选）
        fmt_row = tk.Frame(out_frame, bg=bg)
        fmt_row.pack(fill=tk.X)
        tk.Label(fmt_row, text="输出格式", font=("Microsoft YaHei", 10),
            bg=bg, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))

        self.fmt_excel_var = tk.BooleanVar(value=True)
        self.fmt_png_var = tk.BooleanVar(value=False)
        self.fmt_pdf_var = tk.BooleanVar(value=False)

        self.cb_excel = self._make_format_checkbox(fmt_row, self.fmt_excel_var, "Excel (.xlsx)")
        self.cb_png = self._make_format_checkbox(fmt_row, self.fmt_png_var, "PNG (.png)")
        self.cb_pdf = self._make_format_checkbox(fmt_row, self.fmt_pdf_var, "PDF (.pdf)")

        self.btn_toggle_all = ttk.Button(fmt_row, text="全选", style="Secondary.TButton",
            width=8, command=self._toggle_all_formats)
        self.btn_toggle_all.pack(side=tk.LEFT, padx=(10, 0))

        # ========== 开始按钮 ==========
        btn_frame = tk.Frame(frame, bg=bg)
        btn_frame.pack(pady=16)
        self.btn_start = ttk.Button(btn_frame, text="⚡ 开始处理",
            style="Action.TButton", width=22,
            command=self._start_process)
        self.btn_start.pack()

        # ========== 处理日志 ==========
        log_frame = tk.LabelFrame(frame, text=" 处理日志 ",
            font=("Microsoft YaHei", 11, "bold"), bg=bg, fg="#2c3e50",
            padx=10, pady=10, relief="solid", bd=1)
        log_frame.pack(fill=tk.BOTH, expand=True, pady=(0, 5))

        self.log_text = scrolledtext.ScrolledText(log_frame,
            font=("Consolas", 9), wrap=tk.WORD, state=tk.DISABLED,
            bg="#fafafa", fg="#2c3e50", relief="flat", borderwidth=0,
            padx=10, pady=8, height=10)
        self.log_text.pack(fill=tk.BOTH, expand=True)

        return frame

    # --------------- UI 辅助 ---------------
    def _file_row(self, parent, label, attr, cmd):
        row = tk.Frame(parent, bg=self.app.COLOR_CONTENT_BG)
        row.pack(fill=tk.X, pady=(0, 12))
        tk.Label(row, text=label, font=("Microsoft YaHei", 10),
            bg=self.app.COLOR_CONTENT_BG, width=10, anchor=tk.E).pack(side=tk.LEFT, padx=(0, 8))
        entry = ttk.Entry(row, style="Readonly.TEntry", state="readonly")
        entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8))
        ttk.Button(row, text="📂 浏览", style="Secondary.TButton", command=cmd).pack(side=tk.LEFT)
        setattr(self, attr, entry)

    def _make_format_checkbox(self, parent, var, text):
        """创建一个圆角复选框 + 文字标签"""
        bg = self.app.COLOR_CONTENT_BG
        row = tk.Frame(parent, bg=bg)
        row.pack(side=tk.LEFT, padx=(0, 12))
        cb = RoundedCheckbox(row, variable=var, bg=bg)
        cb.pack(side=tk.LEFT)
        tk.Label(row, text=text, font=("Microsoft YaHei", 10),
            bg=bg, fg="#2c3e50").pack(side=tk.LEFT, padx=(4, 0))
        return cb

    def _set_entry_path(self, entry, path):
        entry.configure(state="normal")
        entry.delete(0, tk.END)
        entry.insert(0, path)
        entry.configure(state="readonly")

    # --------------- 浏览对话框 ---------------
    def _browse_file_in(self):
        path = filedialog.askopenfilename(
            title="选择待处理的 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_file_in, path)
            self.app.log(f"已选择输入文件: {path}")
            # 读取文件基本信息
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                row_count = ws.max_row - 1  # 减去表头
                col_count = ws.max_column
                wb.close()
                self.lbl_file_info.config(
                    text=f"工作表: {ws.title}　|　{row_count} 行数据　|　{col_count} 列"
                )
                # 保存路径，并读取表头填充过滤列名下拉
                self._filter_file_path = path
                self._refresh_filter_columns(path)
            except Exception:
                self.lbl_file_info.config(text="")

    def _refresh_filter_columns(self, file_path):
        """读取文件表头，填充列名下拉框，并预选「业务类型」"""
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            # 找表头行
            hdr_row = 1
            for r in range(1, min(ws.max_row + 1, 10)):
                row_text = " ".join(
                    str(ws.cell(r, c).value or "")
                    for c in range(1, min(ws.max_column + 1, 200))
                )
                if any(kw in row_text for kw in ["标保", "保单号", "分单号", "统计时间"]):
                    hdr_row = r
                    break
            header_list = []
            for col in range(1, ws.max_column + 1):
                h = str(ws.cell(hdr_row, col).value or "").strip()
                if h:
                    header_list.append(h)
            wb.close()
            self.cbo_filter_col["values"] = header_list
            if "业务类型" in header_list:
                self.filter_col_var.set("业务类型")
            elif header_list:
                self.filter_col_var.set(header_list[0])
            self._populate_filter_values(file_path, self.filter_col_var.get())
        except Exception:
            pass

    def _browse_file_personnel(self):
        path = filedialog.askopenfilename(
            title="选择人员清单 Excel 文件",
            filetypes=[("Excel 文件", "*.xlsx *.xls"), ("所有文件", "*.*")])
        if path:
            self._set_entry_path(self.entry_file_personnel, path)
            self.app.log(f"已选择人员清单文件: {path}")
            try:
                wb = load_workbook(path, read_only=True, data_only=True)
                ws = wb.active
                row_count = ws.max_row - 1
                col_count = ws.max_column
                wb.close()
                self.lbl_personnel_info.config(
                    text=f"工作表: {ws.title}　|　{row_count} 行数据　|　{col_count} 列"
                )
            except Exception:
                self.lbl_personnel_info.config(text="")

    def _browse_dir_out(self):
        """选择输出目录"""
        path = filedialog.askdirectory(title="选择输出目录")
        if path:
            self._set_entry_path(self.entry_dir_out, path)
            self.app.log(f"已设置输出目录: {path}")

    def _on_filter_col_changed(self, event):
        """列名下拉切换时，动态刷新值下拉的可选项"""
        col_name = self.filter_col_var.get().strip()
        path = self._filter_file_path
        if col_name and path:
            self._populate_filter_values(path, col_name)

    def _on_filter_enable_changed(self):
        """复选框切换时，启用/禁用列名和值下拉框"""
        enabled = self.filter_enable_var.get()
        new_state = "readonly" if enabled else "disabled"
        self.cbo_filter_col.configure(state=new_state)
        self.cbo_filter_val.configure(state="normal" if enabled else "disabled")

    def _populate_filter_values(self, file_path, col_name):
        """读取文件中指定列的去重值，填充到值下拉框"""
        if not file_path or not col_name:
            return
        try:
            wb = load_workbook(file_path, data_only=True)
            ws = wb.active
            # 找表头行
            hdr_row = 1
            for r in range(1, min(ws.max_row + 1, 10)):
                row_text = " ".join(
                    str(ws.cell(r, c).value or "")
                    for c in range(1, min(ws.max_column + 1, 200))
                )
                if any(kw in row_text for kw in ["标保", "保单号", "分单号", "统计时间"]):
                    hdr_row = r
                    break
            # 找列索引
            col_idx = None
            for c in range(1, ws.max_column + 1):
                if str(ws.cell(hdr_row, c).value or "").strip() == col_name:
                    col_idx = c
                    break
            if col_idx is None:
                wb.close()
                return
            # 读取去重值
            unique_vals = set()
            for r in range(hdr_row + 1, ws.max_row + 1):
                v = str(ws.cell(r, col_idx).value or "").strip()
                if v:
                    unique_vals.add(v)
            wb.close()
            self.cbo_filter_val["values"] = sorted(unique_vals)
        except Exception:
            pass

    def _get_selected_formats(self):
        """返回当前选中的输出格式列表，顺序固定为 [Excel, PNG, PDF]"""
        result = []
        if self.fmt_excel_var.get():
            result.append("Excel")
        if self.fmt_png_var.get():
            result.append("PNG")
        if self.fmt_pdf_var.get():
            result.append("PDF")
        return result

    def _toggle_all_formats(self):
        """全选 / 全不选切换"""
        all_selected = all((self.fmt_excel_var.get(),
                            self.fmt_png_var.get(),
                            self.fmt_pdf_var.get()))
        new_val = not all_selected
        self.fmt_excel_var.set(new_val)
        self.fmt_png_var.set(new_val)
        self.fmt_pdf_var.set(new_val)
        self.btn_toggle_all.config(text="全不选" if new_val else "全选")

    # ---------- Excel COM 导出（PNG / PDF） ----------
    def _make_output_path(self, out_dir, prefix, ext):
        """拼接输出文件路径: {out_dir}/{prefix}.{ext}"""
        return os.path.join(out_dir, f"{prefix}{ext}")

    def _setup_page_for_export(self, ws):
        """设置工作表页面：横向 + 所有列缩放到一页宽度"""
        # 打印区域：所有有内容单元格
        last_row = ws.UsedRange.Rows.Count
        last_col = ws.UsedRange.Columns.Count
        ws.PageSetup.PrintArea = ws.Range(
            ws.Cells(1, 1), ws.Cells(last_row, last_col)
        ).Address

        ws.PageSetup.Orientation = 2        # xlLandscape（横向）
        ws.PageSetup.PaperSize = 9          # xlPaperA4
        # 边距设为最小，让内容尽量撑满页面
        ws.PageSetup.LeftMargin = 0
        ws.PageSetup.RightMargin = 0
        ws.PageSetup.TopMargin = 0
        ws.PageSetup.BottomMargin = 0
        ws.PageSetup.HeaderMargin = 0
        ws.PageSetup.FooterMargin = 0

        # 缩放：所有列适配一页宽度，行数不限制
        ws.PageSetup.Zoom = False
        ws.PageSetup.FitToPagesWide = 1
        ws.PageSetup.FitToPagesTall = False

        # 显式设置标题/时间行单元格的对齐方式（Excel COM 对 openpyxl 合并单元格
        # 的 vertical=center 有时不生效，这里强制再设一次）
        try:
            # 常量：xlHAlignCenter=-4108, xlHAlignRight=-4152
            #       xlVAlignCenter=-4108, xlVAlignTop=-4160, xlVAlignBottom=-4107
            # 第1行：水平居中 + 底端对齐
            r1 = ws.Range(ws.Cells(1, 1), ws.Cells(1, last_col))
            r1.HorizontalAlignment = -4108  # xlHAlignCenter
            r1.VerticalAlignment = -4107    # xlVAlignBottom
            # 第2行：水平右对齐 + 底端对齐
            r2 = ws.Range(ws.Cells(2, 1), ws.Cells(2, last_col))
            r2.HorizontalAlignment = -4152  # xlHAlignRight
            r2.VerticalAlignment = -4107    # xlVAlignBottom
        except Exception as align_err:
            self._log(f"[警告] 显式对齐设置失败: {align_err}")
        ws.PageSetup.CenterVertically = True
        ws.PageSetup.CenterHorizontally = True

    def _get_excel_com_app(self):
        """
        探测可用的 Office COM 组件，优先 Microsoft Excel，其次 WPS 表格。
        返回 (excel对象, 软件名) 元组；都不可用则抛出带友好提示的异常。
        """
        # ProgID 候选列表：Microsoft Excel → WPS 表格（不同版本标识不同）
        candidates = [
            ("Excel.Application", "Microsoft Excel"),
            ("Ket.Application",   "WPS 表格"),       # WPS 旧版/部分版本
            ("et.Application",    "WPS 表格"),       # WPS 表格组件
            ("KWPS.Application",  "WPS Office"),     # WPS Office
        ]
        last_err = None
        for prog_id, soft_name in candidates:
            try:
                excel = win32com.client.DispatchEx(prog_id)
                self._log(f"已连接 {soft_name}（{prog_id}）")
                return excel, soft_name
            except Exception as e:
                last_err = e
                continue
        raise RuntimeError(
            "未检测到 Microsoft Excel 或 WPS Office，PDF/图片导出功能需要安装其中之一。\n"
            "请安装 Microsoft Office Excel 或 WPS Office 后重试。"
        )

    def _export_via_com(self, xlsx_path, pdf_path=None, img_path=None):
        """一次 Excel COM 调用，同时导出 PDF 和/或 PNG（避免重复启动 Excel）"""
        if not _WIN32COM_OK:
            raise RuntimeError("导出需要安装 pywin32：pip install pywin32")
        if img_path and not _PYMUPDF_OK:
            raise RuntimeError("图片导出需要安装 PyMuPDF：pip install PyMuPDF")

        self._log("正在调用 Office COM 导出...")
        import pythoncom
        pythoncom.CoInitialize()
        excel = None
        wb = None
        temp_pdf = None
        try:
            fd, temp_pdf = tempfile.mkstemp(suffix=".pdf", prefix="bbc_export_")
            os.close(fd)

            excel, soft_name = self._get_excel_com_app()
            excel.Visible = False
            excel.DisplayAlerts = False
            excel.ScreenUpdating = False
            try:
                excel.Interactive = False
            except Exception:
                pass  # WPS 部分版本不支持 Interactive 属性，忽略
            wb = excel.Workbooks.Open(os.path.abspath(xlsx_path))
            self._setup_page_for_export(wb.Worksheets(1))
            wb.ExportAsFixedFormat(0, os.path.abspath(temp_pdf))  # 0=xlTypePDF
            wb.Close(False)
            wb = None
            excel.Quit()
            excel = None
            self._log(f"{soft_name} PDF 导出完成")

            # PDF 格式：直接复制临时 PDF
            if pdf_path:
                shutil.copy(temp_pdf, pdf_path)
                self._log("PDF 已保存")

            # PNG 格式：PDF → PNG，并裁剪白边
            if img_path:
                import io as _io
                from PIL import Image as _Image
                doc = fitz.open(temp_pdf)
                if doc.page_count > 0:
                    page = doc[0]
                    pix = page.get_pixmap(dpi=200)

                    # 用 PIL 裁剪四周白边
                    img = _Image.open(_io.BytesIO(pix.tobytes("png")))
                    gray = img.convert("L")
                    # 阈值 240：亮于 240 视为白边，变为黑色(0)；暗于 240 是内容，变为白色(255)
                    content_mask = gray.point(lambda v: 0 if v > 240 else 255)
                    bbox = content_mask.getbbox()
                    if bbox:
                        pad = 20  # 内容周围保留 20px 呼吸空间
                        bbox = (
                            max(0, bbox[0] - pad),
                            max(0, bbox[1] - pad),
                            min(img.width, bbox[2] + pad),
                            min(img.height, bbox[3] + pad),
                        )
                        img = img.crop(bbox)
                    img.save(img_path, "PNG")
                doc.close()
                self._log("PNG 已保存")

            return True
        finally:
            if wb is not None:
                wb.Close(False)
            if excel is not None:
                excel.Quit()
            if temp_pdf and os.path.exists(temp_pdf):
                try:
                    os.remove(temp_pdf)
                except OSError:
                    pass
            pythoncom.CoUninitialize()



    def _log(self, message):
        ts = datetime.now().strftime("%H:%M:%S")
        line = f"[{ts}] {message}\n"
        self.log_text.config(state=tk.NORMAL)
        self.log_text.insert(tk.END, line)
        self.log_text.see(tk.END)
        self.log_text.config(state=tk.DISABLED)
        self.app.log(message)

    # --------------- 核心处理 ---------------
    def _start_process(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("缺少依赖", "请先安装 openpyxl：\n\npip install openpyxl")
            return

        file_in = self.entry_file_in.get().strip()
        out_dir = self.entry_dir_out.get().strip()
        prefix = self.entry_prefix.get().strip()
        selected = self._get_selected_formats()

        if not file_in:
            messagebox.showwarning("参数缺失", "请选择输入文件！")
            return
        if not out_dir:
            messagebox.showwarning("参数缺失", "请选择输出目录！")
            return
        if not prefix:
            prefix = "BBC业绩"
        if not selected:
            messagebox.showwarning("参数缺失", "请至少选择一种输出格式！")
            return
        if not os.path.isfile(file_in):
            messagebox.showerror("文件错误", f"输入文件不存在:\n{file_in}")
            return

        file_personnel = self.entry_file_personnel.get().strip()
        if not file_personnel:
            messagebox.showerror("参数缺失", "人员清单为必传项，请选择人员清单文件！")
            return
        if not os.path.isfile(file_personnel):
            messagebox.showerror("文件错误", f"人员清单文件不存在:\n{file_personnel}")
            return

        self.btn_start.configure(state="disabled", text="处理中...")
        if self.filter_enable_var.get():
            filter_col = self.filter_col_var.get().strip()
            filter_val = self.filter_val_var.get().strip()
        else:
            filter_col = ""
            filter_val = ""
        report_title = self.report_title_var.get().strip()
        t = threading.Thread(
            target=self._do_process,
            args=(file_in, out_dir, prefix, selected, file_personnel, filter_col, filter_val, report_title),
            daemon=True)
        t.start()

    def _do_process(self, file_in, out_dir, prefix, selected, file_personnel="", filter_col="", filter_val="", report_title="BBC业绩报表"):
        """核心处理逻辑：读取原始数据 → 生成报表样式（1）"""
        try:
            self._log("=" * 50)
            self._log("开始处理...")

            # --- 1. 加载原始数据 ---
            self._log(f"加载输入文件: {file_in}")
            wb_raw = load_workbook(file_in, data_only=True)
            ws_raw = wb_raw.active
            self._log(f"工作表: {ws_raw.title}，{ws_raw.max_row} 行 × {ws_raw.max_column} 列")

            # --- 2. 定位表头行和数据起始行 ---
            # 表头在第3行，数据从第4行开始
            header_row = self._find_header_row(ws_raw)
            data_start_row = header_row + 1
            headers = []
            for col in range(1, ws_raw.max_column + 1):
                headers.append(str(ws_raw.cell(header_row, col).value or "").strip())
            self._log(f"表头行: 第{header_row}行, 表头: {headers}")

            # --- 3. 解析列索引（模糊匹配列名）---
            col_map = self._resolve_columns(headers, ws_raw.max_column)
            self._log(f"列映射: {col_map}")

            # --- 4. 提取标题时间（从第2行取时间）---
            title_time = self._extract_title_time(ws_raw, col_map)
            self._log(f"标题时间: {title_time}")
            # 把日期拼到文件名前缀
            if title_time:
                prefix = prefix.rstrip("_") + "_" + title_time

            # --- 5. 从数据起始行开始读取 ---
            self._log("读取数据行...")
            raw_rows = []
            skipped_filter = 0  # 根据过滤条件跳过的行数

            # 解析用户设置的过滤列索引
            filter_ci = None
            if filter_col and filter_val:
                for idx, h in enumerate(headers):
                    if h.strip() == filter_col:
                        filter_ci = idx + 1  # 1-based column index
                        break
                if filter_ci:
                    self._log(f"过滤条件: 列「{filter_col}」=「{filter_val}」时跳过")
                else:
                    self._log(f"[警告] 未找到列「{filter_col}」，过滤条件不生效")

            for row in range(data_start_row, ws_raw.max_row + 1):
                # 根据用户设置的过滤条件跳过
                if filter_ci is not None:
                    cell_val = str(ws_raw.cell(row, filter_ci).value or "").strip()
                    if cell_val == filter_val:
                        skipped_filter += 1
                        continue
                row_data = {}
                for key, ci in col_map.items():
                    if ci is None:
                        row_data[key] = ""
                    else:
                        val = ws_raw.cell(row, ci).value
                        row_data[key] = val
                # 只要标保列有值就算有效行
                if row_data.get("标保") is not None:
                    raw_rows.append(row_data)
            wb_raw.close()
            filter_log = f"（已排除 {filter_col}={filter_val} {skipped_filter} 行）" if skipped_filter else ""
            self._log(f"有效数据行: {len(raw_rows)}" + filter_log)

            # --- 6. 计算昨日日期 ---
            yesterday_str = self._calc_yesterday(title_time)
            self._log(f"昨日日期: {yesterday_str}")

            # --- 7. 按层级分组汇总（本月累计）---
            # 结构: district -> department -> agents -> sum(标保)
            tree = self._build_tree(raw_rows, date_filter=None)
            self._log(f"营业区数量: {len(tree)}")
            for dist_name, depts in tree.items():
                total_dept = len(depts)
                self._log(f"  [{dist_name}] {total_dept} 个营业部")

            # --- 8. 昨日标保+件数汇总（按业务员编码+姓名，保单号去重）---
            yesterday_agent_amt = {}   # key: (编码, 姓名) -> 累计标保
            yesterday_agent_cnt = {}   # key: (编码, 姓名) -> 去重保单数
            if yesterday_str:
                yesterday_tree = self._build_tree(raw_rows, date_filter=yesterday_str)
                # 保单号去重集合
                yesterday_policy_sets = defaultdict(set)
                for depts in yesterday_tree.values():
                    for agents in depts.values():
                        for a in agents:
                            k = (a["业务员编码"], a["业务员姓名"])
                            yesterday_agent_amt[k] = yesterday_agent_amt.get(k, 0.0) + a["标保"]
                            pn = a.get("保单号", "").strip()
                            if pn:
                                yesterday_policy_sets[k].add(pn)
                yesterday_agent_cnt = {k: len(v) for k, v in yesterday_policy_sets.items()}
                self._log(f"昨日有标保的业务员数: {len(yesterday_agent_amt)}")
                # 营业部级别昨日汇总
                yesterday_dept_amt = {}  # key: (营业区, 营业部) -> 累计标保
                yesterday_dept_cnt = {}  # key: (营业区, 营业部) -> 去重保单数
                dept_policy_sets = defaultdict(set)
                for dist_name, depts in yesterday_tree.items():
                    for dept_name, agents in depts.items():
                        dk = (dist_name, dept_name)
                        yesterday_dept_amt[dk] = sum(a["标保"] for a in agents)
                        for a in agents:
                            pn = a.get("保单号", "").strip()
                            if pn:
                                dept_policy_sets[dk].add(pn)
                yesterday_dept_cnt = {k: len(v) for k, v in dept_policy_sets.items()}
            else:
                yesterday_dept_amt = {}
                yesterday_dept_cnt = {}

            # --- 8.5. 人员清单补入 ---
            added_count = self._enrich_zero_agents(tree, file_personnel)
            self._log(f"从人员清单补入零业绩人员: {added_count} 人")

            # --- 8.6. 读取人员清单Sheet2月度目标 ---
            targets = self._read_targets_from_personnel(file_personnel)
            self._log(f"从人员清单Sheet2读取月度目标: {len(targets)} 个营业区")

            # --- 9. 生成报表 ---
            self._log("生成报表样式...")
            wb_out = self._generate_report(tree, title_time, yesterday_agent_amt, yesterday_agent_cnt, report_title, yesterday_dept_amt, yesterday_dept_cnt, targets=targets)

            # --- 10. 输出各格式文件 ---
            ext_map = {"Excel": ".xlsx", "PNG": ".png", "PDF": ".pdf"}

            # 确保输出目录存在
            try:
                os.makedirs(out_dir, exist_ok=True)
            except Exception as e:
                self._log(f"[错误] 创建输出目录失败: {e}")
                self.app.root.after(0, lambda: self.btn_start.configure(
                    state="normal", text="⚡ 开始处理"))
                return

            fd, temp_xlsx = tempfile.mkstemp(suffix=".xlsx", prefix="bbc_temp_")
            os.close(fd)
            try:
                wb_out.save(temp_xlsx)
                generated = []

                # PDF 和 PNG 合并为一次 Excel COM 调用
                need_pdf = "PDF" in selected
                need_png = "PNG" in selected
                pdf_path = self._make_output_path(out_dir, prefix, ".pdf") if need_pdf else None
                png_path = self._make_output_path(out_dir, prefix, ".png") if need_png else None

                if need_pdf or need_png:
                    self._export_via_com(temp_xlsx, pdf_path=pdf_path, img_path=png_path)
                    if need_pdf:
                        self._log(f"结果已保存: {pdf_path}")
                        generated.append(pdf_path)
                    if need_png:
                        self._log(f"结果已保存: {png_path}")
                        generated.append(png_path)

                # Excel 格式：直接复制
                if "Excel" in selected:
                    xlsx_path = self._make_output_path(out_dir, prefix, ".xlsx")
                    shutil.copy(temp_xlsx, xlsx_path)
                    self._log(f"结果已保存: {xlsx_path}")
                    generated.append(xlsx_path)
            finally:
                try:
                    os.remove(temp_xlsx)
                except OSError:
                    pass

            self._log("=" * 50)
            self._log("全部完成！")

            msg = "处理完成！\n\n结果已保存至:\n" + "\n".join(generated)
            messagebox.showinfo("处理完成", msg)

        except Exception as e:
            import traceback
            self._log(f"[错误] {e}")
            self._log(traceback.format_exc())
            messagebox.showerror("处理失败", f"发生错误:\n{e}")
        finally:
            self.app.root.after(0, lambda: self.btn_start.configure(state="normal", text="⚡ 开始处理"))

    # --------------- 列名解析 ---------------
    def _resolve_columns(self, headers, max_col):
        """根据表头文字匹配列索引（精确匹配，strip后比较）"""
        # 关键列名关键词（必须精确匹配）
        patterns = {
            "营业区名称": ["营业区名称"],
            "营业部名称": ["营业部名称"],
            "业务员编码": ["分成业务员代码"],
            "业务员姓名": ["分成业务员姓名"],
            "标保":     ["标保"],
            "时间":     ["统计时间"],
            "保单号":   ["分单号"],  # 用分单号去重
            "业务类型": ["业务类型"],
        }

        result = {}
        for key, keywords in patterns.items():
            found = None
            for kw in keywords:
                for idx, h in enumerate(headers):
                    col_idx = idx + 1
                    if h.strip() == kw:
                        found = col_idx
                        break
                if found:
                    break
            result[key] = found
        return result

    # --------------- 定位表头行 ---------------
    def _find_header_row(self, ws):
        """从第1行开始扫描，找到第一个包含'标保'或'保单号'或'分单号'的行作为表头行"""
        for r in range(1, ws.max_row + 1):
            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))
            if any(kw in row_text for kw in ["标保", "保单号", "分单号", "统计时间"]):
                return r
        return 1  # fallback

    # --------------- 提取标题时间 ---------------
    def _extract_title_time(self, ws, col_map):
        """从第2行提取标题时间（仅保留日期部分，去掉时分秒）"""
        title_time = ""
        # 尝试从第2行的每个非空单元格中提取
        for col in range(1, ws.max_column + 1):
            val = ws.cell(2, col).value
            if val is not None:
                # datetime 对象：直接格式化为日期
                if isinstance(val, datetime):
                    title_time = val.strftime("%Y-%m-%d")
                    break
                s = str(val).strip()
                if s and len(s) > 1:
                    # 包含"月"或"年"或日期格式的单元格
                    if any(c in s for c in ["月", "年", "-", "/"]):
                        title_time = s
                        break
        # 如果上面没找到，拼接所有非空的第2行内容
        if not title_time:
            parts = []
            for col in range(1, ws.max_column + 1):
                val = ws.cell(2, col).value
                if val is not None:
                    if isinstance(val, datetime):
                        parts.append(val.strftime("%Y-%m-%d"))
                        continue
                    s = str(val).strip()
                    if s:
                        parts.append(s)
            title_time = " ".join(parts)

        # 去掉时分秒部分（如 "2026-07-31 09:11:12" → "2026-07-31"）
        # 匹配 YYYY-MM-DD 或 YYYY/MM/DD 后面的时间部分
        title_time = re.sub(
            r"(\d{4}[-/]\d{1,2}[-/]\d{1,2})\s+\d{1,2}:\d{1,2}(:\d{1,2})?",
            r"\1", title_time)
        return title_time

    # --------------- 计算昨日日期 ---------------
    def _calc_yesterday(self, title_time):
        """从标题时间中解析日期，减1天，返回 YYYY-MM-DD 格式的昨日日期"""
        if not title_time:
            return None
        s = title_time.strip()

        # 尝试匹配 YYYY-MM-DD 或 YYYY/MM/DD
        for sep in ["-", "/"]:
            m = re.search(r"(\d{4})" + re.escape(sep) + r"(\d{1,2})" + re.escape(sep) + r"(\d{1,2})", s)
            if m:
                y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
                dt = datetime(y, mo, d)
                return (dt - timedelta(days=1)).strftime("%Y-%m-%d")

        # 尝试匹配 "2024年1月15日" 或 "1月15日" 等中文格式
        m = re.search(r"(\d{4})年(\d{1,2})月(\d{1,2})日", s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            dt = datetime(y, mo, d)
            return (dt - timedelta(days=1)).strftime("%Y-%m-%d")

        # 尝试只匹配年月，取当月1日
        m = re.search(r"(\d{4})年(\d{1,2})月", s)
        if m:
            y, mo = int(m.group(1)), int(m.group(2))
            dt = datetime(y, mo, 1)
            return (dt - timedelta(days=1)).strftime("%Y-%m-%d")

        # 尝试从单元格值中提取日期类型
        return None

    # --------------- 人员清单列名解析 ---------------
    def _resolve_personnel_columns(self, headers):
        """解析人员清单列名，返回 {营业区名称, 营业部名称, 业务员编码, 姓名} 的列索引"""
        patterns = {
            "营业区名称": ["营业区名称"],
            "营业部名称": ["营业部名称"],
            "业务员编码": ["业务员编码"],
            "姓名":       ["姓名", "业务员姓名"],
        }
        result = {}
        for key, keywords in patterns.items():
            found = None
            for kw in keywords:
                for idx, h in enumerate(headers):
                    if h.strip() == kw:
                        found = idx + 1  # 1-based
                        break
                if found:
                    break
            result[key] = found
        return result

    # --------------- 读取人员清单Sheet2月度目标 ---------------
    def _read_targets_from_personnel(self, personnel_path):
        """
        读取人员清单文件的第二个Sheet，提取各营业区的月度目标。
        Sheet2 表头需包含：营业区名称、月度目标
        返回 dict: {营业区名称: 月度目标(float)}
        """
        targets = {}
        try:
            wb = load_workbook(personnel_path, data_only=True)
            sheet_names = wb.sheetnames
            if len(sheet_names) < 2:
                self._log("[警告] 人员清单无第二个Sheet，月度目标全部为0")
                wb.close()
                return targets
            ws = wb[sheet_names[1]]  # 第二个Sheet
            # 读表头
            headers = {}
            for col in range(1, ws.max_column + 1):
                val = str(ws.cell(1, col).value or "").strip()
                if val in ("营业区名称", "月度目标"):
                    headers[val] = col
            if "营业区名称" not in headers or "月度目标" not in headers:
                self._log("[警告] 人员清单Sheet2未找到「营业区名称」或「月度目标」列，月度目标全部为0")
                wb.close()
                return targets
            col_dist = headers["营业区名称"]
            col_target = headers["月度目标"]
            for r in range(2, ws.max_row + 1):
                dist_name = str(ws.cell(r, col_dist).value or "").strip()
                if not dist_name:
                    continue
                try:
                    targets[dist_name] = float(ws.cell(r, col_target).value or 0.0)
                except (ValueError, TypeError):
                    targets[dist_name] = 0.0
            wb.close()
        except Exception as e:
            self._log(f"[警告] 读取人员清单Sheet2失败: {e}，月度目标全部为0")
        return targets

    # --------------- 人员清单补入 ---------------
    def _enrich_zero_agents(self, tree, personnel_path):
        """
        读取人员清单文件，对比 tree 中已有人员：
        - 存在 → 跳过
        - 不存在 → 按营业区/营业部层级追加到 tree，标保=0，保单号=""
        返回补入人数。
        """
        # 读取人员清单（固定取第一个Sheet，不受保存时激活Sheet影响）
        wb = load_workbook(personnel_path, data_only=True)
        ws = wb.worksheets[0]
        # 找表头行
        header_row = 1
        for r in range(1, min(ws.max_row + 1, 10)):
            row_text = " ".join(str(ws.cell(r, c).value or "") for c in range(1, ws.max_column + 1))
            if any(kw in row_text for kw in ["业务员编码", "营业区名称", "营业部名称"]):
                header_row = r
                break
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(str(ws.cell(header_row, col).value or "").strip())
        col_map = self._resolve_personnel_columns(headers)

        if col_map.get("业务员编码") is None:
            self._log("[警告] 人员清单中未找到「业务员编码」列，跳过补入")
            wb.close()
            return 0

        # 构建已有人员集合：key = (营业区名称, 营业部名称, 业务员编码)
        existing = set()
        for dist_name, depts in tree.items():
            for dept_name, agents in depts.items():
                for a in agents:
                    code = str(a.get("业务员编码", "")).strip()
                    existing.add((dist_name, dept_name, code))

        # 遍历人员清单
        added_count = 0
        data_start = header_row + 1
        for row in range(data_start, ws.max_row + 1):
            dist_name = str(ws.cell(row, col_map.get("营业区名称") or 0).value or "").strip() if col_map.get("营业区名称") else ""
            dept_name = str(ws.cell(row, col_map.get("营业部名称") or 0).value or "").strip() if col_map.get("营业部名称") else ""
            code = str(ws.cell(row, col_map.get("业务员编码") or 0).value or "").strip()
            name_val = str(ws.cell(row, col_map.get("姓名") or 0).value or "").strip()

            # 跳过无效行
            if not code or not dist_name:
                continue

            # 补默认值
            if not dept_name:
                dept_name = "(未分类)"
            if not name_val:
                name_val = code

            # 已在业绩数据中 → 跳过
            if (dist_name, dept_name, code) in existing:
                continue

            # 追加到 tree
            if dist_name not in tree:
                tree[dist_name] = {}
            if dept_name not in tree[dist_name]:
                tree[dist_name][dept_name] = []

            tree[dist_name][dept_name].append({
                "业务员编码": code,
                "业务员姓名": name_val,
                "标保": 0.0,
                "保单号": "",
            })
            added_count += 1

        wb.close()
        return added_count

    # --------------- 数据分组汇总 ---------------
    def _build_tree(self, rows, date_filter=None):
        """
        构建层级分组树：
          tree[营业区][营业部] = [ {业务员编码, 业务员姓名, 标保}, ... ]

        date_filter: None=全部数据, "YYYY-MM-DD"=只匹配该日期的行
        """
        tree = defaultdict(lambda: defaultdict(list))
        for r in rows:
            # 日期过滤
            if date_filter:
                row_date = self._extract_date_from_row(r.get("时间", ""))
                if row_date != date_filter:
                    continue

            dist = str(r.get("营业区名称", "") or "").strip()
            dept = str(r.get("营业部名称", "") or "").strip()
            code = str(r.get("业务员编码", "") or "").strip()
            name = str(r.get("业务员姓名", "") or "").strip()
            amt  = self._safe_number(r.get("标保", 0))
            policy_no = str(r.get("保单号", "") or "").strip()

            if not dist:
                dist = "(未分类)"
            if not dept:
                dept = "(未分类)"

            tree[dist][dept].append({
                "业务员编码": code,
                "业务员姓名": name,
                "标保": amt,
                "保单号": policy_no,
            })
        return tree

    def _extract_date_from_row(self, time_val):
        """从行数据的时间列值中提取 YYYY-MM-DD 格式日期"""
        if not time_val:
            return ""
        # 如果本身是 datetime 对象
        if isinstance(time_val, datetime):
            return time_val.strftime("%Y-%m-%d")
        s = str(time_val).strip()

        # 尝试多种日期格式
        for fmt in ["%Y-%m-%d", "%Y/%m/%d", "%Y年%m月%d日", "%Y-%m-%d %H:%M:%S", "%Y/%m/%d %H:%M:%S"]:
            try:
                dt = datetime.strptime(s[:len(fmt)], fmt)
                return dt.strftime("%Y-%m-%d")
            except ValueError:
                continue

        # 正则兜底
        m = re.search(r"(\d{4})[-/年](\d{1,2})[-/月](\d{1,2})", s)
        if m:
            y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
            return datetime(y, mo, d).strftime("%Y-%m-%d")
        return ""

    def _safe_number(self, v):
        """安全转换为浮点数"""
        try:
            return float(v) if v is not None else 0.0
        except (ValueError, TypeError):
            return 0.0

    # --------------- 生成报表 ---------------
    # 报表列映射（12 列结构）
    COL_A = 1   # 营业区
    COL_B = 2   # 月度目标
    COL_C = 3   # 营业区昨日累计业绩
    COL_D = 4   # 营业区本月累计业绩
    COL_E = 5   # 月度达成率
    COL_F = 6   # 营业部
    COL_G = 7   # 营业部昨日出单业绩
    COL_H = 8   # 营业部本月出单业绩
    COL_I = 9   # 业务员编码
    COL_J = 10  # 姓名
    COL_K = 11  # 昨日出单业绩
    COL_L = 12  # 本月累计业绩

    def _generate_report(self, tree, title_time, yesterday_amt=None, yesterday_cnt=None, report_title="BBC业绩报表", yesterday_dept_amt=None, yesterday_dept_cnt=None, targets=None):
        """
        生成业绩报表。
        报表结构（12列）：
          A=营业区  B=月度目标  C=营业区昨日累计业绩  D=营业区本月累计业绩  E=月度达成率  F=营业部
          G=营业部昨日出单业绩  H=营业部本月出单业绩
          I=业务员编码  J=姓名
          K=昨日出单业绩  L=本月累计业绩
        """
        if yesterday_amt is None:
            yesterday_amt = {}
        if yesterday_cnt is None:
            yesterday_cnt = {}
        if yesterday_dept_amt is None:
            yesterday_dept_amt = {}
        if yesterday_dept_cnt is None:
            yesterday_dept_cnt = {}
        if targets is None:
            targets = {}
        import copy
        from openpyxl.utils import get_column_letter

        # ---- 创建工作簿（不再依赖模板文件）----
        wb = Workbook()
        ws = wb.active

        # ---- 列宽定义（标题横幅 + 实际列宽共用）----
        col_widths = {
            self.COL_A: 12,  # 营业区名称
            self.COL_B: 16,  # 月度目标
            self.COL_C: 22,  # 营业区昨日累计业绩
            self.COL_D: 22,  # 营业区本月累计业绩
            self.COL_E: 16,  # 月度达成率
            self.COL_F: 14,  # 营业部名称
            self.COL_G: 22,  # 营业部昨日出单业绩
            self.COL_H: 22,  # 营业部本月出单业绩
            self.COL_I: 14,  # 业务员编码
            self.COL_J: 10,  # 姓名
            self.COL_K: 18,  # 昨日出单业绩
            self.COL_L: 20,  # 本月累计业绩
        }

        # ---- 表头文字（写死为常量，不再从模板读取）----
        header_values = [
            "营业区名称",          # A
            "月度目标",            # B
            "营业区昨日累计业绩",   # C
            "营业区本月累计业绩",   # D
            "月度达成率",          # E
            "营业部名称",          # F
            "营业部昨日出单业绩",   # G
            "营业部本月出单业绩",   # H
            "业务员编码",          # I
            "姓名",               # J
            "昨日出单业绩",        # K
            "本月累计业绩",        # L
        ]

        # ---- 基础样式（写死为常量，不再从模板读取）----
        ref_font = Font(name="微软雅黑", size=9, bold=False)
        ref_align = Alignment(horizontal="center", vertical="center")
        ref_border = Border()
        ref_fill = PatternFill()

        # 暖色系边框和背景
        gold_side = Side(style="thin", color="F4B183")
        thin_border = Border(
            left=gold_side,
            right=gold_side,
            top=gold_side,
            bottom=gold_side,
        )
        cream_fill = PatternFill(start_color="FFFDF5", end_color="FFFDF5", fill_type="solid")
        header_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")

        # 营业部名称列底色（与其他列保持一致）

        center_align = Alignment(horizontal="center", vertical="center")
        left_align   = Alignment(horizontal="left", vertical="center")

        # 合计行样式
        total_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
        total_font = Font(name="微软雅黑", size=9, bold=True)

        # 营业部分隔线
        thick_gold_side = Side(style="medium", color="F4B183")
        thick_bottom_border = Border(
            left=gold_side,
            right=gold_side,
            top=gold_side,
            bottom=thick_gold_side,
        )

        # ---- 行结构：第1行标题 / 第2行时间 / 第3行表头 / 第4行起数据 ----
        # （从空白 Workbook 创建，无需清理模板数据）

        # 把表头写到第3行
        for c in range(1, 13):
            ws.cell(3, c).value = header_values[c - 1]

        # ---- 提前设置列宽和标题行高 ----
        for col, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col)].width = width
        ws.row_dimensions[1].height = 40  # 标题行
        ws.row_dimensions[2].height = 22  # 时间行
        ws.row_dimensions[3].height = 32  # 表头行（加高以容纳文字）

        # ---- 更新标题（第1行主标题 + 第2行时间）----
        if title_time:
            title_text = report_title
            from openpyxl.styles import Font as _Font, Alignment as _Align, PatternFill as _Fill
            title_font = _Font(name="微软雅黑", size=18, bold=True, color="FFFFFF")
            time_font = _Font(name="微软雅黑", size=11, bold=False, color="FFFFFF")
            title_align = _Align(horizontal="center", vertical="bottom")
            time_align = _Align(horizontal="right", vertical="bottom", indent=1)
            title_fill = _Fill(start_color="F5A623", end_color="F5A623", fill_type="solid")
            # 第1行：主标题（合并 A1:L1，居中，亮橙底白字）
            for c in range(1, 13):
                cell = ws.cell(1, c)
                cell.fill = title_fill
                cell.alignment = title_align  # 先设置 alignment（合并前）
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=12)
            ws.cell(1, 1).value = title_text
            ws.cell(1, 1).font = title_font
            # 第2行：时间（合并 A2:L2，右对齐，亮橙底白字，无边框）
            from openpyxl.styles import Border as _Border
            no_border = _Border()  # 空边框，清除模板残留的金色边框
            for c in range(1, 13):
                cell = ws.cell(2, c)
                cell.fill = title_fill
                cell.alignment = time_align  # 先设置 alignment（合并前）
                cell.border = no_border
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=12)
            ws.cell(2, 1).value = title_time
            ws.cell(2, 1).font = time_font
            ws.cell(2, 1).border = no_border

        # ---- 保持默认网格线显示 ----

        # ---- 从第4行开始填充数据（第1行标题/第2行时间/第3行表头）----
        row = 4

        for dist_name, departments in sorted(tree.items()):
            dist_start_row = row

            dept_items = sorted(departments.items())
            for dept_idx, (dept_name, agents) in enumerate(dept_items):
                dept_start_row = row
                agents_sorted = sorted(agents, key=lambda a: a["业务员姓名"])

                # 按业务员编码+姓名汇总：同一个业务员可能有多条记录
                agent_summary = {}  # key: (编码, 姓名) -> {"标保": float, "昨日标保": float, "本月件数": int, "昨日件数": int}
                agent_policy_sets = defaultdict(set)  # 用于本月保单号去重
                for agent in agents_sorted:
                    code = agent["业务员编码"]
                    name = agent["业务员姓名"]
                    k = (code, name)
                    if k not in agent_summary:
                        agent_summary[k] = {
                            "标保": 0.0,
                            "昨日标保": yesterday_amt.get(k, 0.0),
                            "本月件数": 0,
                            "昨日件数": yesterday_cnt.get(k, 0),
                        }
                    agent_summary[k]["标保"] += agent["标保"]
                    pn = agent.get("保单号", "").strip()
                    if pn:
                        agent_policy_sets[k].add(pn)
                # 本月件数 = 去重保单数量
                for k in agent_summary:
                    agent_summary[k]["本月件数"] = len(agent_policy_sets.get(k, set()))

                for (code, name), summary in sorted(agent_summary.items(), key=lambda x: x[0][1]):
                    amt = summary["标保"]
                    y_amt = summary["昨日标保"]
                    month_cnt = summary["本月件数"]
                    yesterday_cnt_val = summary["昨日件数"]

                    # I: 业务员编码
                    c = ws.cell(row, self.COL_I, code)
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.border = thin_border

                    # J: 姓名
                    c = ws.cell(row, self.COL_J, name)
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.border = thin_border

                    # K: 昨日出单业绩
                    c = ws.cell(row, self.COL_K, float(y_amt or 0.0))
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.number_format = '#,##0.00'
                    c.border = thin_border

                    # L: 本月累计业绩
                    c = ws.cell(row, self.COL_L, float(amt or 0.0))
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.number_format = '#,##0.00'
                    c.border = thin_border

                    row += 1

                dept_end_row = row - 1

                # --- 合并营业部列 (F列) ---
                if dept_end_row >= dept_start_row:
                    # 先给合并范围内所有单元格设置背景色和金黄色边框，再合并
                    for rr in range(dept_start_row, dept_end_row + 1):
                        ws.cell(rr, self.COL_F).fill = cream_fill
                        ws.cell(rr, self.COL_F).border = thin_border

                    ws.merge_cells(
                        start_row=dept_start_row, start_column=self.COL_F,
                        end_row=dept_end_row, end_column=self.COL_F
                    )

                    # 合并后给左上角设置值和字体/对齐样式
                    c = ws.cell(dept_start_row, self.COL_F, dept_name)
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.border = thin_border
                    c.fill = cream_fill

                # --- 合并营业部昨日出单业绩列 (G列) ---
                if dept_end_row >= dept_start_row:
                    dk = (dist_name, dept_name)
                    dept_y_amt = float(yesterday_dept_amt.get(dk, 0.0) or 0.0)
                    for rr in range(dept_start_row, dept_end_row + 1):
                        ws.cell(rr, self.COL_G).fill = cream_fill
                        ws.cell(rr, self.COL_G).border = thin_border
                    ws.merge_cells(
                        start_row=dept_start_row, start_column=self.COL_G,
                        end_row=dept_end_row, end_column=self.COL_G
                    )
                    c = ws.cell(dept_start_row, self.COL_G, dept_y_amt)
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.number_format = '#,##0.00'
                    c.border = thin_border
                    c.fill = cream_fill

                # --- 合并营业部本月出单业绩列 (H列) ---
                if dept_end_row >= dept_start_row:
                    dk = (dist_name, dept_name)
                    dept_m_amt = sum(a["标保"] for a in agents)
                    for rr in range(dept_start_row, dept_end_row + 1):
                        ws.cell(rr, self.COL_H).fill = cream_fill
                        ws.cell(rr, self.COL_H).border = thin_border
                    ws.merge_cells(
                        start_row=dept_start_row, start_column=self.COL_H,
                        end_row=dept_end_row, end_column=self.COL_H
                    )
                    c = ws.cell(dept_start_row, self.COL_H, float(dept_m_amt or 0.0))
                    c.font = copy.copy(ref_font)
                    c.alignment = center_align
                    c.number_format = '#,##0.00'
                    c.border = thin_border
                    c.fill = cream_fill

                # --- 营业部之间粗分隔线（最后一个营业部不加） ---
                is_last_dept = (dept_idx == len(dept_items) - 1)
                if not is_last_dept:
                    for cc in range(self.COL_A, self.COL_L + 1):
                        ws.cell(dept_end_row, cc).border = thick_bottom_border



            dist_end_row = row - 1

            # 计算营业区级别汇总（用于D/E列）
            # E列：本月所有行标保累加
            dist_total_amt = sum(agent["标保"] for agents in departments.values() for agent in agents)
            # D列：本区业务员昨日标保汇总（先收集本区唯一业务员，再查昨天数据）
            agent_keys = set()
            for agents in departments.values():
                for agent in agents:
                    agent_keys.add((agent["业务员编码"], agent["业务员姓名"]))
            dist_yesterday_amt = sum(yesterday_amt.get(k, 0.0) for k in agent_keys)
            # 获取本营业区的月度目标
            dist_target = float(targets.get(dist_name, 0.0) or 0.0)

            # --- 合并营业区列 (A列) ---
            if dist_end_row >= dist_start_row:
                # 先给合并范围内所有单元格设置背景色和金黄色边框，再合并
                for rr in range(dist_start_row, dist_end_row + 1):
                    ws.cell(rr, self.COL_A).fill = cream_fill
                    ws.cell(rr, self.COL_A).border = thin_border
                ws.merge_cells(
                    start_row=dist_start_row, start_column=self.COL_A,
                    end_row=dist_end_row, end_column=self.COL_A
                )
                c = ws.cell(dist_start_row, self.COL_A, dist_name)
                c.font = copy.copy(ref_font)
                c.alignment = center_align
                c.border = thin_border
                c.fill = cream_fill

            # --- 合并营业区级别列 (B列: 月度目标, C列: 昨日累计业绩, D列: 本月累计业绩, E列: 月度达成率) ---
            if dist_end_row >= dist_start_row:
                # B列：月度目标
                for rr in range(dist_start_row, dist_end_row + 1):
                    ws.cell(rr, self.COL_B).fill = cream_fill
                    ws.cell(rr, self.COL_B).border = thin_border
                ws.merge_cells(
                    start_row=dist_start_row, start_column=self.COL_B,
                    end_row=dist_end_row, end_column=self.COL_B
                )
                c = ws.cell(dist_start_row, self.COL_B, dist_target)
                c.font = copy.copy(ref_font)
                c.alignment = center_align
                c.number_format = '#,##0.00'
                c.border = thin_border
                c.fill = cream_fill

                # C列：营业区昨日累计业绩（业绩金额）
                for rr in range(dist_start_row, dist_end_row + 1):
                    ws.cell(rr, self.COL_C).fill = cream_fill
                    ws.cell(rr, self.COL_C).border = thin_border
                ws.merge_cells(
                    start_row=dist_start_row, start_column=self.COL_C,
                    end_row=dist_end_row, end_column=self.COL_C
                )
                c = ws.cell(dist_start_row, self.COL_C, float(dist_yesterday_amt or 0.0))
                c.font = copy.copy(ref_font)
                c.alignment = center_align
                c.number_format = '#,##0.00'
                c.border = thin_border
                c.fill = cream_fill

                # D列：营业区本月累计业绩（业绩金额）
                for rr in range(dist_start_row, dist_end_row + 1):
                    ws.cell(rr, self.COL_D).fill = cream_fill
                    ws.cell(rr, self.COL_D).border = thin_border
                ws.merge_cells(
                    start_row=dist_start_row, start_column=self.COL_D,
                    end_row=dist_end_row, end_column=self.COL_D
                )
                c = ws.cell(dist_start_row, self.COL_D, dist_total_amt)
                c.font = copy.copy(ref_font)
                c.alignment = center_align
                c.number_format = '#,##0.00'
                c.border = thin_border
                c.fill = cream_fill

                # E列：月度达成率 = 本月累计业绩 / 月度目标 × 100%
                dist_rate = (float(dist_total_amt or 0.0) / dist_target * 100) if dist_target else 0.0
                for rr in range(dist_start_row, dist_end_row + 1):
                    ws.cell(rr, self.COL_E).fill = cream_fill
                    ws.cell(rr, self.COL_E).border = thin_border
                ws.merge_cells(
                    start_row=dist_start_row, start_column=self.COL_E,
                    end_row=dist_end_row, end_column=self.COL_E
                )
                c = ws.cell(dist_start_row, self.COL_E, round(dist_rate, 2))
                c.font = copy.copy(ref_font)
                c.alignment = center_align
                c.number_format = '0.00"%"'
                c.border = thin_border
                c.fill = cream_fill

            row += 2  # 两个空行，区分不同营业区

        # ---- 统一应用暖米色背景到整个数据区域 ----
        last_data_row = row - 2  # 减去两个空行
        if last_data_row >= 4:
            # 表头行（第3行）
            for c in range(1, 13):
                cell = ws.cell(3, c)
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = thin_border
            # 数据行（第4行起）
            for r in range(4, last_data_row + 1):
                for c in range(1, 13):
                    ws.cell(r, c).fill = cream_fill

        # 列宽和行高已在标题区域之前设置，此处无需重复

        return wb
