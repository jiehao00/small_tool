import tkinter as tk
from tkinter import filedialog, messagebox, ttk
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

COLUMN_MAP = {
    '数电发票号码': '发票号',
    '销售方纳税人名称': '销售方纳税人名称',
    '开票日期*': '开票时间',
    '票种*': '发票票种',
    '金额*': '不含税金额',
    '票面税额*': '税额',
}

DETAIL_COLUMN_MAP = {
    '发票号码': '发票号',
    '开票日期': '开票时间',
    '实物发票种类': '_ticket_kind',
    '发票类型': '_ticket_type',
    '不含税金额': '不含税金额',
    '税额': '税额',
}

OUTPUT_HEADERS = ['序号', '发票号', '销售方纳税人名称', '开票时间', '发票票种', '不含税金额', '税额', '合计数（不含税金额+税额）']


class FinanceConverter:
    NAV_BG = '#2c3e50'
    NAV_ACTIVE_BG = '#3498db'
    NAV_FG = '#ecf0f1'
    NAV_WIDTH = 140

    def __init__(self):
        self.window = tk.Tk()
        self.window.title('进项数据采集票种校验工具')
        self.window.geometry('900x720')
        self.window.minsize(820, 600)
        self.window.lift()
        self.window.attributes('-topmost', True)
        self.window.after(200, lambda: self.window.attributes('-topmost', False))
        self.window.focus_force()
        self._setup_ui()

    def _setup_ui(self):
        # 顶部标题栏
        title_bar = tk.Frame(self.window, bg=self.NAV_BG, height=50)
        title_bar.pack(fill='x')
        title_bar.pack_propagate(False)
        tk.Label(title_bar, text='进项数据采集票种校验工具', font=('Microsoft YaHei', 16, 'bold'),
                 bg=self.NAV_BG, fg='white').pack(side='left', padx=20, pady=10)

        # 主区域：左侧导航 + 右侧内容
        main = tk.Frame(self.window)
        main.pack(fill='both', expand=True)

        # === 左侧导航栏 ===
        self.nav_frame = tk.Frame(main, bg=self.NAV_BG, width=self.NAV_WIDTH)
        self.nav_frame.pack(side='left', fill='y')
        self.nav_frame.pack_propagate(False)

        self.nav_btns = {}
        nav_items = [
            ('merge', '合并数据'),
            ('usage', '发票用途'),
            ('compare', '税额对比'),
            ('convert', '一键核账'),
            ('guide', '使用说明'),
            ('about', '关于'),
        ]
        for key, label in nav_items:
            btn = tk.Button(self.nav_frame, text=label, font=('Microsoft YaHei', 11),
                            bg=self.NAV_BG, fg=self.NAV_FG, bd=0, cursor='hand2',
                            activebackground=self.NAV_ACTIVE_BG, activeforeground='white',
                            anchor='w', padx=18, pady=12)
            btn.pack(fill='x')
            self.nav_btns[key] = btn

        # === 右侧内容区域 ===
        self.content_area = tk.Frame(main, bg='#f0f0f0')
        self.content_area.pack(side='left', fill='both', expand=True)

        self.pages = {}
        self._setup_merge_page()
        self._setup_usage_page()
        self._setup_compare_page()
        self._setup_convert_page()
        self._setup_guide_page()
        self._setup_about_page()

        # 绑定导航事件
        self.nav_btns['merge'].config(command=lambda: self._switch_page('merge'))
        self.nav_btns['usage'].config(command=lambda: self._switch_page('usage'))
        self.nav_btns['compare'].config(command=lambda: self._switch_page('compare'))
        self.nav_btns['convert'].config(command=lambda: self._switch_page('convert'))
        self.nav_btns['guide'].config(command=lambda: self._switch_page('guide'))
        self.nav_btns['about'].config(command=lambda: self._switch_page('about'))

        # 默认选中合并数据
        self._switch_page('merge')

    def _switch_page(self, page_key):
        """切换页面"""
        for key, btn in self.nav_btns.items():
            if key == page_key:
                btn.config(bg=self.NAV_ACTIVE_BG, fg='white')
            else:
                btn.config(bg=self.NAV_BG, fg=self.NAV_FG)
        for key, frame in self.pages.items():
            if key == page_key:
                frame.pack(fill='both', expand=True)
            else:
                frame.pack_forget()

    def _setup_merge_page(self):
        """合并数据页面"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['merge'] = page

        # === 汇总文件 ===
        hui_frame = tk.LabelFrame(page, text='税局进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        hui_frame.pack(fill='x', pady=(0, 10))

        hui_row = tk.Frame(hui_frame)
        hui_row.pack(fill='x')
        self.hui_path_var = tk.StringVar()
        tk.Entry(hui_row, textvariable=self.hui_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(hui_row, text='浏览...', command=self._browse_hui, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 明细文件 ===
        ming_frame = tk.LabelFrame(page, text='税务管理平台进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        ming_frame.pack(fill='x', pady=(0, 10))

        ming_row = tk.Frame(ming_frame)
        ming_row.pack(fill='x')
        self.ming_path_var = tk.StringVar()
        tk.Entry(ming_row, textvariable=self.ming_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(ming_row, text='浏览...', command=self._browse_ming, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 输出目录 ===
        out_frame = tk.LabelFrame(page, text='输出目录', font=('Microsoft YaHei', 10), padx=12, pady=12)
        out_frame.pack(fill='x', pady=(0, 10))

        out_row = tk.Frame(out_frame)
        out_row.pack(fill='x')
        self.output_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(out_row, textvariable=self.output_dir_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(out_row, text='浏览...', command=self._browse_output, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 执行按钮 ===
        btn_frame = tk.Frame(page)
        btn_frame.pack(fill='x', pady=(6, 10))
        self.run_btn = tk.Button(btn_frame, text='开始转换', command=self._run_conversion, font=('Microsoft YaHei', 11, 'bold'), bg='#2196F3', fg='white', width=18, height=2, state='disabled')
        self.run_btn.pack()

        # === 进度条 ===
        self.progress = ttk.Progressbar(page, mode='determinate')
        self.progress.pack(fill='x', pady=(0, 6))

        # === 日志 ===
        log_frame = tk.LabelFrame(page, text='处理日志', font=('Microsoft YaHei', 10), padx=8, pady=8)
        log_frame.pack(fill='both', expand=True)
        self.log_text = tk.Text(log_frame, height=6, font=('Consolas', 9), wrap='word', state='disabled')
        self.log_text.pack(fill='both', expand=True)
        scrollbar = tk.Scrollbar(self.log_text)
        scrollbar.pack(side='right', fill='y')
        self.log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.log_text.yview)

        self.hui_path_var.trace_add('write', lambda *a: self._on_path_change())
        self.ming_path_var.trace_add('write', lambda *a: self._on_path_change())

    def _setup_usage_page(self):
        """发票用途页面"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['usage'] = page

        # === 汇总文件 ===
        hui_frame = tk.LabelFrame(page, text='财务汇总转换结果(合并数据的结果表)', font=('Microsoft YaHei', 10), padx=12, pady=12)
        hui_frame.pack(fill='x', pady=(0, 10))

        hui_row = tk.Frame(hui_frame)
        hui_row.pack(fill='x')
        self.usage_hui_path_var = tk.StringVar()
        tk.Entry(hui_row, textvariable=self.usage_hui_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(hui_row, text='浏览...', command=self._browse_usage_hui, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 明细文件 ===
        ming_frame = tk.LabelFrame(page, text='税务管理平台进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        ming_frame.pack(fill='x', pady=(0, 10))

        ming_row = tk.Frame(ming_frame)
        ming_row.pack(fill='x')
        self.usage_ming_path_var = tk.StringVar()
        tk.Entry(ming_row, textvariable=self.usage_ming_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(ming_row, text='浏览...', command=self._browse_usage_ming, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 输出目录 ===
        out_frame = tk.LabelFrame(page, text='输出目录', font=('Microsoft YaHei', 10), padx=12, pady=12)
        out_frame.pack(fill='x', pady=(0, 10))

        out_row = tk.Frame(out_frame)
        out_row.pack(fill='x')
        self.usage_output_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(out_row, textvariable=self.usage_output_dir_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(out_row, text='浏览...', command=self._browse_usage_output, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 执行按钮 ===
        btn_frame = tk.Frame(page)
        btn_frame.pack(fill='x', pady=(6, 10))
        self.usage_run_btn = tk.Button(btn_frame, text='新增发票用途', command=self._run_usage, font=('Microsoft YaHei', 11, 'bold'), bg='#27AE60', fg='white', width=18, height=2, state='disabled')
        self.usage_run_btn.pack()

        # === 进度条 ===
        self.usage_progress = ttk.Progressbar(page, mode='determinate')
        self.usage_progress.pack(fill='x', pady=(0, 6))

        # === 日志 ===
        log_frame = tk.LabelFrame(page, text='处理日志', font=('Microsoft YaHei', 10), padx=8, pady=8)
        log_frame.pack(fill='both', expand=True)
        self.usage_log_text = tk.Text(log_frame, height=6, font=('Consolas', 9), wrap='word', state='disabled')
        self.usage_log_text.pack(fill='both', expand=True)
        scrollbar = tk.Scrollbar(self.usage_log_text)
        scrollbar.pack(side='right', fill='y')
        self.usage_log_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.usage_log_text.yview)

        self.usage_hui_path_var.trace_add('write', lambda *a: self._on_usage_path_change())
        self.usage_ming_path_var.trace_add('write', lambda *a: self._on_usage_path_change())

    def _setup_compare_page(self):
        """税额对比页面"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['compare'] = page

        # === 含发票用途文件 ===
        f1 = tk.LabelFrame(page, text='财务汇总_含发票用途文件', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f1.pack(fill='x', pady=(0, 10))
        r1 = tk.Frame(f1)
        r1.pack(fill='x')
        self.cmp_usage_path_var = tk.StringVar()
        tk.Entry(r1, textvariable=self.cmp_usage_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r1, text='浏览...', command=self._browse_cmp_usage, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 财务汇总文件 ===
        f2 = tk.LabelFrame(page, text='税局进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f2.pack(fill='x', pady=(0, 10))
        r2 = tk.Frame(f2)
        r2.pack(fill='x')
        self.cmp_hui_path_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.cmp_hui_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r2, text='浏览...', command=self._browse_cmp_hui, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 明细文件 ===
        f3 = tk.LabelFrame(page, text='税务管理平台进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f3.pack(fill='x', pady=(0, 10))
        r3 = tk.Frame(f3)
        r3.pack(fill='x')
        self.cmp_ming_path_var = tk.StringVar()
        tk.Entry(r3, textvariable=self.cmp_ming_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r3, text='浏览...', command=self._browse_cmp_ming, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 输出目录 ===
        f4 = tk.LabelFrame(page, text='输出目录', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f4.pack(fill='x', pady=(0, 10))
        r4 = tk.Frame(f4)
        r4.pack(fill='x')
        self.cmp_output_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(r4, textvariable=self.cmp_output_dir_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r4, text='浏览...', command=self._browse_cmp_output, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 对比按钮 ===
        bf = tk.Frame(page)
        bf.pack(fill='x', pady=(6, 10))
        self.cmp_run_btn = tk.Button(bf, text='开始对比', command=self._run_compare, font=('Microsoft YaHei', 11, 'bold'), bg='#E67E22', fg='white', width=18, height=2, state='disabled')
        self.cmp_run_btn.pack()

        # === 进度条 ===
        self.cmp_progress = ttk.Progressbar(page, mode='determinate')
        self.cmp_progress.pack(fill='x', pady=(0, 6))

        # === 日志 ===
        lf = tk.LabelFrame(page, text='处理日志', font=('Microsoft YaHei', 10), padx=8, pady=8)
        lf.pack(fill='both', expand=True)
        self.cmp_log_text = tk.Text(lf, height=6, font=('Consolas', 9), wrap='word', state='disabled')
        self.cmp_log_text.pack(fill='both', expand=True)
        sb = tk.Scrollbar(self.cmp_log_text)
        sb.pack(side='right', fill='y')
        self.cmp_log_text.config(yscrollcommand=sb.set)
        sb.config(command=self.cmp_log_text.yview)

        self.cmp_usage_path_var.trace_add('write', lambda *a: self._on_cmp_path_change())
        self.cmp_hui_path_var.trace_add('write', lambda *a: self._on_cmp_path_change())
        self.cmp_ming_path_var.trace_add('write', lambda *a: self._on_cmp_path_change())

    def _setup_convert_page(self):
        """财务转换页面（一键串联：合并数据 → 发票用途 → 税额对比）"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['convert'] = page

        # 提示说明
        tip_frame = tk.Frame(page, bg='#FFF3CD', padx=10, pady=6)
        tip_frame.pack(fill='x', pady=(0, 10))
        tk.Label(tip_frame, text='一键完成：合并数据 → 发票用途 → 税额对比，直接输出最终对比结果文件',
                 font=('Microsoft YaHei', 9), fg='#856404', bg='#FFF3CD').pack()

        # === 税局进项明细表 ===
        f1 = tk.LabelFrame(page, text='税局进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f1.pack(fill='x', pady=(0, 10))
        r1 = tk.Frame(f1)
        r1.pack(fill='x')
        self.cvt_hui_path_var = tk.StringVar()
        tk.Entry(r1, textvariable=self.cvt_hui_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r1, text='浏览...', command=self._browse_cvt_hui, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 税务管理平台进项明细表 ===
        f2 = tk.LabelFrame(page, text='税务管理平台进项明细表', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f2.pack(fill='x', pady=(0, 10))
        r2 = tk.Frame(f2)
        r2.pack(fill='x')
        self.cvt_ming_path_var = tk.StringVar()
        tk.Entry(r2, textvariable=self.cvt_ming_path_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r2, text='浏览...', command=self._browse_cvt_ming, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 输出目录 ===
        f3 = tk.LabelFrame(page, text='输出目录', font=('Microsoft YaHei', 10), padx=12, pady=12)
        f3.pack(fill='x', pady=(0, 10))
        r3 = tk.Frame(f3)
        r3.pack(fill='x')
        self.cvt_output_dir_var = tk.StringVar(value=os.getcwd())
        tk.Entry(r3, textvariable=self.cvt_output_dir_var, font=('Microsoft YaHei', 9), state='readonly').pack(side='left', fill='x', expand=True, padx=(0, 8))
        tk.Button(r3, text='浏览...', command=self._browse_cvt_output, font=('Microsoft YaHei', 9), width=8).pack(side='right')

        # === 执行按钮 ===
        bf = tk.Frame(page)
        bf.pack(fill='x', pady=(6, 10))
        self.cvt_run_btn = tk.Button(bf, text='一键转换', command=self._run_convert, font=('Microsoft YaHei', 11, 'bold'), bg='#8E44AD', fg='white', width=18, height=2, state='disabled')
        self.cvt_run_btn.pack()

        # === 进度条 ===
        self.cvt_progress = ttk.Progressbar(page, mode='determinate')
        self.cvt_progress.pack(fill='x', pady=(0, 6))

        # === 日志 ===
        lf = tk.LabelFrame(page, text='处理日志', font=('Microsoft YaHei', 10), padx=8, pady=8)
        lf.pack(fill='both', expand=True)
        self.cvt_log_text = tk.Text(lf, height=6, font=('Consolas', 9), wrap='word', state='disabled')
        self.cvt_log_text.pack(fill='both', expand=True)
        sb = tk.Scrollbar(self.cvt_log_text)
        sb.pack(side='right', fill='y')
        self.cvt_log_text.config(yscrollcommand=sb.set)
        sb.config(command=self.cvt_log_text.yview)

        self.cvt_hui_path_var.trace_add('write', lambda *a: self._on_cvt_path_change())
        self.cvt_ming_path_var.trace_add('write', lambda *a: self._on_cvt_path_change())

    def _setup_guide_page(self):
        """使用说明页面"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['guide'] = page

        title = tk.Label(page, text='使用说明', font=('Microsoft YaHei', 14, 'bold'))
        title.pack(pady=(0, 15))

        # 可滚动的文本框
        tf = tk.Frame(page)
        tf.pack(fill='both', expand=True)
        text = tk.Text(tf, font=('Microsoft YaHei', 10), wrap='word', padx=15, pady=10, spacing1=4, spacing3=4)
        sb = tk.Scrollbar(tf, command=text.yview)
        text.config(yscrollcommand=sb.set)
        sb.pack(side='right', fill='y')
        text.pack(side='left', fill='both', expand=True)

        guide_content = """一、软件概述
本工具用于进项数据的处理和对比，提供四个核心功能模块：
合并数据、发票用途、税额对比、一键核账。

二、合并数据（功能一）
1. 导入"税局进项明细表"（必选）
2. 导入"税务管理平台进项明细表"（必选）
3. 选择输出目录后，点击"开始转换"
4. 生成标准化格式的汇总结果文件

【税局进项明细表 — 字段映射关系】
  原始字段名                      →   输出字段名
  ─────────────────────────────────────────────
  数电发票号码                   →   发票号
  销售方纳税人名称               →   销售方纳税人名称
  开票日期*                      →   开票时间
  票种*                          →   发票票种
  金额*                          →   不含税金额
  票面税额*                      →   税额

【税务管理平台进项明细表 — 字段映射关系】（导入时生效）
  原始字段名                      →   输出字段名
  ─────────────────────────────────────────────
  发票号码                       →   发票号（用于匹配）
  实物发票种类                   →   销售方纳税人名称
  开票日期                       →   开票时间
  发票类型                       →   发票票种
  不含税金额                     →   不含税金额
  税额                           →   税额

【匹配规则】
- 用税局进项明细表的"数电发票号码"与税务管理平台进项明细表的"发票号码"进行匹配
- 以税局进项明细表数据为主表，将匹配到的明细数据补充到对应行
- "合计数（不含税金额+税额）"列由系统自动计算并写入

三、发票用途（功能二）
1. 导入"财务汇总转换结果(合并数据的结果表)"（必选）
2. 导入"税务管理平台进项明细表"（必选，需包含"发票用途"列）
3. 选择输出目录后，点击"新增发票用途"
4. 在结果表末尾追加"发票用途"列

【匹配规则】
- 用结果表的"发票号"与税务管理平台进项明细表的"发票号码"进行匹配
- 同一发票号在明细中存在多条记录时：
  · 若所有记录的"发票用途"相同，仅展示一个用途名称
  · 若存在不同用途，则逐条展示"发票用途（对应税额）"

【输出说明】
- 只有匹配到发票用途的记录才会写入输出文件
- 未匹配到的记录不会出现在输出结果中

四、税额对比（功能三）
1. 导入"财务汇总_含发票用途文件"（必选）
2. 导入"税局进项明细表"（必选）
3. 导入"税务管理平台进项明细表"（必选）
4. 选择输出目录后，点击"开始对比"

【对比逻辑】
- 用税局进项明细表的"数电发票号码"与税务管理平台进项明细表的"发票号码"进行匹配
- 将明细中相同发票号的所有"税额"加总，与税局进项明细表的"票面税额*"进行对比
- 税额不一致的记录，会在输出结果中标记为黄色高亮

【输出结果】
- 基于"财务汇总_含发票用途文件"生成，税额不一致的行标黄

五、一键核账（功能四）★ 一键串联
1. 导入"税局进项明细表"（必选）
2. 导入"税务管理平台进项明细表"（必选）
3. 选择输出目录后，点击"一键转换"
4. 自动依次执行：合并数据 → 发票用途 → 税额对比
5. 直接输出最终的税额对比结果文件

【处理流程】
- 阶段一：合并数据 — 将两个表的字段映射后合并
- 阶段二：发票用途 — 用明细表匹配发票用途，追加到结果中
- 阶段三：税额对比 — 对比税额差异，差异行标黄
- 全程在内存中完成，仅输出最终结果文件

【适用场景】
- 适合日常快速核账，只需两个 Excel 文件即可完成全部处理
- 如需分步查看中间结果，请使用功能一、二、三单独操作

六、注意事项
1. 所有导入的 Excel 文件需为 .xlsx 或 .xls 格式
2. 请确保文件未被其他程序（如 Excel/WPS）占用，否则读取会失败
3. 税务管理平台进项明细表中必须包含"发票号码"列，发票用途/财务转换功能还需包含"发票用途"和"税额"列
4. 税局进项明细表中需包含"数电发票号码"和"票面税额*"列
5. 处理过程中请勿关闭程序窗口
6. 输出文件保存在指定目录下，文件名包含生成时间戳，不会覆盖已有文件
"""

        text.insert('1.0', guide_content)
        text.config(state='disabled')

    def _setup_about_page(self):
        """关于页面"""
        page = tk.Frame(self.content_area, padx=20, pady=15)
        self.pages['about'] = page

        tk.Label(page, text='关于', font=('Microsoft YaHei', 14, 'bold')).pack(pady=(0, 20))

        about_frame = tk.Frame(page, padx=10, pady=10)
        about_frame.pack(fill='both', expand=False)

        info = [
            ('软件名称', '进项数据采集票种校验工具'),
            ('版本号', 'v3.0.0'),
            ('上线日期', '2026年7月15日'),
            ('开发环境', 'Python 3.x + Tkinter'),
            ('主要功能', '进项数据合并、发票用途匹配、税额对比、一键核账'),
            ('适用场景', '进项数据采集与票种校验'),
        ]

        for i, (label, value) in enumerate(info):
            tk.Label(about_frame, text=f'{label}：', font=('Microsoft YaHei', 10, 'bold'), anchor='w').grid(row=i, column=0, sticky='w', padx=5, pady=8)
            tk.Label(about_frame, text=value, font=('Microsoft YaHei', 10), anchor='w').grid(row=i, column=1, sticky='w', padx=5, pady=8)

        # 版权信息
        copy_label = tk.Label(page, text='© 2026 进项数据采集票种校验工具  保留所有权利', font=('Microsoft YaHei', 9), fg='gray')
        copy_label.pack(side='bottom', pady=10)

    def _browse_cmp_usage(self):
        path = filedialog.askopenfilename(title='选择财务汇总_含发票用途文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.cmp_usage_path_var.set(path)

    def _browse_cmp_hui(self):
        path = filedialog.askopenfilename(title='选择财务汇总文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.cmp_hui_path_var.set(path)

    def _browse_cmp_ming(self):
        path = filedialog.askopenfilename(title='选择明细文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.cmp_ming_path_var.set(path)

    def _browse_cmp_output(self):
        path = filedialog.askdirectory(title='选择输出目录')
        if path:
            self.cmp_output_dir_var.set(path)

    def _on_cmp_path_change(self):
        if self.cmp_usage_path_var.get() and self.cmp_hui_path_var.get() and self.cmp_ming_path_var.get():
            self.cmp_run_btn.config(state='normal', bg='#E67E22')
        else:
            self.cmp_run_btn.config(state='disabled', bg='#E67E22')

    def _cmp_log(self, msg):
        self.cmp_log_text.config(state='normal')
        self.cmp_log_text.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.cmp_log_text.see('end')
        self.cmp_log_text.config(state='disabled')
        self.window.update_idletasks()

    def _run_compare(self):
        self.cmp_run_btn.config(state='disabled')
        self.cmp_progress['value'] = 0
        self.cmp_log_text.config(state='normal')
        self.cmp_log_text.delete('1.0', 'end')
        self.cmp_log_text.config(state='disabled')

        try:
            usage_path = self.cmp_usage_path_var.get()
            hui_path = self.cmp_hui_path_var.get()
            ming_path = self.cmp_ming_path_var.get()
            output_dir = self.cmp_output_dir_var.get()

            self._cmp_log(f'[1/4] 正在读取含发票用途文件: {os.path.basename(usage_path)}')
            df_usage = pd.read_excel(usage_path)
            self._cmp_log(f'      共 {len(df_usage)} 条记录')
            self.cmp_progress['value'] = 15

            self._cmp_log(f'[2/4] 正在读取财务汇总: {os.path.basename(hui_path)}')
            df_hui = pd.read_excel(hui_path)
            self._cmp_log(f'      共 {len(df_hui)} 条记录')
            self.cmp_progress['value'] = 35

            self._cmp_log(f'[3/4] 正在读取明细文件: {os.path.basename(ming_path)}')
            df_ming = pd.read_excel(ming_path)
            self._cmp_log(f'      共 {len(df_ming)} 条记录')

            # 明细按发票号码汇总税额
            ming_sum = df_ming.groupby('发票号码')['税额'].sum()
            self._cmp_log(f'      明细含 {len(ming_sum)} 个唯一发票号')
            self.cmp_progress['value'] = 55

            # 对比税额：汇总.数电发票号码 vs 明细.发票号码
            diff_invoices = set()
            match_count = 0
            for _, row in df_hui.iterrows():
                inv = row.get('数电发票号码', '')
                if pd.isna(inv):
                    continue
                inv = str(inv)
                hui_tax = self._float_or_zero(row.get('票面税额*', 0))
                if inv in ming_sum.index:
                    match_count += 1
                    ming_tax = round(ming_sum[inv], 2)
                    if round(hui_tax, 2) != ming_tax:
                        diff_invoices.add(inv)

            self._cmp_log(f'      汇总与明细匹配: {match_count} 条')
            self._cmp_log(f'      税额不一致: {len(diff_invoices)} 条')
            self.cmp_progress['value'] = 75

            self._cmp_log(f'[4/4] 正在生成对比结果文件...')
            output_path = self._write_compare_output(df_usage, diff_invoices, output_dir)
            self.cmp_progress['value'] = 100

            self._cmp_log(f'【完成】已生成文件: {output_path}')
            self._cmp_log(f'      标黄行数: {len(diff_invoices)} / 共 {len(df_usage)} 条')

            messagebox.showinfo('对比完成',
                                f'文件已生成:\n{output_path}\n\n'
                                f'总记录: {len(df_usage)} 条\n'
                                f'汇总与明细匹配: {match_count} 条\n'
                                f'税额不一致（已标黄）: {len(diff_invoices)} 条')

        except Exception as e:
            self._cmp_log(f'【异常】{e}')
            messagebox.showerror('错误', f'处理过程中发生错误:\n{e}')
        finally:
            self._on_cmp_path_change()

    def _write_compare_output(self, df_usage, diff_invoices, output_dir):
        """基于含发票用途文件，将差异行标黄输出"""
        wb = Workbook()
        ws = wb.active
        ws.title = '税额对比结果'

        yellow_fill = PatternFill('solid', fgColor='FFFF00')
        header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        data_font = Font(name='Microsoft YaHei', size=10)
        data_align = Alignment(horizontal='center', vertical='center')

        headers = list(df_usage.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = header_font
            c.fill = header_fill
            c.alignment = header_align
            c.border = thin_border

        for i, (_, row) in enumerate(df_usage.iterrows()):
            rn = i + 2
            inv = self._str_or_empty(row.get('发票号', ''))
            is_diff = inv in diff_invoices
            for j, col_name in enumerate(headers):
                val = row[col_name]
                if pd.isna(val):
                    val = ''
                elif hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                c = ws.cell(row=rn, column=j + 1, value=val)
                c.font = data_font
                c.alignment = data_align
                c.border = thin_border
                if is_diff:
                    c.fill = yellow_fill

        for ci in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(ci)].width = 18
        last_col = get_column_letter(len(headers))
        ws.auto_filter.ref = f'A1:{last_col}{len(df_usage) + 1}'
        ws.freeze_panes = 'A2'

        output_path = os.path.join(output_dir, f'税额对比结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(output_path)
        return output_path

    # ==================== 财务转换（一键串联） ====================

    def _merge_in_memory(self, df_hui, df_ming):
        """内存中执行合并数据，返回 DataFrame"""
        hui_rows, hui_invoices = self._extract_hui(df_hui)
        all_rows = list(hui_rows)

        if df_ming is not None:
            ming_rows, _, _ = self._extract_ming(df_ming, hui_invoices)
            all_rows.extend(ming_rows)

        data = []
        for i, row in enumerate(all_rows):
            data.append([i + 1] + row + [round(row[4] + row[5], 2)])

        return pd.DataFrame(data, columns=OUTPUT_HEADERS)

    def _add_usage_in_memory(self, df_merge, df_ming):
        """内存中为合并结果追加发票用途列，返回 DataFrame"""
        usage_map = self._build_usage_map(df_ming)
        usages = []
        for _, row in df_merge.iterrows():
            inv = self._str_or_empty(row.get('发票号', ''))
            if inv and inv in usage_map:
                usages.append('_'.join(usage_map[inv]))
            else:
                usages.append('')
        df_merge['发票用途'] = usages
        return df_merge

    def _browse_cvt_hui(self):
        path = filedialog.askopenfilename(title='选择税局进项明细表', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.cvt_hui_path_var.set(path)

    def _browse_cvt_ming(self):
        path = filedialog.askopenfilename(title='选择税务管理平台进项明细表', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.cvt_ming_path_var.set(path)

    def _browse_cvt_output(self):
        path = filedialog.askdirectory(title='选择输出目录')
        if path:
            self.cvt_output_dir_var.set(path)

    def _on_cvt_path_change(self):
        if self.cvt_hui_path_var.get() and self.cvt_ming_path_var.get():
            self.cvt_run_btn.config(state='normal', bg='#8E44AD')
        else:
            self.cvt_run_btn.config(state='disabled', bg='#8E44AD')

    def _cvt_log(self, msg):
        self.cvt_log_text.config(state='normal')
        self.cvt_log_text.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.cvt_log_text.see('end')
        self.cvt_log_text.config(state='disabled')
        self.window.update_idletasks()

    def _run_convert(self):
        """一键执行：合并数据 → 发票用途 → 税额对比"""
        self.cvt_run_btn.config(state='disabled')
        self.cvt_progress['value'] = 0
        self.cvt_log_text.config(state='normal')
        self.cvt_log_text.delete('1.0', 'end')
        self.cvt_log_text.config(state='disabled')

        try:
            hui_path = self.cvt_hui_path_var.get()
            ming_path = self.cvt_ming_path_var.get()
            output_dir = self.cvt_output_dir_var.get()

            # 阶段一：合并数据
            self._cvt_log('━' * 40)
            self._cvt_log('【阶段 1/3】合并数据')
            self._cvt_log(f'  读取税局进项明细表: {os.path.basename(hui_path)}')
            df_hui = pd.read_excel(hui_path)
            self._cvt_log(f'  共 {len(df_hui)} 条记录')
            self._cvt_log(f'  读取税务管理平台进项明细表: {os.path.basename(ming_path)}')
            df_ming = pd.read_excel(ming_path)
            self._cvt_log(f'  共 {len(df_ming)} 条记录')

            df_merge = self._merge_in_memory(df_hui, df_ming)
            self._cvt_log(f'  合并完成，共 {len(df_merge)} 条记录')
            self.cvt_progress['value'] = 33

            # 阶段二：发票用途
            self._cvt_log('━' * 40)
            self._cvt_log('【阶段 2/3】匹配发票用途')
            df_with_usage = self._add_usage_in_memory(df_merge, df_ming)
            matched = (df_with_usage['发票用途'] != '').sum()
            self._cvt_log(f'  匹配到发票用途: {matched} 条 / 共 {len(df_with_usage)} 条')
            self.cvt_progress['value'] = 66

            # 阶段三：税额对比
            self._cvt_log('━' * 40)
            self._cvt_log('【阶段 3/3】税额对比')
            ming_sum = df_ming.groupby('发票号码')['税额'].sum()
            self._cvt_log(f'  明细含 {len(ming_sum)} 个唯一发票号')

            diff_invoices = set()
            match_count = 0
            for _, row in df_hui.iterrows():
                inv = str(row.get('数电发票号码', ''))
                if pd.isna(row.get('数电发票号码')):
                    continue
                hui_tax = self._float_or_zero(row.get('票面税额*', 0))
                if inv in ming_sum.index:
                    match_count += 1
                    ming_tax = round(ming_sum[inv], 2)
                    if round(hui_tax, 2) != ming_tax:
                        diff_invoices.add(inv)

            self._cvt_log(f'  汇总与明细匹配: {match_count} 条')
            self._cvt_log(f'  税额不一致: {len(diff_invoices)} 条')
            self.cvt_progress['value'] = 90

            # 生成最终结果文件
            self._cvt_log(f'  正在生成最终对比结果文件...')
            output_path = self._write_compare_output(df_with_usage, diff_invoices, output_dir)
            self.cvt_progress['value'] = 100

            self._cvt_log('━' * 40)
            self._cvt_log(f'【完成】已生成文件: {output_path}')
            self._cvt_log(f'  总记录: {len(df_with_usage)} 条')
            self._cvt_log(f'  匹配到发票用途: {matched} 条')
            self._cvt_log(f'  税额不一致（已标黄）: {len(diff_invoices)} 条')

            messagebox.showinfo('一键转换完成',
                                f'文件已生成:\n{output_path}\n\n'
                                f'总记录: {len(df_with_usage)} 条\n'
                                f'匹配到发票用途: {matched} 条\n'
                                f'税额不一致（已标黄）: {len(diff_invoices)} 条')

        except Exception as e:
            self._cvt_log(f'【异常】{e}')
            messagebox.showerror('错误', f'处理过程中发生错误:\n{e}')
        finally:
            self._on_cvt_path_change()

    # ==================== 发票用途 ====================

    def _browse_usage_hui(self):
        path = filedialog.askopenfilename(title='选择财务汇总文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.usage_hui_path_var.set(path)

    def _browse_usage_ming(self):
        path = filedialog.askopenfilename(title='选择明细文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.usage_ming_path_var.set(path)

    def _browse_usage_output(self):
        path = filedialog.askdirectory(title='选择输出目录')
        if path:
            self.usage_output_dir_var.set(path)

    def _on_usage_path_change(self):
        if self.usage_hui_path_var.get() and self.usage_ming_path_var.get():
            self.usage_run_btn.config(state='normal', bg='#27AE60')
        else:
            self.usage_run_btn.config(state='disabled', bg='#27AE60')

    def _usage_log(self, msg):
        self.usage_log_text.config(state='normal')
        self.usage_log_text.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.usage_log_text.see('end')
        self.usage_log_text.config(state='disabled')
        self.window.update_idletasks()

    def _get_invoice_no(self, row):
        """获取汇总行的发票号"""
        invoice_no = row.get('发票号', '')
        if pd.isna(invoice_no):
            invoice_no = row.get('数电发票号码', '')
            if pd.isna(invoice_no):
                invoice_no = row.get('发票号码', '')
        return self._str_or_empty(invoice_no)

    def _build_usage_map(self, df_ming):
        """构建明细发票用途映射: {发票号: ['展示文本1', '展示文本2', ...]}"""
        if '发票用途' not in df_ming.columns or '税额' not in df_ming.columns:
            raise ValueError('明细文件缺少必要字段: 发票用途, 税额')

        # 先收集原始数据
        raw_map = {}  # {发票号: [{'usage': xxx, 'tax': xxx}, ...]}
        for _, row in df_ming.iterrows():
            inv = self._str_or_empty(row.get('发票号码', ''))
            if not inv:
                continue
            usage = self._str_or_empty(row.get('发票用途', ''))
            tax = self._float_or_zero(row.get('税额', 0))
            if inv not in raw_map:
                raw_map[inv] = []
            raw_map[inv].append({'usage': usage, 'tax': tax})

        # 根据发票用途是否一致，决定展示格式
        usage_map = {}
        for inv, items in raw_map.items():
            # 提取所有不同的发票用途
            usage_set = set(it['usage'] for it in items)
            if len(usage_set) == 1:
                # 所有发票用途一样，只展示发票用途
                usage_map[inv] = [usage_set.pop()]
            else:
                # 发票用途不同，展示 发票用途（税额）格式
                usage_map[inv] = [f'{it["usage"]}（{it["tax"]}）' for it in items]

        return usage_map

    def _write_usage_output(self, df_hui, usage_map, output_dir):
        """基于原汇总DataFrame，新增发票用途列，写入新Excel"""
        wb = Workbook()
        ws = wb.active
        ws.title = '财务汇总（含发票用途）'

        header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        data_font = Font(name='Microsoft YaHei', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')

        # 原汇总的所有列 + 发票用途
        original_cols = list(df_hui.columns)
        all_headers = original_cols + ['发票用途']

        for col_idx, header in enumerate(all_headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        matched_count = 0
        for i, (_, row) in enumerate(df_hui.iterrows()):
            row_num = i + 2
            for j, col_name in enumerate(original_cols):
                val = row[col_name]
                if pd.isna(val):
                    val = ''
                elif hasattr(val, 'strftime'):
                    val = val.strftime('%Y-%m-%d %H:%M:%S')
                cell = ws.cell(row=row_num, column=j + 1, value=val)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            inv = self._get_invoice_no(row)
            usage_cell = ws.cell(row=row_num, column=len(all_headers))
            if inv and inv in usage_map:
                usage_cell.value = '_'.join(usage_map[inv])
                matched_count += 1
            else:
                usage_cell.value = ''
            usage_cell.font = data_font
            usage_cell.alignment = data_alignment
            usage_cell.border = thin_border

        # 设置列宽
        for i in range(1, len(all_headers) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 18

        last_col_idx = len(all_headers)
        last_col_letter = get_column_letter(last_col_idx)
        ws.column_dimensions[last_col_letter].width = 50  # 发票用途列宽一些

        total_rows = len(df_hui) + 1
        ws.auto_filter.ref = f'A1:{last_col_letter}{total_rows}'
        ws.freeze_panes = 'A2'

        output_path = os.path.join(output_dir, f'财务汇总_含发票用途_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(output_path)
        return output_path, matched_count

    def _run_usage(self):
        self.usage_run_btn.config(state='disabled')
        self.usage_progress['value'] = 0
        self.usage_log_text.config(state='normal')
        self.usage_log_text.delete('1.0', 'end')
        self.usage_log_text.config(state='disabled')

        try:
            hui_path = self.usage_hui_path_var.get()
            ming_path = self.usage_ming_path_var.get()
            output_dir = self.usage_output_dir_var.get()

            self._usage_log(f'[1/4] 正在读取财务汇总: {os.path.basename(hui_path)}')
            df_hui = pd.read_excel(hui_path)
            self._usage_log(f'      共 {len(df_hui)} 条记录')
            self.usage_progress['value'] = 20

            self._usage_log(f'[2/4] 正在读取明细文件: {os.path.basename(ming_path)}')
            df_ming = pd.read_excel(ming_path)
            self._usage_log(f'      共 {len(df_ming)} 条记录')
            self.usage_progress['value'] = 45

            self._usage_log(f'[3/4] 正在构建发票用途映射...')
            usage_map = self._build_usage_map(df_ming)
            self._usage_log(f'      构建 {len(usage_map)} 个发票号的用途映射')
            self.usage_progress['value'] = 70

            self._usage_log(f'[4/4] 正在生成输出文件...')
            output_path, matched_count = self._write_usage_output(df_hui, usage_map, output_dir)

            self.usage_progress['value'] = 100
            self._usage_log(f'【完成】已生成文件: {output_path}')
            self._usage_log(f'      匹配到发票用途: {matched_count} 条 / 共 {len(df_hui)} 条')

            messagebox.showinfo('处理完成',
                                f'文件已生成:\n{output_path}\n\n'
                                f'总记录: {len(df_hui)} 条\n'
                                f'匹配到发票用途: {matched_count} 条\n'
                                f'未匹配: {len(df_hui) - matched_count} 条')

        except Exception as e:
            self._usage_log(f'【异常】{e}')
            messagebox.showerror('错误', f'处理过程中发生错误:\n{e}')
        finally:
            self._on_usage_path_change()

    def _browse_hui(self):
        path = filedialog.askopenfilename(title='选择财务汇总文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.hui_path_var.set(path)

    def _browse_ming(self):
        path = filedialog.askopenfilename(title='选择明细文件', filetypes=[('Excel文件', '*.xlsx *.xls'), ('所有文件', '*.*')])
        if path:
            self.ming_path_var.set(path)

    def _browse_output(self):
        path = filedialog.askdirectory(title='选择输出目录')
        if path:
            self.output_dir_var.set(path)

    def _on_path_change(self):
        if self.hui_path_var.get():
            self.run_btn.config(state='normal', bg='#2196F3')
        else:
            self.run_btn.config(state='disabled', bg='#2196F3')

    def _log(self, msg):
        self.log_text.config(state='normal')
        self.log_text.insert('end', f'[{datetime.now().strftime("%H:%M:%S")}] {msg}\n')
        self.log_text.see('end')
        self.log_text.config(state='disabled')
        self.window.update_idletasks()

    @staticmethod
    def _fmt_date(val):
        if pd.isna(val):
            return ''
        if hasattr(val, 'strftime'):
            return val.strftime('%Y-%m-%d %H:%M:%S')
        return str(val)

    @staticmethod
    def _str_or_empty(val):
        return str(val) if not pd.isna(val) else ''

    @staticmethod
    def _float_or_zero(val):
        return float(val) if not pd.isna(val) else 0.0

    def _extract_hui(self, df):
        missing = [c for c in COLUMN_MAP if c not in df.columns]
        if missing:
            raise ValueError(f'财务汇总文件缺少必要字段: {missing}')

        rows = []
        invoice_numbers = set()
        for _, row in df.iterrows():
            invoice_no = row.get('数电发票号码', '')
            if pd.isna(invoice_no):
                invoice_no = row.get('发票号码', '')
                if pd.isna(invoice_no):
                    invoice_no = row.get('发票代码', '')
            invoice_no = self._str_or_empty(invoice_no)
            if invoice_no:
                invoice_numbers.add(invoice_no)

            rows.append([
                invoice_no,
                self._str_or_empty(row.get('销售方纳税人名称', '')),
                self._fmt_date(row.get('开票日期*', '')),
                self._str_or_empty(row.get('票种*', '')),
                self._float_or_zero(row.get('金额*', 0)),
                self._float_or_zero(row.get('票面税额*', 0)),
            ])
        return rows, invoice_numbers

    def _extract_ming(self, df, exclude_invoices):
        missing = [c for c in DETAIL_COLUMN_MAP if c not in df.columns]
        if missing:
            raise ValueError(f'明细文件缺少必要字段: {missing}')

        agg = {}
        merged_count = 0
        skipped_exist = 0

        for _, row in df.iterrows():
            invoice_no = self._str_or_empty(row.get('发票号码', ''))
            if not invoice_no:
                continue

            if invoice_no in exclude_invoices:
                skipped_exist += 1
                continue

            amount = self._float_or_zero(row.get('不含税金额', 0))
            tax = self._float_or_zero(row.get('税额', 0))

            if invoice_no in agg:
                merged_count += 1
                agg[invoice_no]['amount'] += amount
                agg[invoice_no]['tax'] += tax
            else:
                agg[invoice_no] = {
                    'ticket_kind': self._str_or_empty(row.get('实物发票种类', '')),
                    'date': self._fmt_date(row.get('开票日期', '')),
                    'ticket_type': self._str_or_empty(row.get('发票类型', '')),
                    'amount': amount,
                    'tax': tax,
                }

        rows = []
        for invoice_no, data in agg.items():
            rows.append([
                invoice_no,
                data['ticket_kind'],
                data['date'],
                data['ticket_type'],
                data['amount'],
                data['tax'],
            ])

        return rows, merged_count, skipped_exist

    def _write_output(self, all_rows, output_dir):
        wb = Workbook()
        ws = wb.active
        ws.title = '财务汇总转换结果'

        header_font = Font(name='Microsoft YaHei', bold=True, size=11, color='FFFFFF')
        header_fill = PatternFill('solid', fgColor='4472C4')
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin'),
        )
        data_font = Font(name='Microsoft YaHei', size=10)
        data_alignment = Alignment(horizontal='center', vertical='center')

        for col_idx, header in enumerate(OUTPUT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

        money_fmt = '#,##0.00'

        for i, row_data in enumerate(all_rows):
            row_num = i + 2
            ws.cell(row=row_num, column=1, value=i + 1).font = data_font
            ws.cell(row=row_num, column=1).alignment = data_alignment
            ws.cell(row=row_num, column=1).border = thin_border

            for j, val in enumerate(row_data):
                cell = ws.cell(row=row_num, column=j + 2, value=val)
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border

            for col in (6, 7, 8):
                cell = ws.cell(row=row_num, column=col)
                cell.number_format = money_fmt

            sum_cell = ws.cell(row=row_num, column=8)
            sum_cell.value = round(row_data[4] + row_data[5], 2)
            sum_cell.font = data_font
            sum_cell.alignment = data_alignment
            sum_cell.border = thin_border

        col_widths = [8, 26, 30, 22, 36, 16, 16, 26]
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

        total_rows = len(all_rows) + 1
        ws.auto_filter.ref = f'A1:H{total_rows}'
        ws.freeze_panes = 'A2'

        output_path = os.path.join(output_dir, f'财务汇总转换结果_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(output_path)
        return output_path, len(all_rows)

    def _run_conversion(self):
        self.run_btn.config(state='disabled')
        self.progress['value'] = 0
        self.log_text.config(state='normal')
        self.log_text.delete('1.0', 'end')
        self.log_text.config(state='disabled')

        try:
            hui_path = self.hui_path_var.get()
            ming_path = self.ming_path_var.get().strip()
            output_dir = self.output_dir_var.get()

            self._log(f'[1/4] 正在读取财务汇总: {os.path.basename(hui_path)}')
            df_hui = pd.read_excel(hui_path)
            self._log(f'      共 {len(df_hui)} 条记录')
            self.progress['value'] = 15

            self._log(f'[2/4] 正在提取汇总数据...')
            hui_rows, hui_invoices = self._extract_hui(df_hui)
            self._log(f'      提取 {len(hui_rows)} 条有效记录, {len(hui_invoices)} 个唯一发票号')
            all_rows = list(hui_rows)
            self.progress['value'] = 40

            if ming_path and os.path.isfile(ming_path):
                self._log(f'[3/4] 正在读取明细文件: {os.path.basename(ming_path)}')
                df_ming = pd.read_excel(ming_path)
                self._log(f'      共 {len(df_ming)} 条记录')

                ming_rows, merged_count, skipped_exist = self._extract_ming(df_ming, hui_invoices)
                all_rows.extend(ming_rows)
                self._log(f'      明细内部合并 {merged_count} 条为 {len(ming_rows)} 条, 与汇总重复跳过 {skipped_exist} 条')
                self._log(f'      新增明细记录 {len(ming_rows)} 条')
            else:
                self._log(f'[3/4] 未指定明细文件，跳过合并步骤')
                ming_path = None

            self.progress['value'] = 65

            self._log(f'[4/4] 正在生成输出文件...')
            output_path, total = self._write_output(all_rows, output_dir)

            self.progress['value'] = 100
            self._log(f'【完成】已生成文件: {output_path}')

            summary = f'财务汇总: {len(hui_rows)} 条'
            if ming_path:
                summary += f'\n明细新增: {len(all_rows) - len(hui_rows)} 条'
            summary += f'\n合计: {total} 条'

            self._log(f'      {summary.replace(chr(10), " / ")}')
            messagebox.showinfo('转换完成', f'文件已生成:\n{output_path}\n\n{summary}')

        except Exception as e:
            self._log(f'【异常】{e}')
            messagebox.showerror('错误', f'转换过程中发生错误:\n{e}')
        finally:
            self.run_btn.config(state='normal')

    def run(self):
        self.window.mainloop()


if __name__ == '__main__':
    FinanceConverter().run()
